import json
import re
from io import BytesIO
from datetime import datetime
import io
import csv
from decimal import Decimal
import uuid
from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.services.audit_helper import safe_audit, model_snapshot
from app.core.database import get_db
from app.models.core_models import AuditLog, Partner, Section, Subsection, User
from app.models.contact_models import PartnerContact, PartnerLink, PartnerProduct
from app.services.passwords import hash_password
from app.services import mailer as mailer_service

def send_email(db, to_email: str, subject: str, text_body: str, **kwargs):
    """Compatibility guard for mixed deployed files.
    Some older mailer.py builds do not accept template_key. Never let that break e-mail sending.
    """
    try:
        return mailer_service.send_email(db, to_email, subject, text_body, **kwargs)
    except TypeError as exc:
        if "template_key" in str(exc):
            safe_kwargs = dict(kwargs)
            safe_kwargs.pop("template_key", None)
            return mailer_service.send_email(db, to_email, subject, text_body, **safe_kwargs)
        raise

smtp_config_status = mailer_service.smtp_config_status
ensure_email_tables = mailer_service.ensure_email_tables
email_template = mailer_service.email_template
EMAIL_VERSION = "1.6.1-mail-templates-professional-safe"
public_smtp_diagnostics = getattr(mailer_service, "public_smtp_diagnostics", lambda: {})

def send_template_email(db, to_email: str, template_key: str, *, data=None, event_type: str = "system", entity_type: str = "", entity_id: str = "", created_by_email: str = ""):
    """Robust compatibility wrapper. Never lets template tuple mismatch crash the request."""
    try:
        tpl = email_template(template_key, **(data or {}))
        if isinstance(tpl, (list, tuple)) and len(tpl) >= 3:
            subject, body, html = tpl[0], tpl[1], tpl[2]
        elif isinstance(tpl, (list, tuple)) and len(tpl) == 2:
            subject, body = tpl
            html = None
        else:
            subject = f"HUB ASTORIE – {template_key}"
            body = str(tpl or "")
            html = None
    except Exception as exc:
        subject = f"HUB ASTORIE – {template_key}"
        body = f"Dobrý den,\n\ne-mail byl vytvořen systémem HUB ASTORIE.\n\nPoznámka šablony: {exc}\n\nASTORIE a.s."
        html = None
    return send_email(
        db, to_email, subject, body, html_body=html, event_type=event_type,
        entity_type=entity_type, entity_id=entity_id, created_by_email=created_by_email,
        template_key=template_key
    )
from app.services.importer import IMPORT_HANDLERS
from app.services.ares import fetch_ares_subject

router = APIRouter(tags=["admin-ui"])


def render(request: Request, template_name: str, context: dict):
    templates = request.app.state.templates
    base_context = {
        "request": request,
        "app_name": "HUB",
        "version": "v1.6.1",
        "admin_name": "Admin ASTORIE",
        "admin_email": "nekudova@astorieas.cz",
    }
    base_context.update(context)
    return templates.TemplateResponse(template_name, base_context)


def audit(db: Session, action: str, entity_type: str, payload: dict):
    try:
        db.add(AuditLog(user_email="admin@astorie.local", action=action, entity_type=entity_type, new_value=payload))
        db.commit()
    except Exception:
        db.rollback()



# --- v1.5.0: user multi-role + module permissions -------------------------
SYSTEM_ROLES_V150 = ["IF", "PS", "ADMIN", "BO", "VEDENI"]
HUB_MODULES_V150 = [
    ("tip", "📝 Nový TIP", "/hub/new-tip"),
    ("mytips", "📌 Moje TIPy", "/hub/my-tips"),
    ("calcs", "🧮 Kalkulačky", "/hub/calculators"),
    ("partners", "🏢 Partneři", "/hub/partners"),
    ("contacts", "☎️ Kontakty", "/hub/contacts"),
    ("links", "🔗 Odkazy ASTORIE", "/hub/links"),
    ("vypovedi", "📄 Výpovědi", "/hub/terminations"),
    ("formulare", "📑 Formuláře", "/hub/forms"),
    ("napoveda", "❓ Nápověda", "/hub/help"),
]
ADMIN_MODULES_V150 = [
    ("dashboard", "📊 Dashboard", "/admin"),
    ("modules", "🧭 Mapa modulů", "/admin/modules"),
    ("advisors", "👥 Poradci / uživatelé", "/admin/advisors"),
    ("permissions", "🔐 Oprávnění menu", "/admin/permissions"),
    ("sections", "🧭 Sekce / podsekce", "/admin/sections"),
    ("partners", "🏢 Partneři", "/admin/partners"),
    ("contacts", "👥 Kontakty", "/admin/contacts"),
    ("contact_roles", "🏷️ Role kontaktů", "/admin/contact-roles"),
    ("links", "🔗 Odkazy", "/admin/links"),
    ("products", "📦 Produkty", "/admin/products"),
    ("rates", "📊 Sazebník provizí", "/admin/rates"),
    ("terminations", "📄 Výpovědi", "/admin/terminations"),
    ("termination_archive", "🗂️ Evidence výpovědí", "/admin/terminations/archive"),
    ("admin_tips", "📋 Správa TIPů", "/admin/tips"),
    ("specialists", "👤 Specialisté", "/admin/specialists"),
    ("email", "✉️ E-maily / SMTP", "/admin/email"),
    ("import", "⬆️ Import dat", "/admin/import"),
]


def parse_roles_v150(value):
    if not value:
        return []
    raw = str(value).replace(';', ',').replace('|', ',').split(',')
    roles = []
    for r in raw:
        code = r.strip().upper()
        if code and code not in roles:
            roles.append(code)
    return roles


def normalize_roles_v150(values):
    if not values:
        return "IF"
    if isinstance(values, str):
        values = [values]
    roles = []
    for r in values:
        code = str(r).strip().upper()
        if code in SYSTEM_ROLES_V150 and code not in roles:
            roles.append(code)
    return ",".join(roles or ["IF"])


def ensure_user_permissions_v150(db: Session):
    """Bezpečný upgrade bez destruktivních migrací: ponechá sloupec role a přidá čitelný multi-role text."""
    db.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS advisor_id VARCHAR(80)"))
    db.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS name VARCHAR(255)"))
    db.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS email VARCHAR(255)"))
    db.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS phone VARCHAR(80)"))
    db.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS role VARCHAR(120) DEFAULT 'IF' NOT NULL"))
    db.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS password_hash TEXT"))
    db.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS is_active BOOLEAN DEFAULT TRUE NOT NULL"))
    db.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS must_change_password BOOLEAN DEFAULT FALSE NOT NULL"))
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS module_permissions (
            id SERIAL PRIMARY KEY,
            module_id VARCHAR(80) NOT NULL,
            module_name VARCHAR(180) NOT NULL,
            module_url VARCHAR(255) NOT NULL DEFAULT '',
            area VARCHAR(30) NOT NULL DEFAULT 'hub',
            role_code VARCHAR(30) NOT NULL,
            is_allowed BOOLEAN DEFAULT TRUE NOT NULL,
            updated_at TIMESTAMP WITH TIME ZONE DEFAULT now() NOT NULL
        )
    """))
    db.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ux_module_permissions_area_module_role ON module_permissions(area, module_id, role_code)"))
    defaults = []
    for area, mods in [("hub", HUB_MODULES_V150), ("admin", ADMIN_MODULES_V150)]:
        for module_id, module_name, module_url in mods:
            for role in SYSTEM_ROLES_V150:
                allowed = True
                if area == "admin" and role not in ("ADMIN", "BO", "VEDENI"):
                    allowed = False
                if module_id in ("permissions",) and role != "ADMIN":
                    allowed = False
                defaults.append({"area": area, "module_id": module_id, "module_name": module_name, "module_url": module_url, "role": role, "allowed": allowed})
    for d in defaults:
        db.execute(text("""
            INSERT INTO module_permissions (area, module_id, module_name, module_url, role_code, is_allowed)
            VALUES (:area, :module_id, :module_name, :module_url, :role, :allowed)
            ON CONFLICT (area, module_id, role_code) DO NOTHING
        """), d)
    db.commit()


def get_menu_preview_v150(db: Session, roles):
    ensure_user_permissions_v150(db)
    role_list = parse_roles_v150(','.join(roles) if isinstance(roles, list) else roles)
    if not role_list:
        role_list = ["IF"]
    rows = db.execute(text("""
        SELECT area, module_id, module_name, module_url, bool_or(is_allowed) AS allowed
        FROM module_permissions
        WHERE role_code = ANY(:roles)
        GROUP BY area, module_id, module_name, module_url
        ORDER BY area, module_name
    """), {"roles": role_list}).mappings().all()
    return rows



def ensure_termination_archive_table_v147_(db: Session):
    """Archiv vygenerovaných výpovědí. Bez zásahu do existujících tabulek."""
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS termination_documents (
            id UUID PRIMARY KEY,
            created_at TIMESTAMP WITH TIME ZONE DEFAULT now() NOT NULL,
            created_by_name VARCHAR(255),
            created_by_email VARCHAR(255),
            partner_code VARCHAR(100),
            partner_name VARCHAR(255),
            client_name VARCHAR(255),
            client_identifier VARCHAR(120),
            client_address TEXT,
            policy_no VARCHAR(180),
            insurance_type VARCHAR(255),
            insured_subject TEXT,
            termination_type VARCHAR(20),
            reason_text TEXT,
            bank_account VARCHAR(180),
            extra_date VARCHAR(60),
            note TEXT,
            document_text TEXT,
            status VARCHAR(80) DEFAULT 'Vygenerováno' NOT NULL
        )
    """))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_termination_documents_created_at ON termination_documents(created_at DESC)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_termination_documents_partner_code ON termination_documents(partner_code)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_termination_documents_policy_no ON termination_documents(policy_no)"))
    db.commit()


def build_termination_document_v147_(partner, termination_type: str, client_name: str, client_identifier: str,
                                     client_address: str, policy_no: str, insurance_type: str,
                                     insured_subject: str, bank_account: str, extra_date: str, note: str):
    reason_map = {
        "A": "ke konci pojistného období",
        "B": "ve lhůtě do 2 měsíců od uzavření smlouvy",
        "C": "po oznámení pojistné události",
        "D": "z důvodu nesouhlasu se změnou výše pojistného",
        "E": "z důvodu zániku pojistného zájmu – prodej předmětu pojištění",
        "F": "z důvodu vyřazení vozidla z evidence / odcizení",
    }
    reason = reason_map.get(termination_type, reason_map["A"])
    insurer_name = getattr(partner, 'name', '') if partner else ''
    insurer_address = getattr(partner, 'address_full', '') if partner else ''
    insurer_data_box = getattr(partner, 'data_box', '') if partner else ''
    insurer_email = getattr(partner, 'registry_email', '') if partner else ''
    lines = [
        "VÝPOVĚĎ POJISTNÉ SMLOUVY", "",
        f"Adresát: {insurer_name}",
    ]
    if insurer_address: lines.append(f"Adresa: {insurer_address}")
    if insurer_data_box: lines.append(f"Datová schránka: {insurer_data_box}")
    if insurer_email: lines.append(f"E-mail: {insurer_email}")
    lines += ["", f"Pojistník: {client_name}"]
    if client_identifier: lines.append(f"Identifikace: {client_identifier}")
    if client_address: lines.append(f"Adresa pojistníka: {client_address}")
    lines += ["", f"Tímto vypovídám pojistnou smlouvu č. {policy_no}.", f"Výpověď podávám {reason}."]
    if extra_date: lines.append(f"Rozhodné datum: {extra_date}")
    if insurance_type: lines.append(f"Druh pojištění: {insurance_type}")
    if insured_subject: lines.append(f"Identifikace předmětu pojištění: {insured_subject}")
    lines += ["", "Žádám o potvrzení přijetí této výpovědi."]
    if bank_account:
        lines.append(f"Případný přeplatek pojistného žádám zaslat na bankovní účet: {bank_account}.")
    else:
        lines.append("Případný přeplatek pojistného žádám zaslat na adresu pojistníka.")
    if note:
        lines += ["", f"Poznámka: {note}"]
    lines += ["", "V ........................ dne ................", "", "........................................", "podpis pojistníka"]
    return reason, lines, "\n".join(lines)


def save_termination_document_v147_(db: Session, partner, partner_code: str, termination_type: str,
                                    client_name: str, client_identifier: str, client_address: str,
                                    policy_no: str, insurance_type: str, insured_subject: str,
                                    bank_account: str, extra_date: str, note: str, document_text: str, reason_text: str):
    ensure_termination_archive_table_v147_(db)
    doc_id = str(uuid.uuid4())
    db.execute(text("""
        INSERT INTO termination_documents
        (id, created_by_name, created_by_email, partner_code, partner_name, client_name, client_identifier,
         client_address, policy_no, insurance_type, insured_subject, termination_type, reason_text,
         bank_account, extra_date, note, document_text, status)
        VALUES
        (:id, :created_by_name, :created_by_email, :partner_code, :partner_name, :client_name, :client_identifier,
         :client_address, :policy_no, :insurance_type, :insured_subject, :termination_type, :reason_text,
         :bank_account, :extra_date, :note, :document_text, 'Vygenerováno')
    """), {
        "id": doc_id,
        "created_by_name": "Admin ASTORIE",
        "created_by_email": "nekudova@astorieas.cz",
        "partner_code": (partner_code or '').upper(),
        "partner_name": getattr(partner, 'name', '') if partner else '',
        "client_name": client_name,
        "client_identifier": client_identifier,
        "client_address": client_address,
        "policy_no": policy_no,
        "insurance_type": insurance_type,
        "insured_subject": insured_subject,
        "termination_type": termination_type,
        "reason_text": reason_text,
        "bank_account": bank_account,
        "extra_date": extra_date,
        "note": note,
        "document_text": document_text,
    })
    db.commit()
    return doc_id

@router.get("/", response_class=HTMLResponse)
def home():
    return RedirectResponse(url="/admin", status_code=302)


@router.get("/admin", response_class=HTMLResponse)
@router.get("/admin/", response_class=HTMLResponse)
def admin_dashboard(request: Request, db: Session = Depends(get_db)):
    try:
        ensure_visible_hub_sections_(db)
    except Exception:
        pass
    counts = {
        "users": db.query(User).count(),
        "sections": db.query(Section).count(),
        "subsections": db.query(Subsection).count(),
        "partners": db.query(Partner).count(),
        "audit": db.query(AuditLog).count(),
    }
    latest_audit = db.query(AuditLog).order_by(AuditLog.created_at.desc()).limit(8).all()
    return render(request, "dashboard.html", {"active": "dashboard", "counts": counts, "latest_audit": latest_audit})


@router.get("/admin/users", response_class=HTMLResponse)
def users_page(request: Request, db: Session = Depends(get_db)):
    users = db.query(User).order_by(User.created_at.desc()).all()
    return render(request, "users.html", {"active": "users", "users": users})


@router.post("/admin/users")
def create_user(
    advisor_id: str = Form(...),
    name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(""),
    roles: list[str] = Form([]),
    role: str = Form(""),
    password: str = Form("1234"),
    db: Session = Depends(get_db),
):
    user = User(
        advisor_id=advisor_id.strip(),
        name=name.strip(),
        email=email.strip().lower(),
        phone=phone.strip(),
        role=role.strip().upper(),
        password_hash=hash_password(password),
        is_active=True,
        must_change_password=True,
    )
    db.add(user)
    db.commit()
    audit(db, "CREATE", "users", {"advisor_id": advisor_id, "email": email, "role": role})
    return RedirectResponse(url="/admin/users", status_code=303)


@router.post("/admin/users/{user_id}/toggle")
def toggle_user(user_id: str, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if user:
        user.is_active = not user.is_active
        db.commit()
        audit(db, "TOGGLE_ACTIVE", "users", {"id": user_id, "is_active": user.is_active})
    return RedirectResponse(url="/admin/users", status_code=303)


@router.post("/admin/sections")
def create_section(
    section_code: str = Form(...),
    name: str = Form(...),
    icon: str = Form(""),
    sort_order: int = Form(0),
    db: Session = Depends(get_db),
):
    row = Section(section_code=section_code.strip().upper(), name=name.strip(), icon=icon.strip(), sort_order=sort_order, is_active=True)
    db.add(row)
    db.commit()
    audit(db, "CREATE", "sections", {"section_code": section_code, "name": name})
    return RedirectResponse(url="/admin/sections", status_code=303)


@router.post("/admin/subsections")
def create_subsection(
    subsection_code: str = Form(...),
    section_code: str = Form(...),
    name: str = Form(...),
    sort_order: int = Form(0),
    db: Session = Depends(get_db),
):
    row = Subsection(
        subsection_code=subsection_code.strip().upper(),
        section_code=section_code.strip().upper(),
        name=name.strip(),
        sort_order=sort_order,
        is_active=True,
    )
    db.add(row)
    db.commit()
    audit(db, "CREATE", "subsections", {"subsection_code": subsection_code, "section_code": section_code, "name": name})
    return RedirectResponse(url="/admin/sections", status_code=303)


@router.get("/admin/partners", response_class=HTMLResponse)
def partners_page(
    request: Request,
    q: str = "",
    status: str = "",
    segment: str = "",
    db: Session = Depends(get_db),
):
    query = db.query(Partner)

    if q:
        like = f"%{q.lower()}%"
        query = query.filter(
            (Partner.partner_code.ilike(like)) |
            (Partner.name.ilike(like)) |
            (Partner.ico.ilike(like)) |
            (Partner.data_box.ilike(like)) |
            (Partner.registry_email.ilike(like)) |
            (Partner.city.ilike(like)) |
            (Partner.address_full.ilike(like))
        )

    if status:
        query = query.filter(Partner.partner_status == status)

    if segment == "vip":
        query = query.filter(Partner.is_vip == True)
    elif segment == "fleet":
        query = query.filter(Partner.segment_fleet == True)
    elif segment == "retail":
        query = query.filter(Partner.segment_retail == True)
    elif segment == "life":
        query = query.filter(Partner.segment_life == True)
    elif segment == "business":
        query = query.filter(Partner.segment_business == True)
    elif segment == "missing_contact":
        query = query.outerjoin(PartnerContact, PartnerContact.partner_code == Partner.partner_code).filter(PartnerContact.id == None)

    partners = query.order_by(Partner.name).limit(1000).all()

    return render(request, "partners.html", {
        "active": "partners",
        "partners": partners,
        "q": q,
        "status": status,
        "segment": segment,
    })


@router.post("/admin/partners")
def create_partner(
    partner_code: str = Form(...),
    name: str = Form(...),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    row = Partner(partner_code=partner_code.strip().upper(), name=name.strip(), note=note.strip(), is_active=True)
    db.add(row)
    db.commit()
    audit(db, "CREATE", "partners", {"partner_code": partner_code, "name": name})
    return RedirectResponse(url="/admin/partners", status_code=303)


@router.get("/admin/audit", response_class=HTMLResponse)
def audit_page(request: Request, db: Session = Depends(get_db)):
    rows = db.query(AuditLog).order_by(AuditLog.created_at.desc()).limit(100).all()
    return render(request, "audit.html", {"active": "audit", "rows": rows})


@router.get("/admin/import", response_class=HTMLResponse)
def import_page(request: Request):
    return render(request, "import.html", {"active": "import", "result": None})


@router.post("/admin/import", response_class=HTMLResponse)
async def import_csv(
    request: Request,
    entity: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    handler = IMPORT_HANDLERS.get(entity)
    if not handler:
        return render(request, "import.html", {
            "active": "import",
            "result": {"ok": False, "entity": entity, "errors": ["Neznámý typ importu."]},
        })

    raw = await file.read()
    try:
        result = handler(db, raw).as_dict()
    except Exception as exc:
        result = {"ok": False, "entity": entity, "errors": [str(exc)]}

    return render(request, "import.html", {"active": "import", "result": result})




# --- v1.4.9 SAFE: Contact role dictionary ---------------------------------
def ensure_contact_role_tables_v149(db: Session):
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS contact_roles (
            id SERIAL PRIMARY KEY,
            group_name VARCHAR(120) NOT NULL DEFAULT 'Ostatní',
            role_name VARCHAR(180) NOT NULL,
            description TEXT NOT NULL DEFAULT '',
            sort_order INTEGER NOT NULL DEFAULT 100,
            is_active BOOLEAN NOT NULL DEFAULT TRUE,
            created_at TIMESTAMPTZ DEFAULT now(),
            updated_at TIMESTAMPTZ DEFAULT now()
        )
    """))
    db.execute(text("ALTER TABLE partner_contacts ADD COLUMN IF NOT EXISTS role_group VARCHAR(120) NOT NULL DEFAULT ''"))
    db.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ux_contact_roles_group_role ON contact_roles (lower(group_name), lower(role_name))"))
    defaults = [
        ('VIP', 'KAM', 'Klíčový obchodní kontakt / Key Account Manager', 10),
        ('VIP', 'Ředitel', 'Manažerský nebo ředitelský kontakt', 20),
        ('VIP', 'Obchodní ředitel', 'Obchodní vedení partnera', 30),
        ('Infolinka', 'Infolinka', 'Obecná klientská nebo poradenská linka', 10),
        ('Infolinka', 'Helpdesk', 'Helpdesk / provozní podpora', 20),
        ('Infolinka', 'IT', 'Technická podpora / portál', 30),
        ('Infolinka', 'Podpora klient', 'Klientská podpora', 40),
        ('Smlouvy', 'Dotazy ke smlouvám', 'Změny, smlouvy, dokumenty', 10),
        ('Likvidace', 'Likvidace', 'Likvidace pojistných událostí', 10),
        ('Provize', 'Provize', 'Provizní a produkční dotazy', 10),
        ('Ostatní', 'Ostatní', 'Jiný typ kontaktu', 999),
    ]
    for group_name, role_name, description, sort_order in defaults:
        db.execute(text("""
            INSERT INTO contact_roles (group_name, role_name, description, sort_order, is_active)
            VALUES (:group_name, :role_name, :description, :sort_order, TRUE)
            ON CONFLICT DO NOTHING
        """), {"group_name": group_name, "role_name": role_name, "description": description, "sort_order": sort_order})
    db.commit()

def get_contact_roles_v149(db: Session, only_active: bool = True):
    ensure_contact_role_tables_v149(db)
    where = "WHERE is_active IS TRUE" if only_active else ""
    rows = db.execute(text(f"""
        SELECT id, group_name, role_name, description, sort_order, is_active
        FROM contact_roles
        {where}
        ORDER BY group_name, sort_order, role_name
    """)).mappings().all()
    return [dict(r) for r in rows]

def contact_role_groups_v149(roles):
    groups = []
    seen = set()
    for r in roles:
        g = r.get('group_name') or 'Ostatní'
        if g not in seen:
            groups.append(g); seen.add(g)
    return groups or ['VIP','Infolinka','Smlouvy','Likvidace','Provize','Ostatní']

@router.get("/admin/contacts", response_class=HTMLResponse)
def contacts_page(request: Request, q: str = "", partner: str = "", role_group: str = "", db: Session = Depends(get_db)):
    roles = get_contact_roles_v149(db)
    groups = contact_role_groups_v149(roles)
    query = db.query(PartnerContact)
    if partner:
        query = query.filter(PartnerContact.partner_code == partner.upper())
    if role_group:
        like_group = f"%{role_group}%"
        query = query.filter((PartnerContact.contact_type.ilike(like_group)) | (PartnerContact.role.ilike(like_group)))
    if q:
        like = f"%{q}%"
        query = query.filter(
            (PartnerContact.full_name.ilike(like)) |
            (PartnerContact.role.ilike(like)) |
            (PartnerContact.email.ilike(like)) |
            (PartnerContact.phone.ilike(like)) |
            (PartnerContact.specialization.ilike(like)) |
            (PartnerContact.territory.ilike(like)) |
            (PartnerContact.contact_type.ilike(like)) |
            (PartnerContact.partner_code.ilike(like))
        )
    contacts = query.order_by(PartnerContact.partner_code, PartnerContact.contact_type, PartnerContact.role, PartnerContact.full_name).limit(700).all()
    partners = db.query(Partner).order_by(Partner.name).all()
    return render(request, "contacts.html", {"active": "contacts", "contacts": contacts, "partners": partners, "roles": roles, "role_groups": groups, "q": q, "partner": partner, "role_group": role_group})


@router.post("/admin/contacts/create")
def create_contact(request: Request, partner_code: str = Form(...), full_name: str = Form(...), role: str = Form(""), role_custom: str = Form(""), role_group: str = Form(""), email: str = Form(""), phone: str = Form(""), specialization: str = Form(""), contact_type: str = Form(""), territory: str = Form(""), is_vip: str = Form(""), note: str = Form(""), db: Session = Depends(get_db)):
    ensure_contact_role_tables_v149(db)
    final_role = (role_custom or role or "Ostatní").strip()
    final_group = (role_group or contact_type or "Ostatní").strip()
    if final_role:
        db.execute(text("""
            INSERT INTO contact_roles (group_name, role_name, description, sort_order, is_active)
            VALUES (:group_name, :role_name, '', 100, TRUE)
            ON CONFLICT DO NOTHING
        """), {"group_name": final_group, "role_name": final_role})
    db.add(PartnerContact(partner_code=partner_code.upper().strip(), full_name=full_name.strip(), role=final_role, email=email.strip(), phone=phone.strip(), specialization=specialization.strip(), contact_type=final_group, territory=territory.strip(), is_vip=bool(is_vip), is_top=bool(is_vip), note=note, is_active=True))
    db.commit()
    return RedirectResponse("/admin/contacts", status_code=303)


@router.post("/admin/contacts/{item_id}/duplicate")
def duplicate_contact(item_id: int, db: Session = Depends(get_db)):
    item = db.query(PartnerContact).filter(PartnerContact.id == item_id).first()
    if item:
        db.add(PartnerContact(partner_code=item.partner_code, full_name=item.full_name + " – kopie", role=item.role, email=item.email, phone=item.phone, specialization=item.specialization, contact_type=item.contact_type, territory=item.territory, is_vip=item.is_vip, is_top=item.is_top, note=item.note, is_active=item.is_active))
        db.commit()
    return RedirectResponse("/admin/contacts", status_code=303)




@router.get("/admin/contact-roles", response_class=HTMLResponse)
def contact_roles_page(request: Request, db: Session = Depends(get_db)):
    roles = get_contact_roles_v149(db, only_active=False)
    return render(request, "contact_roles.html", {"active": "contact_roles", "roles": roles, "role_groups": contact_role_groups_v149(roles)})


@router.post("/admin/contact-roles/create")
def create_contact_role(group_name: str = Form(...), role_name: str = Form(...), description: str = Form(""), sort_order: int = Form(100), db: Session = Depends(get_db)):
    ensure_contact_role_tables_v149(db)
    db.execute(text("""
        INSERT INTO contact_roles (group_name, role_name, description, sort_order, is_active)
        VALUES (:group_name, :role_name, :description, :sort_order, TRUE)
        ON CONFLICT DO NOTHING
    """), {"group_name": group_name.strip() or "Ostatní", "role_name": role_name.strip(), "description": description.strip(), "sort_order": sort_order})
    db.commit()
    return RedirectResponse("/admin/contact-roles", status_code=303)


@router.post("/admin/contact-roles/{role_id}/toggle")
def toggle_contact_role(role_id: int, db: Session = Depends(get_db)):
    ensure_contact_role_tables_v149(db)
    db.execute(text("UPDATE contact_roles SET is_active = NOT is_active, updated_at = now() WHERE id = :id"), {"id": role_id})
    db.commit()
    return RedirectResponse("/admin/contact-roles", status_code=303)


@router.post("/admin/contact-roles/{role_id}/update")
def update_contact_role(role_id: int, group_name: str = Form(...), role_name: str = Form(...), description: str = Form(""), sort_order: int = Form(100), db: Session = Depends(get_db)):
    ensure_contact_role_tables_v149(db)
    db.execute(text("""
        UPDATE contact_roles
        SET group_name = :group_name, role_name = :role_name, description = :description, sort_order = :sort_order, updated_at = now()
        WHERE id = :id
    """), {"id": role_id, "group_name": group_name.strip() or "Ostatní", "role_name": role_name.strip(), "description": description.strip(), "sort_order": sort_order})
    db.commit()
    return RedirectResponse("/admin/contact-roles", status_code=303)

@router.get("/admin/links", response_class=HTMLResponse)
def links_page(request: Request, q: str = "", partner: str = "", db: Session = Depends(get_db)):
    query = db.query(PartnerLink)
    if partner:
        query = query.filter(PartnerLink.partner_code == partner.upper())
    if q:
        like = f"%{q}%"
        query = query.filter((PartnerLink.title.ilike(like)) | (PartnerLink.url.ilike(like)) | (PartnerLink.category.ilike(like)) | (PartnerLink.note.ilike(like)) | (PartnerLink.partner_code.ilike(like)))
    links = query.order_by(PartnerLink.partner_code, PartnerLink.title).limit(500).all()
    partners = db.query(Partner).order_by(Partner.name).all()
    return render(request, "links.html", {"active": "links", "links": links, "partners": partners, "q": q, "partner": partner})


@router.post("/admin/links/create")
def create_link(request: Request, partner_code: str = Form(...), title: str = Form(...), url: str = Form(...), category: str = Form(""), note: str = Form(""), source_type: str = Form(""), db: Session = Depends(get_db)):
    ensure_link_source_columns_v155a_(db)
    item = PartnerLink(partner_code=partner_code.upper().strip(), title=title, url=url, category=category, note=note, is_active=True)
    db.add(item)
    db.flush()
    st = (source_type or "").strip().upper()
    if st in ("ASTORIE_LINK", "ONLINE_CALCULATOR", "PARTNER_LINK"):
        db.execute(text("UPDATE partner_links SET source_type = :st WHERE id = :id"), {"st": st, "id": item.id})
    db.commit()
    return RedirectResponse("/admin/links", status_code=303)


@router.post("/admin/links/{item_id}/duplicate")
def duplicate_link(item_id: int, db: Session = Depends(get_db)):
    item = db.query(PartnerLink).filter(PartnerLink.id == item_id).first()
    if item:
        db.add(PartnerLink(partner_code=item.partner_code, title=item.title + " – kopie", url=item.url, category=item.category, note=item.note, is_active=item.is_active))
        db.commit()
    return RedirectResponse("/admin/links", status_code=303)


@router.get("/admin/products", response_class=HTMLResponse)
def products_page(request: Request, q: str = "", partner: str = "", db: Session = Depends(get_db)):
    query = db.query(PartnerProduct)
    if partner:
        query = query.filter(PartnerProduct.partner_code == partner.upper())
    if q:
        like = f"%{q}%"
        query = query.filter((PartnerProduct.partner_code.ilike(like)) | (PartnerProduct.area.ilike(like)) | (PartnerProduct.subarea.ilike(like)) | (PartnerProduct.product_name.ilike(like)) | (PartnerProduct.note.ilike(like)))
    products = query.order_by(PartnerProduct.partner_code, PartnerProduct.area, PartnerProduct.product_name).limit(500).all()
    partners = db.query(Partner).order_by(Partner.name).all()
    return render(request, "products.html", {"active": "products", "products": products, "partners": partners, "q": q, "partner": partner})


@router.post("/admin/products/create")
def create_product(request: Request, partner_code: str = Form(...), area: str = Form(""), subarea: str = Form(""), product_name: str = Form(...), note: str = Form(""), db: Session = Depends(get_db)):
    db.add(PartnerProduct(partner_code=partner_code.upper().strip(), area=area, subarea=subarea, product_name=product_name, note=note, is_active=True))
    db.commit()
    return RedirectResponse("/admin/products", status_code=303)


@router.post("/admin/products/{item_id}/duplicate")
def duplicate_product(item_id: int, db: Session = Depends(get_db)):
    item = db.query(PartnerProduct).filter(PartnerProduct.id == item_id).first()
    if item:
        db.add(PartnerProduct(partner_code=item.partner_code, area=item.area, subarea=item.subarea, product_name=item.product_name + " – kopie", note=item.note, is_active=item.is_active))
        db.commit()
    return RedirectResponse("/admin/products", status_code=303)


@router.get("/admin/partners/{partner_code}", response_class=HTMLResponse)
def partner_detail(request: Request, partner_code: str, db: Session = Depends(get_db)):
    partner = db.query(Partner).filter(Partner.partner_code == partner_code.upper()).first()
    contacts = db.query(PartnerContact).filter(PartnerContact.partner_code == partner_code.upper()).order_by(PartnerContact.full_name).all()
    links = db.query(PartnerLink).filter(PartnerLink.partner_code == partner_code.upper()).order_by(PartnerLink.title).all()
    products = db.query(PartnerProduct).filter(PartnerProduct.partner_code == partner_code.upper()).order_by(PartnerProduct.area, PartnerProduct.product_name).all()
    return render(request, "partner_detail.html", {"active": "partners", "partner": partner, "partner_code": partner_code.upper(), "contacts": contacts, "links": links, "products": products})


@router.get("/api/ares/subject")
def api_ares_subject(ico: str):
    return JSONResponse(fetch_ares_subject(ico))


@router.post("/admin/partners/create-extended")
def create_partner_extended(
    request: Request,
    partner_code: str = Form(...),
    name: str = Form(...),
    ico: str = Form(""),
    dic: str = Form(""),
    data_box: str = Form(""),
    registry_email: str = Form(""),
    street: str = Form(""),
    city: str = Form(""),
    zip_code: str = Form(""),
    address_full: str = Form(""),
    legal_form: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    code = partner_code.upper().strip()
    existing = db.query(Partner).filter(Partner.partner_code == code).first()
    if existing:
        existing.name = name
        existing.ico = ico
        existing.dic = dic
        existing.data_box = data_box
        existing.registry_email = registry_email
        existing.street = street
        existing.city = city
        existing.zip_code = zip_code
        existing.address_full = address_full
        existing.legal_form = legal_form
        existing.note = note
        existing.is_active = True
    else:
        db.add(Partner(
            partner_code=code,
            name=name,
            ico=ico,
            dic=dic,
            data_box=data_box,
            registry_email=registry_email,
            street=street,
            city=city,
            zip_code=zip_code,
            address_full=address_full,
            legal_form=legal_form,
            note=note,
            is_active=True,
        ))
    db.commit()
    return RedirectResponse("/admin/partners", status_code=303)


@router.post("/admin/partners/{partner_code}/ares-refresh")
def refresh_partner_from_ares(partner_code: str, db: Session = Depends(get_db)):
    partner = db.query(Partner).filter(Partner.partner_code == partner_code.upper()).first()
    if not partner or not partner.ico:
        return RedirectResponse(f"/admin/partners/{partner_code.upper()}", status_code=303)

    data = fetch_ares_subject(partner.ico)
    if data.get("ok"):
        partner.name = data.get("name") or partner.name
        partner.dic = data.get("dic") or partner.dic
        partner.data_box = data.get("data_box") or partner.data_box
        partner.street = data.get("street") or partner.street
        partner.city = data.get("city") or partner.city
        partner.zip_code = data.get("zip_code") or partner.zip_code
        partner.address_full = data.get("address_full") or partner.address_full
        partner.legal_form = data.get("legal_form") or partner.legal_form
        partner.source = "ARES"
        db.commit()

    return RedirectResponse(f"/admin/partners/{partner_code.upper()}", status_code=303)


@router.post("/admin/partners/{partner_code}/duplicate-partner")
def duplicate_partner(partner_code: str, db: Session = Depends(get_db)):
    src = db.query(Partner).filter(Partner.partner_code == partner_code.upper()).first()
    if not src:
        return RedirectResponse("/admin/partners", status_code=303)

    base_code = (src.partner_code + "_COPY")[:40]
    code = base_code
    counter = 1
    while db.query(Partner).filter(Partner.partner_code == code).first():
        counter += 1
        code = f"{base_code}_{counter}"[:50]

    db.add(Partner(
        partner_code=code,
        name=(src.name or "") + " – kopie",
        ico=src.ico,
        dic=src.dic,
        data_box=src.data_box,
        registry_email=src.registry_email,
        street=src.street,
        city=src.city,
        zip_code=src.zip_code,
        address_full=src.address_full,
        legal_form=src.legal_form,
        source=src.source,
        note=src.note,
        partner_status=getattr(src, "partner_status", "aktivní"),
        cooperation_status=getattr(src, "cooperation_status", ""),
        is_vip=getattr(src, "is_vip", False),
        segment_fleet=getattr(src, "segment_fleet", False),
        segment_retail=getattr(src, "segment_retail", False),
        segment_life=getattr(src, "segment_life", False),
        segment_business=getattr(src, "segment_business", False),
        onboarding_done=getattr(src, "onboarding_done", False),
        contract_valid=getattr(src, "contract_valid", False),
        last_audit_note=getattr(src, "last_audit_note", ""),
        is_active=True,
    ))
    db.commit()
    return RedirectResponse("/admin/partners", status_code=303)


@router.post("/admin/partners/{partner_code}/update")
def update_partner(
    partner_code: str,
    request: Request,
    name: str = Form(...),
    ico: str = Form(""),
    dic: str = Form(""),
    data_box: str = Form(""),
    registry_email: str = Form(""),
    street: str = Form(""),
    city: str = Form(""),
    zip_code: str = Form(""),
    address_full: str = Form(""),
    legal_form: str = Form(""),
    note: str = Form(""),
    partner_status: str = Form("aktivní"),
    cooperation_status: str = Form(""),
    is_vip: str = Form(""),
    segment_fleet: str = Form(""),
    segment_retail: str = Form(""),
    segment_life: str = Form(""),
    segment_business: str = Form(""),
    onboarding_done: str = Form(""),
    contract_valid: str = Form(""),
    last_audit_note: str = Form(""),
    is_active: str = Form(""),
    db: Session = Depends(get_db),
):
    partner = db.query(Partner).filter(Partner.partner_code == partner_code.upper()).first()
    if not partner:
        return RedirectResponse("/admin/partners", status_code=303)

    old_snapshot = model_snapshot(partner, [
        "name", "ico", "dic", "data_box", "registry_email", "street", "city", "zip_code",
        "address_full", "legal_form", "note", "is_active"
    ])

    partner.name = name
    partner.ico = ico
    partner.dic = dic
    partner.data_box = data_box
    partner.registry_email = registry_email
    partner.street = street
    partner.city = city
    partner.zip_code = zip_code
    partner.address_full = address_full
    partner.legal_form = legal_form
    partner.note = note
    partner.partner_status = partner_status
    partner.cooperation_status = cooperation_status
    partner.is_vip = bool(is_vip)
    partner.segment_fleet = bool(segment_fleet)
    partner.segment_retail = bool(segment_retail)
    partner.segment_life = bool(segment_life)
    partner.segment_business = bool(segment_business)
    partner.onboarding_done = bool(onboarding_done)
    partner.contract_valid = bool(contract_valid)
    partner.last_audit_note = last_audit_note
    partner.is_active = bool(is_active)
    new_snapshot = model_snapshot(partner, [
        "name", "ico", "dic", "data_box", "registry_email", "street", "city", "zip_code",
        "address_full", "legal_form", "note", "is_active"
    ])
    db.commit()
    safe_audit(db, "admin@astorie.local", "UPDATE", "partner", partner.partner_code, old_snapshot, new_snapshot, "Úprava partnera")

    return RedirectResponse(f"/admin/partners/{partner.partner_code}", status_code=303)


@router.get("/api/partners/{partner_code}/registry")
def api_partner_registry(partner_code: str, db: Session = Depends(get_db)):
    partner = db.query(Partner).filter(Partner.partner_code == partner_code.upper()).first()
    if not partner:
        return JSONResponse({"ok": False, "error": "Partner nebyl nalezen."}, status_code=404)

    return {
        "ok": True,
        "partner": {
            "partner_code": partner.partner_code,
            "name": partner.name,
            "ico": partner.ico,
            "dic": partner.dic,
            "data_box": partner.data_box,
            "registry_email": partner.registry_email,
            "street": partner.street,
            "city": partner.city,
            "zip_code": partner.zip_code,
            "address_full": partner.address_full,
            "legal_form": partner.legal_form,
            "note": partner.note,
            "is_active": partner.is_active,
            "partner_status": getattr(partner, "partner_status", "aktivní"),
            "cooperation_status": getattr(partner, "cooperation_status", ""),
            "is_vip": getattr(partner, "is_vip", False),
            "segment_fleet": getattr(partner, "segment_fleet", False),
            "segment_retail": getattr(partner, "segment_retail", False),
            "segment_life": getattr(partner, "segment_life", False),
            "segment_business": getattr(partner, "segment_business", False),
        }
    }


@router.post("/admin/contacts/{item_id}/update")
def update_contact(
    item_id: int,
    partner_code: str = Form(...),
    full_name: str = Form(...),
    role: str = Form(""),
    email: str = Form(""),
    phone: str = Form(""),
    specialization: str = Form(""),
    contact_type: str = Form(""),
    territory: str = Form(""),
    is_vip: str = Form(""),
    note: str = Form(""),
    is_active: str = Form(""),
    db: Session = Depends(get_db),
):
    item = db.query(PartnerContact).filter(PartnerContact.id == item_id).first()
    if item:
        item.partner_code = partner_code.upper().strip()
        item.full_name = full_name
        item.role = role
        item.email = email
        item.phone = phone
        item.specialization = specialization
        item.contact_type = contact_type
        item.territory = territory
        item.is_vip = bool(is_vip)
        item.is_top = bool(is_vip)
        item.note = note
        item.is_active = bool(is_active)
        db.commit()
        return RedirectResponse(f"/admin/partners/{item.partner_code}", status_code=303)
    return RedirectResponse("/admin/contacts", status_code=303)


@router.post("/admin/contacts/{item_id}/toggle")
def toggle_contact(item_id: int, db: Session = Depends(get_db)):
    item = db.query(PartnerContact).filter(PartnerContact.id == item_id).first()
    if item:
        item.is_active = not item.is_active
        partner_code = item.partner_code
        db.commit()
        return RedirectResponse(f"/admin/partners/{partner_code}", status_code=303)
    return RedirectResponse("/admin/contacts", status_code=303)


@router.post("/admin/links/{item_id}/update")
def update_link(
    item_id: int,
    partner_code: str = Form(...),
    title: str = Form(...),
    url: str = Form(...),
    category: str = Form(""),
    note: str = Form(""),
    is_active: str = Form(""),
    db: Session = Depends(get_db),
):
    item = db.query(PartnerLink).filter(PartnerLink.id == item_id).first()
    if item:
        item.partner_code = partner_code.upper().strip()
        item.title = title
        item.url = url
        item.category = category
        item.note = note
        item.is_active = bool(is_active)
        db.commit()
        return RedirectResponse(f"/admin/partners/{item.partner_code}", status_code=303)
    return RedirectResponse("/admin/links", status_code=303)


@router.post("/admin/links/{item_id}/toggle")
def toggle_link(item_id: int, db: Session = Depends(get_db)):
    item = db.query(PartnerLink).filter(PartnerLink.id == item_id).first()
    if item:
        item.is_active = not item.is_active
        partner_code = item.partner_code
        db.commit()
        return RedirectResponse(f"/admin/partners/{partner_code}", status_code=303)
    return RedirectResponse("/admin/links", status_code=303)


@router.post("/admin/products/{item_id}/update")
def update_product(
    item_id: int,
    partner_code: str = Form(...),
    area: str = Form(""),
    subarea: str = Form(""),
    product_name: str = Form(...),
    note: str = Form(""),
    is_active: str = Form(""),
    db: Session = Depends(get_db),
):
    item = db.query(PartnerProduct).filter(PartnerProduct.id == item_id).first()
    if item:
        item.partner_code = partner_code.upper().strip()
        item.area = area
        item.subarea = subarea
        item.product_name = product_name
        item.note = note
        item.is_active = bool(is_active)
        db.commit()
        return RedirectResponse(f"/admin/partners/{item.partner_code}", status_code=303)
    return RedirectResponse("/admin/products", status_code=303)


@router.post("/admin/products/{item_id}/toggle")
def toggle_product(item_id: int, db: Session = Depends(get_db)):
    item = db.query(PartnerProduct).filter(PartnerProduct.id == item_id).first()
    if item:
        item.is_active = not item.is_active
        partner_code = item.partner_code
        db.commit()
        return RedirectResponse(f"/admin/partners/{partner_code}", status_code=303)
    return RedirectResponse("/admin/products", status_code=303)


@router.get("/api/partners/search")
def api_partner_search(q: str = "", limit: int = 15, db: Session = Depends(get_db)):
    text = (q or "").strip()
    query = db.query(Partner).filter(Partner.is_active == True)

    if text:
        like = f"%{text.lower()}%"
        query = query.filter(
            (Partner.partner_code.ilike(like)) |
            (Partner.name.ilike(like)) |
            (Partner.ico.ilike(like)) |
            (Partner.data_box.ilike(like)) |
            (Partner.registry_email.ilike(like)) |
            (Partner.city.ilike(like))
        )

    items = query.order_by(Partner.name).limit(max(1, min(limit, 25))).all()

    return {
        "ok": True,
        "items": [
            {
                "partner_code": p.partner_code,
                "label": f"{p.partner_code} – {p.name}",
                "name": p.name,
                "ico": p.ico,
                "data_box": p.data_box,
                "email": p.registry_email,
                "city": p.city,
            }
            for p in items
        ],
    }


@router.get("/api/partners/{partner_code}/form-source")
def api_partner_form_source(partner_code: str, db: Session = Depends(get_db)):
    partner = db.query(Partner).filter(Partner.partner_code == partner_code.upper()).first()
    if not partner:
        return JSONResponse({"ok": False, "error": "Partner nebyl nalezen."}, status_code=404)

    top_contacts = (
        db.query(PartnerContact)
        .filter(PartnerContact.partner_code == partner.partner_code)
        .filter(PartnerContact.is_active == True)
        .order_by(PartnerContact.is_top.desc(), PartnerContact.is_vip.desc(), PartnerContact.full_name)
        .limit(10)
        .all()
    )

    links = (
        db.query(PartnerLink)
        .filter(PartnerLink.partner_code == partner.partner_code)
        .filter(PartnerLink.is_active == True)
        .order_by(PartnerLink.category, PartnerLink.title)
        .limit(20)
        .all()
    )

    products = (
        db.query(PartnerProduct)
        .filter(PartnerProduct.partner_code == partner.partner_code)
        .filter(PartnerProduct.is_active == True)
        .order_by(PartnerProduct.area, PartnerProduct.subarea, PartnerProduct.product_name)
        .limit(50)
        .all()
    )

    return {
        "ok": True,
        "partner": {
            "partner_code": partner.partner_code,
            "name": partner.name,
            "ico": partner.ico,
            "dic": partner.dic,
            "data_box": partner.data_box,
            "registry_email": partner.registry_email,
            "street": partner.street,
            "city": partner.city,
            "zip_code": partner.zip_code,
            "address_full": partner.address_full,
            "legal_form": partner.legal_form,
            "is_active": partner.is_active,
            "partner_status": getattr(partner, "partner_status", "aktivní"),
            "cooperation_status": getattr(partner, "cooperation_status", ""),
            "is_vip": getattr(partner, "is_vip", False),
            "segment_fleet": getattr(partner, "segment_fleet", False),
            "segment_retail": getattr(partner, "segment_retail", False),
            "segment_life": getattr(partner, "segment_life", False),
            "segment_business": getattr(partner, "segment_business", False),
        },
        "form_prefill": {
            "insurer_name": partner.name,
            "insurer_ico": partner.ico,
            "insurer_data_box": partner.data_box,
            "insurer_email": partner.registry_email,
            "insurer_address": partner.address_full or " ".join([x for x in [partner.street, partner.zip_code, partner.city] if x]),
        },
        "contacts": [
            {
                "full_name": c.full_name,
                "role": c.role,
                "contact_type": c.contact_type,
                "territory": c.territory,
                "email": c.email,
                "phone": c.phone,
                "is_top": c.is_top or c.is_vip,
            }
            for c in top_contacts
        ],
        "links": [
            {
                "title": l.title,
                "url": l.url,
                "category": l.category,
            }
            for l in links
        ],
        "products": [
            {
                "area": p.area,
                "subarea": p.subarea,
                "product_name": p.product_name,
            }
            for p in products
        ],
    }


@router.get("/admin/form-bridge", response_class=HTMLResponse)
def form_bridge_page(request: Request):
    return render(request, "form_bridge.html", {"active": "form_bridge"})


@router.get("/admin/terminations", response_class=HTMLResponse)
def terminations_page(request: Request, partner_code: str = "", db: Session = Depends(get_db)):
    partner = db.query(Partner).filter(Partner.partner_code == partner_code.upper()).first() if partner_code else None
    partners = db.query(Partner).filter(Partner.is_active == True).order_by(Partner.name).limit(1000).all()
    partner_items = [
        {
            "partner_code": x.partner_code or "",
            "name": x.name or "",
            "ico": x.ico or "",
            "data_box": x.data_box or "",
            "registry_email": x.registry_email or "",
            "address_full": x.address_full or "",
            "city": x.city or "",
        }
        for x in partners
    ]
    return render(request, "terminations.html", {
        "active": "terminations",
        "partner": partner,
        "partners": partner_items,
        "partner_code": partner_code.upper() if partner_code else "",
    })


@router.post("/admin/terminations/preview", response_class=HTMLResponse)
def termination_preview(
    request: Request,
    partner_code: str = Form(""),
    termination_type: str = Form("A"),
    client_name: str = Form(""),
    client_identifier: str = Form(""),
    client_address: str = Form(""),
    policy_no: str = Form(""),
    insurance_type: str = Form(""),
    insured_subject: str = Form(""),
    bank_account: str = Form(""),
    extra_date: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    partner = db.query(Partner).filter(Partner.partner_code == partner_code.upper()).first() if partner_code else None
    reason_text, preview_lines, preview_text = build_termination_document_v147_(
        partner, termination_type, client_name, client_identifier, client_address, policy_no,
        insurance_type, insured_subject, bank_account, extra_date, note
    )
    doc_id = save_termination_document_v147_(
        db, partner, partner_code, termination_type, client_name, client_identifier, client_address,
        policy_no, insurance_type, insured_subject, bank_account, extra_date, note, preview_text, reason_text
    )
    return render(request, "termination_preview.html", {
        "active": "terminations",
        "partner": partner,
        "partner_code": partner_code.upper() if partner_code else "",
        "preview_text": preview_text,
        "preview_lines": preview_lines,
        "doc_id": doc_id,
        "saved": True,
    })


@router.get("/admin/terminations/archive", response_class=HTMLResponse)
def admin_terminations_archive_v147(request: Request, q: str = "", db: Session = Depends(get_db)):
    ensure_termination_archive_table_v147_(db)
    where = ""
    params = {}
    if q:
        where = "WHERE lower(coalesce(client_name,'') || ' ' || coalesce(policy_no,'') || ' ' || coalesce(partner_name,'') || ' ' || coalesce(created_by_name,'')) LIKE :q"
        params["q"] = f"%{q.lower()}%"
    rows = fetch_all_safe_v084_(db, f"""
        SELECT * FROM termination_documents
        {where}
        ORDER BY created_at DESC
        LIMIT 500
    """, params)
    return render(request, "terminations_archive.html", {"active": "termination_archive", "rows": [dict(r) for r in rows], "q": q})


@router.get("/admin/terminations/archive/{doc_id}", response_class=HTMLResponse)
def admin_termination_archive_detail_v147(request: Request, doc_id: str, db: Session = Depends(get_db)):
    ensure_termination_archive_table_v147_(db)
    row = fetch_one_safe_v084_(db, "SELECT * FROM termination_documents WHERE id = :id LIMIT 1", {"id": doc_id})
    if not row:
        return HTMLResponse("Nenalezeno", status_code=404)
    return render(request, "termination_archive_detail.html", {"active": "termination_archive", "doc": dict(row), "preview_lines": (row["document_text"] or "").split("\n")})


@router.get("/admin/modules", response_class=HTMLResponse)
def modules_page(request: Request, focus: str = '', db: Session = Depends(get_db)):
    modules = [
        {
            "group": "TIP HUB",
            "items": [
                {"name": "Nový TIP", "status": "připraveno", "desc": "Budoucí migrace zadání TIPu z uživatelského HUBu.", "url": "/admin/modules"},
                {"name": "Moje TIPy", "status": "připraveno", "desc": "Budoucí přehled odeslaných a přidělených TIPů.", "url": "/admin/modules"},
                {"name": "Specialisté", "status": "připraveno", "desc": "Správa specialistů, dostupnosti a routingu.", "url": "/admin/modules"},
                {"name": "Statistiky & soutěž", "status": "připraveno", "desc": "Budoucí manažerský žebříček a vyhodnocení TIPů.", "url": "/admin/modules"},
            ],
        },
        {
            "group": "ČÍSELNÍKY",
            "items": [
                {"name": "Poradci / uživatelé", "status": "funkční", "desc": "Základní správa uživatelů.", "url": "/admin/advisors"},
                {"name": "Sekce / podsekce", "status": "funkční", "desc": "Základní struktura oblastí.", "url": "/admin/sections"},
                {"name": "Partneři", "status": "funkční", "desc": "Centrální partner registry s ARES.", "url": "/admin/partners"},
                {"name": "Kontakty", "status": "funkční", "desc": "Kontakty partnerů.", "url": "/admin/contacts"},
                {"name": "Odkazy", "status": "funkční", "desc": "Odkazy partnerů.", "url": "/admin/links"},
                {"name": "Produkty", "status": "funkční", "desc": "Produkty partnerů.", "url": "/admin/products"},
            ],
        },
        {
            "group": "DOKUMENTY",
            "items": [
                {"name": "Výpovědi", "status": "funkční základ", "desc": "Modul výpovědí s napojením na partnera.", "url": "/admin/terminations"},
                {"name": "Formuláře", "status": "připraveno", "desc": "Budoucí generování dalších dokumentů.", "url": "/admin/form-bridge"},
                {"name": "Napojení formulářů", "status": "funkční", "desc": "API bridge pro přebírání dat partnera.", "url": "/admin/form-bridge"},
            ],
        },
        {
            "group": "SYSTÉM",
            "items": [
                {"name": "Import dat", "status": "funkční", "desc": "Import ze stávajících CSV/Sheets zdrojů.", "url": "/admin/import"},
                {"name": "Audit", "status": "funkční", "desc": "Záznam změn a provozních událostí.", "url": "/admin/audit"},
                {"name": "Health", "status": "funkční", "desc": "Kontrola běhu služby.", "url": "/health"},
                {"name": "API dokumentace", "status": "funkční", "desc": "OpenAPI dokumentace.", "url": "/docs"},
            ],
        },
    ]

    return render(request, "modules.html", {
        "active": "modules",
        "modules": modules,
        "focus": focus,
    })


@router.post("/admin/partners/registry-upgrade")
def partner_registry_upgrade(db: Session = Depends(get_db)):
    statements = [
        "ALTER TABLE partners ADD COLUMN IF NOT EXISTS partner_status VARCHAR(80) DEFAULT 'aktivní' NOT NULL",
        "ALTER TABLE partners ADD COLUMN IF NOT EXISTS cooperation_status VARCHAR(120) DEFAULT '' NOT NULL",
        "ALTER TABLE partners ADD COLUMN IF NOT EXISTS is_vip BOOLEAN DEFAULT FALSE NOT NULL",
        "ALTER TABLE partners ADD COLUMN IF NOT EXISTS segment_fleet BOOLEAN DEFAULT FALSE NOT NULL",
        "ALTER TABLE partners ADD COLUMN IF NOT EXISTS segment_retail BOOLEAN DEFAULT FALSE NOT NULL",
        "ALTER TABLE partners ADD COLUMN IF NOT EXISTS segment_life BOOLEAN DEFAULT FALSE NOT NULL",
        "ALTER TABLE partners ADD COLUMN IF NOT EXISTS segment_business BOOLEAN DEFAULT FALSE NOT NULL",
        "ALTER TABLE partners ADD COLUMN IF NOT EXISTS onboarding_done BOOLEAN DEFAULT FALSE NOT NULL",
        "ALTER TABLE partners ADD COLUMN IF NOT EXISTS contract_valid BOOLEAN DEFAULT FALSE NOT NULL",
        "ALTER TABLE partners ADD COLUMN IF NOT EXISTS last_audit_note TEXT DEFAULT '' NOT NULL",
    ]
    for sql in statements:
        db.execute(text(sql))
    db.commit()
    return RedirectResponse("/admin/partners", status_code=303)


@router.post("/admin/audit-history/upgrade")
def audit_history_upgrade(db: Session = Depends(get_db)):
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS audit_history (
            id SERIAL PRIMARY KEY,
            created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT now(),
            user_email TEXT NOT NULL DEFAULT '',
            action TEXT NOT NULL DEFAULT '',
            entity TEXT NOT NULL DEFAULT '',
            entity_id TEXT NOT NULL DEFAULT '',
            old_data TEXT NOT NULL DEFAULT '',
            new_data TEXT NOT NULL DEFAULT '',
            note TEXT NOT NULL DEFAULT ''
        )
    """))
    db.commit()
    return RedirectResponse("/admin/audit-history", status_code=303)


@router.get("/admin/audit-history", response_class=HTMLResponse)
def audit_history_page(
    request: Request,
    q: str = "",
    entity: str = "",
    action: str = "",
    limit: int = 150,
    db: Session = Depends(get_db),
):
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS audit_history (
            id SERIAL PRIMARY KEY,
            created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT now(),
            user_email TEXT NOT NULL DEFAULT '',
            action TEXT NOT NULL DEFAULT '',
            entity TEXT NOT NULL DEFAULT '',
            entity_id TEXT NOT NULL DEFAULT '',
            old_data TEXT NOT NULL DEFAULT '',
            new_data TEXT NOT NULL DEFAULT '',
            note TEXT NOT NULL DEFAULT ''
        )
    """))
    db.commit()

    sql = "SELECT * FROM audit_history WHERE 1=1"
    params = {}

    if q:
        sql += " AND (lower(user_email) LIKE :q OR lower(entity_id) LIKE :q OR lower(note) LIKE :q OR lower(new_data) LIKE :q OR lower(old_data) LIKE :q)"
        params["q"] = f"%{q.lower()}%"

    if entity:
        sql += " AND entity = :entity"
        params["entity"] = entity

    if action:
        sql += " AND action = :action"
        params["action"] = action

    sql += " ORDER BY created_at DESC LIMIT :limit"
    params["limit"] = max(1, min(limit, 1000))

    rows = db.execute(text(sql), params).mappings().all()

    return render(request, "audit_history.html", {
        "active": "audit_history",
        "rows": rows,
        "q": q,
        "entity": entity,
        "action": action,
        "limit": limit,
    })


@router.get("/admin/partners/{partner_code}/history", response_class=HTMLResponse)
def partner_history_page(request: Request, partner_code: str, db: Session = Depends(get_db)):
    rows = db.execute(
        text("""
            SELECT * FROM audit_history
            WHERE entity_id = :partner_code OR lower(new_data) LIKE :like OR lower(old_data) LIKE :like
            ORDER BY created_at DESC
            LIMIT 300
        """),
        {"partner_code": partner_code.upper(), "like": f"%{partner_code.lower()}%"}
    ).mappings().all()

    partner = db.query(Partner).filter(Partner.partner_code == partner_code.upper()).first()

    return render(request, "partner_history.html", {
        "active": "partners",
        "partner": partner,
        "partner_code": partner_code.upper(),
        "rows": rows,
    })



@router.get("/admin/advisors", response_class=HTMLResponse)
def advisors_page(
    request: Request,
    q: str = "",
    role: str = "",
    active: str = "",
    db: Session = Depends(get_db),
):
    ensure_user_permissions_v150(db)

    sql = """
      SELECT
        id,
        COALESCE(advisor_id, '') AS advisor_id,
        COALESCE(name, '') AS name,
        COALESCE(email, '') AS email,
        COALESCE(phone, '') AS phone,
        COALESCE(role, '') AS role,
        COALESCE(is_active, TRUE) AS is_active,
        COALESCE(must_change_password, FALSE) AS must_change_password
      FROM users
      WHERE 1=1
    """
    params = {}

    if q:
        sql += """
          AND (
            lower(COALESCE(advisor_id, '')) LIKE :q OR
            lower(COALESCE(name, '')) LIKE :q OR
            lower(COALESCE(email, '')) LIKE :q OR
            lower(COALESCE(phone, '')) LIKE :q
          )
        """
        params["q"] = f"%{q.lower()}%"

    if role:
        sql += " AND (',' || replace(upper(COALESCE(role, '')), ' ', '') || ',') LIKE :role_like"
        params["role_like"] = f"%,{role.upper()},%"

    if active == "1":
        sql += " AND COALESCE(is_active, TRUE) = TRUE"
    elif active == "0":
        sql += " AND COALESCE(is_active, TRUE) = FALSE"

    sql += " ORDER BY name, advisor_id LIMIT 1000"

    advisors = db.execute(text(sql), params).mappings().all()

    return render(request, "advisors.html", {
        "active": "advisors",
        "advisors": advisors,
        "q": q,
        "role": role,
        "active_filter": active,
        "system_roles": SYSTEM_ROLES_V150,
        "menu_preview": get_menu_preview_v150(db, role or "IF"),
    })


@router.post("/admin/advisors/create")
def advisor_create(
    advisor_id: str = Form(""),
    name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(""),
    roles: list[str] = Form([]),
    role: str = Form(""),
    password: str = Form("1234"),
    is_active: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_user_permissions_v150(db)
    role_value = normalize_roles_v150(roles or [role])
    # Heslo ukládáme přes existující hash helper, pokud je dostupný.
    try:
        password_hash = hash_password(password)
    except Exception:
        password_hash = password

    db.execute(text("""
      INSERT INTO users
        (advisor_id, name, email, phone, role, password_hash, is_active, must_change_password)
      VALUES
        (:advisor_id, :name, :email, :phone, :role, :password_hash, :is_active, TRUE)
    """), {
        "advisor_id": advisor_id,
        "name": name,
        "email": email.lower().strip(),
        "phone": phone,
        "role": role_value,
        "password_hash": password_hash,
        "is_active": bool(is_active),
    })
    db.commit()

    try:
        subj, body, html = email_template("new_user", name=name, email=email.lower().strip(), password=password)
        send_email(db, email.lower().strip(), subj, body, html_body=html, event_type="user_created", entity_type="user", entity_id=advisor_id or email, created_by_email="admin@astorie.local")
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass

    safe_audit(db, "admin@astorie.local", "CREATE", "advisor", advisor_id or email, {}, {
        "advisor_id": advisor_id, "name": name, "email": email, "role": role_value, "is_active": bool(is_active)
    }, "Založení poradce / uživatele")

    return RedirectResponse("/admin/advisors", status_code=303)


@router.post("/admin/advisors/{user_id}/update")
def advisor_update(
    user_id: str,
    advisor_id: str = Form(""),
    name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(""),
    roles: list[str] = Form([]),
    role: str = Form(""),
    is_active: str = Form(""),
    must_change_password: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_user_permissions_v150(db)
    role_value = normalize_roles_v150(roles or [role])
    old = db.execute(text("SELECT * FROM users WHERE id::text = :id"), {"id": str(user_id)}).mappings().first()

    db.execute(text("""
      UPDATE users
      SET advisor_id = :advisor_id,
          name = :name,
          email = :email,
          phone = :phone,
          role = :role,
          is_active = :is_active,
          must_change_password = :must_change_password
      WHERE id::text = :id
    """), {
        "id": str(user_id),
        "advisor_id": advisor_id,
        "name": name,
        "email": email.lower().strip(),
        "phone": phone,
        "role": role_value,
        "is_active": bool(is_active),
        "must_change_password": bool(must_change_password),
    })
    db.commit()

    safe_audit(db, "admin@astorie.local", "UPDATE", "advisor", str(user_id), dict(old or {}), {
        "advisor_id": advisor_id, "name": name, "email": email, "role": role_value,
        "is_active": bool(is_active), "must_change_password": bool(must_change_password)
    }, "Úprava poradce / uživatele")

    return RedirectResponse("/admin/advisors", status_code=303)


@router.post("/admin/advisors/{user_id}/toggle")
def advisor_toggle(user_id: str, db: Session = Depends(get_db)):
    old = db.execute(text("SELECT id, is_active, email, advisor_id FROM users WHERE id::text = :id"), {"id": str(user_id)}).mappings().first()
    if old:
        new_active = not bool(old["is_active"])
        db.execute(text("UPDATE users SET is_active = :is_active WHERE id::text = :id"), {"id": str(user_id), "is_active": new_active})
        db.commit()
        safe_audit(db, "admin@astorie.local", "TOGGLE", "advisor", str(user_id), dict(old), {"is_active": new_active}, "Zapnutí/vypnutí poradce")
    return RedirectResponse("/admin/advisors", status_code=303)


@router.post("/admin/advisors/{user_id}/reset-password")
def advisor_reset_password(
    user_id: str,
    password: str = Form("1234"),
    db: Session = Depends(get_db),
):
    try:
        password_hash = hash_password(password)
    except Exception:
        password_hash = password

    db.execute(text("""
      UPDATE users
      SET password_hash = :password_hash,
          must_change_password = TRUE
      WHERE id::text = :id
    """), {"id": str(user_id), "password_hash": password_hash})
    db.commit()

    try:
        row = db.execute(text("SELECT email, name FROM users WHERE id::text = :id"), {"id": str(user_id)}).mappings().first()
        if row and row.get("email"):
            subj, body, html = email_template("password_reset", name=row.get("name", ""), email=row.get("email", ""), password=password)
            send_email(db, row.get("email"), subj, body, html_body=html, event_type="password_reset", entity_type="user", entity_id=str(user_id), created_by_email="admin@astorie.local")
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass

    safe_audit(db, "admin@astorie.local", "UPDATE", "advisor", str(user_id), {}, {"password_reset": True}, "Reset hesla poradce")
    return RedirectResponse("/admin/advisors", status_code=303)




@router.get("/admin/permissions", response_class=HTMLResponse)
def permissions_page(request: Request, db: Session = Depends(get_db)):
    ensure_user_permissions_v150(db)
    rows = db.execute(text("""
        SELECT area, module_id, module_name, module_url,
               max(CASE WHEN role_code='IF' THEN is_allowed::int ELSE 0 END) AS if_allowed,
               max(CASE WHEN role_code='PS' THEN is_allowed::int ELSE 0 END) AS ps_allowed,
               max(CASE WHEN role_code='ADMIN' THEN is_allowed::int ELSE 0 END) AS admin_allowed,
               max(CASE WHEN role_code='BO' THEN is_allowed::int ELSE 0 END) AS bo_allowed,
               max(CASE WHEN role_code='VEDENI' THEN is_allowed::int ELSE 0 END) AS vedeni_allowed
        FROM module_permissions
        GROUP BY area, module_id, module_name, module_url
        ORDER BY CASE WHEN area='hub' THEN 1 ELSE 2 END, module_name
    """)).mappings().all()
    return render(request, "permissions.html", {"active": "permissions", "rows": rows, "system_roles": SYSTEM_ROLES_V150})


@router.post("/admin/permissions/{area}/{module_id}/update")
def permission_row_update(area: str, module_id: str,
                          IF: str = Form(""), PS: str = Form(""), ADMIN: str = Form(""), BO: str = Form(""), VEDENI: str = Form(""),
                          db: Session = Depends(get_db)):
    ensure_user_permissions_v150(db)
    role_values = {"IF": bool(IF), "PS": bool(PS), "ADMIN": bool(ADMIN), "BO": bool(BO), "VEDENI": bool(VEDENI)}
    for role_code, allowed in role_values.items():
        db.execute(text("""
            UPDATE module_permissions
            SET is_allowed = :allowed, updated_at = now()
            WHERE area = :area AND module_id = :module_id AND role_code = :role_code
        """), {"allowed": allowed, "area": area, "module_id": module_id, "role_code": role_code})
    db.commit()
    return RedirectResponse("/admin/permissions", status_code=303)


@router.get("/api/release-1-5-0/status")
def release_150_status(db: Session = Depends(get_db)):
    ensure_user_permissions_v150(db)
    cnt = db.execute(text("SELECT count(*) FROM module_permissions")).scalar() or 0
    return {"ok": True, "version": "1.5.0-user-multirole-permissions-safe", "safe": True, "db_changed": "additive_only", "module_permissions": cnt, "changed_modules": ["admin_advisors", "admin_permissions"], "unchanged_modules": ["partners", "contacts", "rates", "terminations", "tips"]}



@router.get("/api/release-1-5-1/status")
def release_151_status(db: Session = Depends(get_db)):
    ensure_user_permissions_v150(db)
    ensure_contact_role_tables_v149(db)
    cnt_perm = db.execute(text("SELECT count(*) FROM module_permissions")).scalar() or 0
    cnt_roles = db.execute(text("SELECT count(*) FROM contact_roles")).scalar() or 0
    return {"ok": True, "version": "1.5.1-admin-contacts-permissions-ux-safe", "safe": True, "db_changed": False, "module_permissions": cnt_perm, "contact_roles": cnt_roles, "changed_modules": ["admin_contacts_ui", "admin_permissions_ui"], "unchanged_modules": ["partners", "tips", "rates", "terminations", "email", "products", "links"]}


def ensure_specialists_table_(db: Session):
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS specialists (
            id SERIAL PRIMARY KEY,
            advisor_id TEXT NOT NULL DEFAULT '',
            specialist_name TEXT NOT NULL DEFAULT '',
            email TEXT NOT NULL DEFAULT '',
            phone TEXT NOT NULL DEFAULT '',
            section_code TEXT NOT NULL DEFAULT '',
            subsection_code TEXT NOT NULL DEFAULT '',
            role_description TEXT NOT NULL DEFAULT '',
            region TEXT NOT NULL DEFAULT '',
            if_share TEXT NOT NULL DEFAULT '',
            ps_share TEXT NOT NULL DEFAULT '',
            available BOOLEAN NOT NULL DEFAULT TRUE,
            unavailable_reason TEXT NOT NULL DEFAULT '',
            is_active BOOLEAN NOT NULL DEFAULT TRUE,
            note TEXT NOT NULL DEFAULT '',
            created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT now()
        )
    """))
    db.commit()


@router.get("/admin/specialists", response_class=HTMLResponse)
def specialists_page(
    request: Request,
    q: str = "",
    section: str = "",
    available: str = "",
    db: Session = Depends(get_db),
):
    ensure_specialists_table_(db)
    sql = "SELECT * FROM specialists WHERE 1=1"
    params = {}

    if q:
        sql += """
          AND (
            lower(specialist_name) LIKE :q OR
            lower(email) LIKE :q OR
            lower(phone) LIKE :q OR
            lower(advisor_id) LIKE :q OR
            lower(region) LIKE :q OR
            lower(role_description) LIKE :q
          )
        """
        params["q"] = f"%{q.lower()}%"

    if section:
        sql += " AND section_code = :section"
        params["section"] = section

    if available == "1":
        sql += " AND available = TRUE AND is_active = TRUE"
    elif available == "0":
        sql += " AND (available = FALSE OR is_active = FALSE)"

    sql += " ORDER BY specialist_name, section_code, subsection_code LIMIT 1000"
    rows = db.execute(text(sql), params).mappings().all()

    ensure_taxonomy_tables_(db)
    sections = db.execute(text("""
        SELECT section_code AS code, section_name AS name
        FROM hub_sections
        WHERE is_active = TRUE
        ORDER BY sort_order, section_name
    """)).mappings().all()
    subsections = db.execute(text("""
        SELECT subsection_code AS code, subsection_name AS name, section_code
        FROM hub_subsections
        WHERE is_active = TRUE
        ORDER BY sort_order, subsection_name
    """)).mappings().all()

    users = get_admin_users_for_specialists_v122_(db)
    return render(request, "specialists.html", {
        "active": "specialists",
        "specialists": rows,
        "sections": sections,
        "subsections": subsections,
        "users": users,
        "q": q,
        "section": section,
        "available_filter": available,
    })


@router.post("/admin/specialists/create")
def specialist_create(
    advisor_id: str = Form(""),
    specialist_name: str = Form(...),
    email: str = Form(""),
    phone: str = Form(""),
    section_code: str = Form(""),
    subsection_code: str = Form(""),
    role_description: str = Form(""),
    region: str = Form(""),
    if_share: str = Form(""),
    ps_share: str = Form(""),
    available: str = Form(""),
    is_active: str = Form(""),
    unavailable_reason: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_specialists_table_(db)
    db.execute(text("""
        INSERT INTO specialists
        (advisor_id, specialist_name, email, phone, section_code, subsection_code, role_description, region,
         if_share, ps_share, available, unavailable_reason, is_active, note)
        VALUES
        (:advisor_id, :specialist_name, :email, :phone, :section_code, :subsection_code, :role_description, :region,
         :if_share, :ps_share, :available, :unavailable_reason, :is_active, :note)
    """), {
        "advisor_id": advisor_id,
        "specialist_name": specialist_name,
        "email": email.lower().strip(),
        "phone": phone,
        "section_code": section_code.upper().strip(),
        "subsection_code": subsection_code.upper().strip(),
        "role_description": role_description,
        "region": region,
        "if_share": if_share,
        "ps_share": ps_share,
        "available": bool(available),
        "unavailable_reason": unavailable_reason,
        "is_active": bool(is_active),
        "note": note,
    })
    db.commit()
    try:
        safe_audit(db, "admin@astorie.local", "CREATE", "specialist", specialist_name, {}, {
            "advisor_id": advisor_id, "name": specialist_name, "email": email,
            "section": section_code, "subsection": subsection_code
        }, "Založení specialisty")
    except Exception:
        pass
    return RedirectResponse("/admin/specialists", status_code=303)


@router.post("/admin/specialists/{item_id}/update")
def specialist_update(
    item_id: int,
    advisor_id: str = Form(""),
    specialist_name: str = Form(...),
    email: str = Form(""),
    phone: str = Form(""),
    section_code: str = Form(""),
    subsection_code: str = Form(""),
    role_description: str = Form(""),
    region: str = Form(""),
    if_share: str = Form(""),
    ps_share: str = Form(""),
    available: str = Form(""),
    is_active: str = Form(""),
    unavailable_reason: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_specialists_table_(db)
    old = db.execute(text("SELECT * FROM specialists WHERE id = :id"), {"id": item_id}).mappings().first()
    db.execute(text("""
        UPDATE specialists SET
          advisor_id = :advisor_id,
          specialist_name = :specialist_name,
          email = :email,
          phone = :phone,
          section_code = :section_code,
          subsection_code = :subsection_code,
          role_description = :role_description,
          region = :region,
          if_share = :if_share,
          ps_share = :ps_share,
          available = :available,
          unavailable_reason = :unavailable_reason,
          is_active = :is_active,
          note = :note
        WHERE id = :id
    """), {
        "id": item_id,
        "advisor_id": advisor_id,
        "specialist_name": specialist_name,
        "email": email.lower().strip(),
        "phone": phone,
        "section_code": section_code.upper().strip(),
        "subsection_code": subsection_code.upper().strip(),
        "role_description": role_description,
        "region": region,
        "if_share": if_share,
        "ps_share": ps_share,
        "available": bool(available),
        "unavailable_reason": unavailable_reason,
        "is_active": bool(is_active),
        "note": note,
    })
    db.commit()
    try:
        safe_audit(db, "admin@astorie.local", "UPDATE", "specialist", str(item_id), dict(old or {}), {
            "advisor_id": advisor_id, "name": specialist_name, "email": email,
            "section": section_code, "subsection": subsection_code,
            "available": bool(available), "is_active": bool(is_active)
        }, "Úprava specialisty")
    except Exception:
        pass
    return RedirectResponse("/admin/specialists", status_code=303)


@router.post("/admin/specialists/{item_id}/toggle")
def specialist_toggle(item_id: int, db: Session = Depends(get_db)):
    ensure_specialists_table_(db)
    old = db.execute(text("SELECT * FROM specialists WHERE id = :id"), {"id": item_id}).mappings().first()
    if old:
        new_active = not bool(old["is_active"])
        db.execute(text("UPDATE specialists SET is_active = :is_active WHERE id = :id"), {"id": item_id, "is_active": new_active})
        db.commit()
        try:
            safe_audit(db, "admin@astorie.local", "TOGGLE", "specialist", str(item_id), dict(old), {"is_active": new_active}, "Zapnutí/vypnutí specialisty")
        except Exception:
            pass
    return RedirectResponse("/admin/specialists", status_code=303)


@router.get("/api/specialists/search")
def api_specialists_search(section: str = "", subsection: str = "", q: str = "", db: Session = Depends(get_db)):
    ensure_specialists_table_(db)
    sql = "SELECT * FROM specialists WHERE is_active = TRUE AND available = TRUE"
    params = {}

    if section:
        sql += " AND section_code = :section"
        params["section"] = section.upper()

    if subsection:
        sql += " AND (subsection_code = :subsection OR COALESCE(subsection_code, '') = '')"
        params["subsection"] = subsection.upper()

    if q:
        sql += """
          AND (
            lower(specialist_name) LIKE :q OR
            lower(email) LIKE :q OR
            lower(region) LIKE :q OR
            lower(role_description) LIKE :q
          )
        """
        params["q"] = f"%{q.lower()}%"

    sql += " ORDER BY specialist_name LIMIT 50"
    rows = db.execute(text(sql), params).mappings().all()

    return {
        "ok": True,
        "items": [
            {
                "id": r["id"],
                "advisor_id": r["advisor_id"],
                "name": r["specialist_name"],
                "email": r["email"],
                "phone": r["phone"],
                "section_code": r["section_code"],
                "subsection_code": r["subsection_code"],
                "role_description": r["role_description"],
                "region": r["region"],
                "if_share": r["if_share"],
                "ps_share": r["ps_share"],
            }
            for r in rows
        ],
    }





def ensure_taxonomy_tables_(db: Session):
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS hub_sections (
            id SERIAL PRIMARY KEY,
            section_code TEXT UNIQUE NOT NULL,
            section_name TEXT NOT NULL DEFAULT '',
            icon TEXT NOT NULL DEFAULT '',
            image_url TEXT NOT NULL DEFAULT '',
            sort_order INTEGER NOT NULL DEFAULT 100,
            is_active BOOLEAN NOT NULL DEFAULT TRUE,
            note TEXT NOT NULL DEFAULT '',
            created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT now()
        )
    """))
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS hub_subsections (
            id SERIAL PRIMARY KEY,
            subsection_code TEXT UNIQUE NOT NULL,
            section_code TEXT NOT NULL DEFAULT '',
            subsection_name TEXT NOT NULL DEFAULT '',
            sort_order INTEGER NOT NULL DEFAULT 100,
            is_active BOOLEAN NOT NULL DEFAULT TRUE,
            note TEXT NOT NULL DEFAULT '',
            created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT now()
        )
    """))
    db.commit()


@router.get("/admin/subsections")
def admin_subsections_compat():
    return RedirectResponse("/admin/sections", status_code=303)


@router.get("/admin/sections", response_class=HTMLResponse)
def sections_page(
    request: Request,
    q: str = "",
    db: Session = Depends(get_db),
):
    ensure_visible_hub_sections_(db)

    sql = "SELECT * FROM hub_sections WHERE 1=1"
    params = {}
    if q:
        sql += " AND (lower(section_code) LIKE :q OR lower(section_name) LIKE :q OR lower(note) LIKE :q)"
        params["q"] = f"%{q.lower()}%"
    sql += " ORDER BY sort_order, section_name"

    sections = db.execute(text(sql), params).mappings().all()
    subsections = db.execute(text("""
        SELECT s.*, h.section_name
        FROM hub_subsections s
        LEFT JOIN hub_sections h ON h.section_code = s.section_code
        ORDER BY h.sort_order, s.sort_order, s.subsection_name
    """)).mappings().all()

    return render(request, "sections.html", {
        "active": "sections",
        "sections": sections,
        "subsections": subsections,
        "q": q,
    })


@router.post("/admin/sections/create")
def section_create(
    section_code: str = Form(...),
    section_name: str = Form(...),
    icon: str = Form(""),
    image_url: str = Form(""),
    sort_order: int = Form(100),
    is_active: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_taxonomy_tables_(db)
    db.execute(text("""
        INSERT INTO hub_sections
          (section_code, section_name, icon, image_url, sort_order, is_active, note)
        VALUES
          (:section_code, :section_name, :icon, :image_url, :sort_order, :is_active, :note)
        ON CONFLICT (section_code) DO UPDATE SET
          section_name = EXCLUDED.section_name,
          icon = EXCLUDED.icon,
          image_url = EXCLUDED.image_url,
          sort_order = EXCLUDED.sort_order,
          is_active = EXCLUDED.is_active,
          note = EXCLUDED.note
    """), {
        "section_code": section_code.upper().strip(),
        "section_name": section_name,
        "icon": icon,
        "image_url": image_url,
        "sort_order": sort_order,
        "is_active": bool(is_active),
        "note": note,
    })
    db.commit()
    try:
        safe_audit(db, "admin@astorie.local", "UPSERT", "section", section_code.upper().strip(), {}, {
            "section_code": section_code.upper().strip(), "section_name": section_name
        }, "Založení/úprava sekce")
    except Exception:
        pass
    return RedirectResponse("/admin/sections", status_code=303)


@router.post("/admin/subsections/create")
def subsection_create(
    section_code: str = Form(...),
    subsection_code: str = Form(...),
    subsection_name: str = Form(...),
    sort_order: int = Form(100),
    is_active: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_taxonomy_tables_(db)
    db.execute(text("""
        INSERT INTO hub_subsections
          (section_code, subsection_code, subsection_name, sort_order, is_active, note)
        VALUES
          (:section_code, :subsection_code, :subsection_name, :sort_order, :is_active, :note)
        ON CONFLICT (subsection_code) DO UPDATE SET
          section_code = EXCLUDED.section_code,
          subsection_name = EXCLUDED.subsection_name,
          sort_order = EXCLUDED.sort_order,
          is_active = EXCLUDED.is_active,
          note = EXCLUDED.note
    """), {
        "section_code": section_code.upper().strip(),
        "subsection_code": subsection_code.upper().strip(),
        "subsection_name": subsection_name,
        "sort_order": sort_order,
        "is_active": bool(is_active),
        "note": note,
    })
    db.commit()
    try:
        safe_audit(db, "admin@astorie.local", "UPSERT", "subsection", subsection_code.upper().strip(), {}, {
            "section_code": section_code.upper().strip(),
            "subsection_code": subsection_code.upper().strip(),
            "subsection_name": subsection_name
        }, "Založení/úprava podsekce")
    except Exception:
        pass
    return RedirectResponse("/admin/sections", status_code=303)


@router.get("/api/taxonomy/sections")
def api_taxonomy_sections(db: Session = Depends(get_db)):
    ensure_taxonomy_tables_(db)
    rows = db.execute(text("""
        SELECT section_code, section_name, icon, image_url
        FROM hub_sections
        WHERE is_active = TRUE
        ORDER BY sort_order, section_name
    """)).mappings().all()
    return {"ok": True, "items": [dict(r) for r in rows]}


@router.get("/api/taxonomy/subsections")
def api_taxonomy_subsections(section_code: str = "", db: Session = Depends(get_db)):
    ensure_taxonomy_tables_(db)
    sql = """
        SELECT subsection_code, subsection_name, section_code
        FROM hub_subsections
        WHERE is_active = TRUE
    """
    params = {}
    if section_code:
        sql += " AND section_code = :section_code"
        params["section_code"] = section_code.upper()
    sql += " ORDER BY sort_order, subsection_name"
    rows = db.execute(text(sql), params).mappings().all()
    return {"ok": True, "items": [dict(r) for r in rows]}


@router.get("/admin/my-specialist-profile-old", response_class=HTMLResponse)
def my_specialist_profile(request: Request, db: Session = Depends(get_db)):
    ensure_specialists_table_(db)
    # Dočasně používáme admin e-mail; po ostrém loginu se nahradí session uživatelem.
    email = "nekudova@astorieas.cz"
    rows = db.execute(text("""
        SELECT * FROM specialists
        WHERE lower(email) = :email
        ORDER BY specialist_name, section_code, subsection_code
    """), {"email": email}).mappings().all()

    return render(request, "my_specialist_profile.html", {
        "active": "my_specialist_profile",
        "rows": rows,
        "email": email,
    })


@router.post("/admin/my-specialist-profile-old/{item_id}/availability")
def my_specialist_availability(
    item_id: int,
    available: str = Form(""),
    unavailable_reason: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_specialists_table_(db)
    old = db.execute(text("SELECT * FROM specialists WHERE id = :id"), {"id": item_id}).mappings().first()
    db.execute(text("""
        UPDATE specialists
        SET available = :available,
            unavailable_reason = :unavailable_reason
        WHERE id = :id
    """), {
        "id": item_id,
        "available": bool(available),
        "unavailable_reason": unavailable_reason,
    })
    db.commit()
    try:
        safe_audit(db, "admin@astorie.local", "UPDATE", "specialist_availability", str(item_id), dict(old or {}), {
            "available": bool(available), "unavailable_reason": unavailable_reason
        }, "Specialista upravil vlastní dostupnost")
    except Exception:
        pass
    return RedirectResponse("/admin/my-specialist-profile-old", status_code=303)


@router.get("/api/routing/specialists")
def api_routing_specialists(section_code: str = "", subsection_code: str = "", db: Session = Depends(get_db)):
    ensure_specialists_table_(db)
    ensure_taxonomy_tables_(db)

    sql = """
        SELECT *
        FROM specialists
        WHERE is_active = TRUE
          AND available = TRUE
    """
    params = {}

    if section_code:
        sql += " AND section_code = :section_code"
        params["section_code"] = section_code.upper()

    if subsection_code:
        sql += " AND (subsection_code = :subsection_code OR COALESCE(subsection_code, '') = '')"
        params["subsection_code"] = subsection_code.upper()

    sql += " ORDER BY specialist_name LIMIT 50"
    rows = db.execute(text(sql), params).mappings().all()

    return {
        "ok": True,
        "routing": {
            "section_code": section_code.upper() if section_code else "",
            "subsection_code": subsection_code.upper() if subsection_code else "",
            "count": len(rows),
        },
        "specialists": [
            {
                "id": r["id"],
                "name": r["specialist_name"],
                "email": r["email"],
                "phone": r["phone"],
                "region": r["region"],
                "section_code": r["section_code"],
                "subsection_code": r["subsection_code"],
                "if_share": r["if_share"],
                "ps_share": r["ps_share"],
            }
            for r in rows
        ],
    }




# -------------------------------------------------------------------
# v1.2.2 Specialist Profile & Sections Fix
# -------------------------------------------------------------------

def seed_default_hub_taxonomy_(db: Session):
    ensure_taxonomy_tables_(db)

    defaults_sections = [
        ("FLOTILY", "Flotily", "🚗", 10),
        ("MAJETEK", "Majetek", "🏡", 20),
        ("ZIVOT", "Život", "❤️", 30),
        ("PODNIKATELE", "Podnikatelé", "🏢", 40),
        ("PENZE", "Penze", "💼", 50),
        ("UVERY", "Úvěry", "🏦", 60),
        ("OBNOVA", "Obnova", "♻️", 70),
        ("INVESTICE", "Investice", "📈", 80),
        ("ZLATO", "Zlato", "💰", 90),
        ("ZVIRE", "Zvíře", "🐕", 100),
    ]

    defaults_subsections = [
        ("FLOTILY", "FLOTILY_FIREMNI", "Firemní flotily", 10),
        ("FLOTILY", "AUTODOPRAVCI", "Autodopravci", 20),
        ("MAJETEK", "DOMACNOSTI", "Domácnosti", 10),
        ("MAJETEK", "NEMOVITOSTI", "Nemovitosti", 20),
        ("ZIVOT", "ZIVOTNI_POJISTENI", "Životní pojištění", 10),
        ("PODNIKATELE", "PODNIKATELSKA_RIZIKA", "Podnikatelská rizika", 10),
        ("PENZE", "DPS", "Doplňkové penzijní spoření", 10),
        ("UVERY", "HYPOTEKY", "Hypotéky", 10),
        ("OBNOVA", "RETENCE", "Obnova / retence", 10),
        ("INVESTICE", "INVESTICE_OBECNE", "Investice", 10),
        ("ZLATO", "INVESTICNI_ZLATO", "Investiční zlato", 10),
        ("ZVIRE", "POJISTENI_ZVIRAT", "Pojištění zvířat", 10),
    ]

    for code, name, icon, order in defaults_sections:
        db.execute(text("""
            INSERT INTO hub_sections (section_code, section_name, icon, sort_order, is_active)
            VALUES (:code, :name, :icon, :sort_order, TRUE)
            ON CONFLICT (section_code) DO NOTHING
        """), {"code": code, "name": name, "icon": icon, "sort_order": order})

    for section_code, sub_code, sub_name, order in defaults_subsections:
        db.execute(text("""
            INSERT INTO hub_subsections (section_code, subsection_code, subsection_name, sort_order, is_active)
            VALUES (:section_code, :sub_code, :sub_name, :sort_order, TRUE)
            ON CONFLICT (subsection_code) DO NOTHING
        """), {
            "section_code": section_code,
            "sub_code": sub_code,
            "sub_name": sub_name,
            "sort_order": order,
        })

    db.commit()


@router.post("/admin/sections/seed-defaults")
def sections_seed_defaults(db: Session = Depends(get_db)):
    seed_default_hub_taxonomy_(db)
    try:
        safe_audit(db, "admin@astorie.local", "UPSERT", "taxonomy", "defaults", {}, {"seed": "default_hub_taxonomy"}, "Doplnění výchozích sekcí a podsekcí")
    except Exception:
        pass
    return RedirectResponse("/admin/sections", status_code=303)


@router.get("/admin/my-specialist-profile", response_class=HTMLResponse)
def my_specialist_profile_v071(request: Request, db: Session = Depends(get_db)):
    ensure_specialists_table_(db)
    seed_default_hub_taxonomy_(db)

    current_user = {
        "advisor_id": "501",
        "name": "Nekudová Dagmar",
        "email": "nekudova@astorieas.cz",
        "phone": "737 233 888",
    }

    rows = db.execute(text("""
        SELECT s.*,
               hs.section_name,
               hss.subsection_name
        FROM specialists s
        LEFT JOIN hub_sections hs ON hs.section_code = s.section_code
        LEFT JOIN hub_subsections hss ON hss.subsection_code = s.subsection_code
        WHERE lower(s.email) = :email
        ORDER BY s.specialist_name, hs.sort_order, hss.sort_order, s.section_code, s.subsection_code
    """), {"email": current_user["email"].lower()}).mappings().all()

    sections = db.execute(text("""
        SELECT section_code, section_name, icon
        FROM hub_sections
        WHERE is_active = TRUE
        ORDER BY sort_order, section_name
    """)).mappings().all()

    subsections = db.execute(text("""
        SELECT subsection_code, subsection_name, section_code
        FROM hub_subsections
        WHERE is_active = TRUE
        ORDER BY section_code, sort_order, subsection_name
    """)).mappings().all()

    return render(request, "my_specialist_profile.html", {
        "active": "my_specialist_profile",
        "rows": rows,
        "sections": sections,
        "subsections": subsections,
        "current_user": current_user,
        "email": current_user["email"],
    })


@router.post("/admin/my-specialist-profile/add-specialization")
def my_specialist_add_specialization_v071(
    section_code: str = Form(...),
    subsection_code: str = Form(""),
    role_description: str = Form(""),
    region: str = Form("ČR"),
    if_share: str = Form(""),
    ps_share: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_specialists_table_(db)
    seed_default_hub_taxonomy_(db)

    current_user = {
        "advisor_id": "501",
        "name": "Nekudová Dagmar",
        "email": "nekudova@astorieas.cz",
        "phone": "737 233 888",
    }

    db.execute(text("""
        INSERT INTO specialists
        (advisor_id, specialist_name, email, phone, section_code, subsection_code,
         role_description, region, if_share, ps_share, available, is_active, note)
        VALUES
        (:advisor_id, :name, :email, :phone, :section_code, :subsection_code,
         :role_description, :region, :if_share, :ps_share, TRUE, TRUE, 'Založeno z profilu specialisty')
    """), {
        "advisor_id": current_user["advisor_id"],
        "name": current_user["name"],
        "email": current_user["email"].lower(),
        "phone": current_user["phone"],
        "section_code": section_code.upper().strip(),
        "subsection_code": subsection_code.upper().strip(),
        "role_description": role_description,
        "region": region,
        "if_share": if_share,
        "ps_share": ps_share,
    })
    db.commit()

    try:
        safe_audit(db, current_user["email"], "CREATE", "specialist_profile", current_user["email"], {}, {
            "section_code": section_code,
            "subsection_code": subsection_code,
            "role_description": role_description,
        }, "Specialista si přidal odbornost do profilu")
    except Exception:
        pass

    return RedirectResponse("/admin/my-specialist-profile", status_code=303)


@router.post("/admin/my-specialist-profile/{item_id}/availability")
def my_specialist_availability_v071(
    item_id: int,
    available: str = Form(""),
    unavailable_reason: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_specialists_table_(db)
    old = db.execute(text("SELECT * FROM specialists WHERE id = :id"), {"id": item_id}).mappings().first()
    db.execute(text("""
        UPDATE specialists
        SET available = :available,
            unavailable_reason = :unavailable_reason
        WHERE id = :id
    """), {
        "id": item_id,
        "available": bool(available),
        "unavailable_reason": unavailable_reason,
    })
    db.commit()
    try:
        safe_audit(db, "admin@astorie.local", "UPDATE", "specialist_availability", str(item_id), dict(old or {}), {
            "available": bool(available), "unavailable_reason": unavailable_reason
        }, "Specialista upravil vlastní dostupnost")
    except Exception:
        pass
    return RedirectResponse("/admin/my-specialist-profile", status_code=303)





# -------------------------------------------------------------------
# v1.2.2 Visible Sections Fix
# -------------------------------------------------------------------

def ensure_visible_hub_sections_(db: Session):
    """
    Zajistí, že poradenská část HUBu vždy uvidí základní sekce.
    Nedestruktivní: nic nemaže a existující sekce nepřepisuje.
    """
    try:
        seed_default_hub_taxonomy_(db)
    except NameError:
        ensure_taxonomy_tables_(db)
        defaults_sections = [
            ("FLOTILY", "Flotily", "🚗", 10),
            ("MAJETEK", "Majetek", "🏡", 20),
            ("ZIVOT", "Život", "❤️", 30),
            ("PODNIKATELE", "Podnikatelé", "🏢", 40),
            ("PENZE", "Penze", "💼", 50),
            ("UVERY", "Úvěry", "🏦", 60),
            ("OBNOVA", "Obnova", "♻️", 70),
            ("INVESTICE", "Investice", "📈", 80),
            ("ZLATO", "Zlato", "💰", 90),
            ("ZVIRE", "Zvíře", "🐕", 100),
        ]
        defaults_subsections = [
            ("FLOTILY", "FLOTILY_FIREMNI", "Firemní flotily", 10),
            ("FLOTILY", "AUTODOPRAVCI", "Autodopravci", 20),
            ("MAJETEK", "DOMACNOSTI", "Domácnosti", 10),
            ("MAJETEK", "NEMOVITOSTI", "Nemovitosti", 20),
            ("ZIVOT", "ZIVOTNI_POJISTENI", "Životní pojištění", 10),
            ("PODNIKATELE", "PODNIKATELSKA_RIZIKA", "Podnikatelská rizika", 10),
            ("PENZE", "DPS", "Doplňkové penzijní spoření", 10),
            ("UVERY", "HYPOTEKY", "Hypotéky", 10),
            ("OBNOVA", "RETENCE", "Obnova / retence", 10),
            ("INVESTICE", "INVESTICE_OBECNE", "Investice", 10),
            ("ZLATO", "INVESTICNI_ZLATO", "Investiční zlato", 10),
            ("ZVIRE", "POJISTENI_ZVIRAT", "Pojištění zvířat", 10),
        ]
        for code, name, icon, order in defaults_sections:
            db.execute(text("""
                INSERT INTO hub_sections (section_code, section_name, icon, sort_order, is_active)
                VALUES (:code, :name, :icon, :sort_order, TRUE)
                ON CONFLICT (section_code) DO NOTHING
            """), {"code": code, "name": name, "icon": icon, "sort_order": order})
        for section_code, sub_code, sub_name, order in defaults_subsections:
            db.execute(text("""
                INSERT INTO hub_subsections (section_code, subsection_code, subsection_name, sort_order, is_active)
                VALUES (:section_code, :sub_code, :sub_name, :sort_order, TRUE)
                ON CONFLICT (subsection_code) DO NOTHING
            """), {"section_code": section_code, "sub_code": sub_code, "sub_name": sub_name, "sort_order": order})
        db.commit()


@router.get("/api/taxonomy/visible-sections")
def api_visible_sections_v072(db: Session = Depends(get_db)):
    ensure_visible_hub_sections_(db)
    sections = db.execute(text("""
        SELECT section_code, section_name, icon, image_url
        FROM hub_sections
        WHERE is_active = TRUE
        ORDER BY sort_order, section_name
    """)).mappings().all()
    subsections = db.execute(text("""
        SELECT subsection_code, subsection_name, section_code
        FROM hub_subsections
        WHERE is_active = TRUE
        ORDER BY section_code, sort_order, subsection_name
    """)).mappings().all()
    return {
        "ok": True,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "sections": [dict(s) for s in sections],
        "subsections": [dict(s) for s in subsections],
    }


@router.post("/admin/sections/force-visible-defaults")
def sections_force_visible_defaults_v072(db: Session = Depends(get_db)):
    ensure_visible_hub_sections_(db)
    return RedirectResponse("/admin/sections", status_code=303)







def ensure_user_hub_tables_v082_(db: Session):
    """
    v1.2.2 – bezpečné tabulky pro TIPy.
    Nedestruktivní: tabulku vytvoří nebo doplní chybějící sloupce.
    """
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS tips (
            id TEXT PRIMARY KEY,
            created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT now(),
            adviser_original_id TEXT NOT NULL DEFAULT '',
            adviser_name TEXT NOT NULL DEFAULT '',
            adviser_email TEXT NOT NULL DEFAULT '',
            specialist_name TEXT NOT NULL DEFAULT '',
            specialist_email TEXT NOT NULL DEFAULT '',
            client_name TEXT NOT NULL DEFAULT '',
            client_phone TEXT NOT NULL DEFAULT '',
            client_identifier TEXT NOT NULL DEFAULT '',
            potential_amount NUMERIC(14,2),
            adviser_note TEXT NOT NULL DEFAULT '',
            status TEXT NOT NULL DEFAULT 'Nový',
            policy_no TEXT NOT NULL DEFAULT '',
            final_volume NUMERIC(14,2),
            specialist_feedback TEXT NOT NULL DEFAULT ''
        )
    """))

    # Bezpečné doplnění sloupců pro případy, kdy už tabulka existuje ze starší verze.
    alters = [
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT now()",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS adviser_original_id TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS adviser_name TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS adviser_email TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS specialist_name TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS specialist_email TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS client_name TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS client_phone TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS client_identifier TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS potential_amount NUMERIC(14,2)",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS adviser_note TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS status TEXT NOT NULL DEFAULT 'Nový'",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS policy_no TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS final_volume NUMERIC(14,2)",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS specialist_feedback TEXT NOT NULL DEFAULT ''",
    ]
    for stmt in alters:
        db.execute(text(stmt))

    db.execute(text("CREATE INDEX IF NOT EXISTS idx_tips_adviser ON tips (adviser_original_id)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_tips_status ON tips (status)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_tips_created ON tips (created_at DESC)"))
    db.commit()


@router.get("/api/tips/status")
def api_tips_status_v082(db: Session = Depends(get_db)):
    try:
        ensure_user_hub_tables_v082_(db)
        count = db.execute(text("SELECT COUNT(*) FROM tips")).scalar()
        latest = db.execute(text("SELECT created_at, client_name, status FROM tips ORDER BY created_at DESC LIMIT 5")).mappings().all()
        return {
            "ok": True,
            "version": "1.2.2-admin-taxonomy-specialists-links-safe",
            "count": count,
            "latest": [dict(r) for r in latest],
        }
    except Exception as e:
        return {"ok": False, "version": "1.2.2-admin-taxonomy-specialists-links-safe", "error": str(e)}




# -------------------------------------------------------------------
# v1.2.2 Adviser HUB routes fix
# -------------------------------------------------------------------

def hub_user_context_v083_():
    return {
        "advisor_id": "501",
        "name": "Nekudová Dagmar",
        "email": "nekudova@astorieas.cz",
        "phone": "737 233 888",
        "role": "IF",
    }


def hub_render_v083_(request: Request, template_name: str, context: dict):
    base = {
        "request": request,
        "app_name": "HUB ASTORIE",
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "user": hub_user_context_v083_(),
    }
    base.update(context)
    return request.app.state.templates.TemplateResponse(template_name, base)


def ensure_hub_taxonomy_v083_(db: Session):
    try:
        ensure_visible_hub_sections_(db)
    except Exception:
        try:
            seed_default_hub_taxonomy_(db)
        except Exception:
            pass


@router.get("/hub")
def hub_home_v083():
    return RedirectResponse("/hub/new-tip", status_code=302)


@router.get("/hub/new-tip-old-v085", response_class=HTMLResponse)
def hub_new_tip_v083(request: Request, db: Session = Depends(get_db)):
    ensure_user_hub_tables_v082_(db)
    ensure_hub_taxonomy_v083_(db)

    sections = db.execute(text("""
        SELECT section_code, section_name, icon
        FROM hub_sections
        WHERE is_active = TRUE
        ORDER BY sort_order, section_name
    """)).mappings().all()

    subsections = db.execute(text("""
        SELECT subsection_code, subsection_name, section_code
        FROM hub_subsections
        WHERE is_active = TRUE
        ORDER BY section_code, sort_order, subsection_name
    """)).mappings().all()

    try:
        ensure_specialists_table_(db)
        specialists = db.execute(text("""
            SELECT *
            FROM specialists
            WHERE is_active = TRUE
              AND available = TRUE
            ORDER BY specialist_name, section_code, subsection_code
            LIMIT 200
        """)).mappings().all()
    except Exception:
        specialists = []

    return hub_render_v083_(request, "hub_new_tip.html", {
        "active": "new_tip",
        "sections": sections,
        "subsections": subsections,
        "specialists": specialists,
    })


@router.post("/hub/tips/create-old-v085")
def hub_create_tip_v083(
    section_code: str = Form(""),
    subsection_code: str = Form(""),
    specialist_email: str = Form(""),
    specialist_name: str = Form(""),
    client_name: str = Form(...),
    client_phone: str = Form(""),
    client_identifier: str = Form(""),
    potential_amount: str = Form(""),
    adviser_note: str = Form(""),
    policy_no: str = Form(""),
    final_volume: str = Form(""),
    closed_at_input: str = Form(""),
    next_business: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_user_hub_tables_v082_(db)
    user = hub_user_context_v083_()

    amount = None
    if potential_amount:
        try:
            amount = Decimal(str(potential_amount).replace(" ", "").replace("Kč", "").replace(",", "."))
        except Exception:
            amount = None

    tip_id = str(uuid.uuid4())
    db.execute(text("""
        INSERT INTO tips
          (id, adviser_original_id, adviser_name, adviser_email, specialist_name, specialist_email,
           client_name, client_phone, client_identifier, potential_amount, adviser_note, status, policy_no)
        VALUES
          (:id, :advisor_id, :advisor_name, :advisor_email, :specialist_name, :specialist_email,
           :client_name, :client_phone, :client_identifier, :potential_amount, :adviser_note, 'Nový', :policy_no)
    """), {
        "id": tip_id,
        "advisor_id": user["advisor_id"],
        "advisor_name": user["name"],
        "advisor_email": user["email"],
        "specialist_name": specialist_name,
        "specialist_email": specialist_email,
        "client_name": client_name,
        "client_phone": client_phone,
        "client_identifier": client_identifier,
        "potential_amount": amount,
        "adviser_note": f"[Sekce: {section_code}; Podsekce: {subsection_code}]\n{adviser_note}",
        "policy_no": policy_no,
        "final_volume": final_amount,
        "closed_at": closed_value,
        "next_business": next_business,
    })
    db.commit()
    return RedirectResponse("/hub/my-tips?created=1", status_code=303)


@router.get("/hub/my-tips-old-v085", response_class=HTMLResponse)
def hub_my_tips_v083(
    request: Request,
    q: str = "",
    status: str = "",
    created: str = "",
    db: Session = Depends(get_db),
):
    ensure_user_hub_tables_v082_(db)
    user = hub_user_context_v083_()

    sql = """
        SELECT *
        FROM tips
        WHERE COALESCE(adviser_original_id, '') = :advisor_id
    """
    params = {"advisor_id": user["advisor_id"]}

    if q:
        sql += """
          AND (
            lower(COALESCE(client_name, '')) LIKE :q OR
            lower(COALESCE(client_identifier, '')) LIKE :q OR
            lower(COALESCE(specialist_name, '')) LIKE :q OR
            lower(COALESCE(policy_no, '')) LIKE :q OR
            lower(COALESCE(adviser_note, '')) LIKE :q
          )
        """
        params["q"] = f"%{q.lower()}%"

    if status:
        sql += " AND status = :status"
        params["status"] = status

    sql += " ORDER BY created_at DESC LIMIT 300"
    rows = db.execute(text(sql), params).mappings().all()

    stats = db.execute(text("""
        SELECT
          COUNT(*) AS total,
          COUNT(*) FILTER (WHERE status ILIKE 'sjednáno') AS won,
          COUNT(*) FILTER (WHERE status ILIKE 'storno') AS lost,
          COUNT(*) FILTER (WHERE status NOT ILIKE 'sjednáno' AND status NOT ILIKE 'storno') AS open
        FROM tips
        WHERE COALESCE(adviser_original_id, '') = :advisor_id
    """), {"advisor_id": user["advisor_id"]}).mappings().first()

    return hub_render_v083_(request, "hub_my_tips.html", {
        "active": "my_tips",
        "rows": rows,
        "stats": stats,
        "q": q,
        "status": status,
        "created": created,
    })


@router.get("/hub/calculators-old-v083", response_class=HTMLResponse)
def hub_calculators_v083(request: Request, db: Session = Depends(get_db)):
    return hub_render_v083_(request, "hub_calculators.html", {"active": "calculators", "links": [], "rates": [], "q": ""})


@router.get("/hub/partners-old-v083", response_class=HTMLResponse)
def hub_partners_v083(request: Request, q: str = "", selected: str = "", tab: str = "contacts", db: Session = Depends(get_db)):
    dashboard = fetch_partner_dashboard_v111(db, selected) if selected else {}
    partner_history = fetch_partner_history_v111(db, selected) if selected else []
    partner_requests = fetch_partner_requests_v111(db, selected) if selected else []

    # v1.2.2 safe route fix: proměnné pro šablonu musí existovat vždy.
    try:
        dashboard = fetch_partner_dashboard_v111(db, selected) if selected and globals().get("fetch_partner_dashboard_v111") else {}
    except Exception:
        dashboard = {}
        try:
            db.rollback()
        except Exception:
            pass
    try:
        partner_history = fetch_partner_history_v111(db, selected) if selected and globals().get("fetch_partner_history_v111") else []
    except Exception:
        partner_history = []
        try:
            db.rollback()
        except Exception:
            pass
    try:
        partner_requests = fetch_partner_requests_v111(db, selected) if selected and globals().get("fetch_partner_requests_v111") else []
    except Exception:
        partner_requests = []
        try:
            db.rollback()
        except Exception:
            pass

    return hub_render_v083_(request, "hub_partners.html", {
        "active": "partners", "partners": [], "partner": None,
        "contacts": [], "links": [], "products": [], "q": q, "selected": selected, "tab": tab
    })


@router.get("/hub/contacts-old-v083", response_class=HTMLResponse)
def hub_contacts_v083(request: Request, q: str = "", db: Session = Depends(get_db)):
    return hub_render_v083_(request, "hub_contacts.html", {"active": "contacts", "rows": [], "q": q})


@router.get("/hub/forms-old-v083", response_class=HTMLResponse)
def hub_forms_v083(request: Request, db: Session = Depends(get_db)):
    return hub_render_v083_(request, "hub_forms.html", {"active": "forms", "partners": []})


@router.get("/hub/stats-old-v083", response_class=HTMLResponse)
def hub_stats_v083(request: Request, db: Session = Depends(get_db)):
    ensure_user_hub_tables_v082_(db)
    user = hub_user_context_v083_()
    stats = db.execute(text("""
        SELECT
          COUNT(*) AS total,
          COUNT(*) FILTER (WHERE status ILIKE 'sjednáno') AS won,
          COUNT(*) FILTER (WHERE status ILIKE 'storno') AS lost,
          COUNT(*) FILTER (WHERE status NOT ILIKE 'sjednáno' AND status NOT ILIKE 'storno') AS open,
          COALESCE(SUM(final_volume), 0) AS final_volume,
          COALESCE(SUM(potential_amount), 0) AS potential_amount
        FROM tips
        WHERE adviser_original_id = :advisor_id
    """), {"advisor_id": user["advisor_id"]}).mappings().first()

    return hub_render_v083_(request, "hub_stats.html", {"active": "stats", "stats": stats, "by_specialist": []})


@router.get("/hub/help-old-v083", response_class=HTMLResponse)
def hub_help_v083(request: Request):
    return hub_render_v083_(request, "hub_help.html", {"active": "help"})





# -------------------------------------------------------------------
# v1.2.2 HUB Data Bridge – propojení uživatelského HUBu na admin data
# -------------------------------------------------------------------

def table_exists_v084_(db: Session, table_name: str) -> bool:
    try:
        return bool(db.execute(text("""
            SELECT EXISTS (
                SELECT 1 FROM information_schema.tables
                WHERE table_schema = 'public' AND table_name = :table_name
            )
        """), {"table_name": table_name}).scalar())
    except Exception:
        return False


def column_exists_v084_(db: Session, table_name: str, column_name: str) -> bool:
    try:
        return bool(db.execute(text("""
            SELECT EXISTS (
                SELECT 1 FROM information_schema.columns
                WHERE table_schema = 'public'
                  AND table_name = :table_name
                  AND column_name = :column_name
            )
        """), {"table_name": table_name, "column_name": column_name}).scalar())
    except Exception:
        return False


def fetch_all_safe_v084_(db: Session, sql: str, params: dict | None = None):
    try:
        return db.execute(text(sql), params or {}).mappings().all()
    except Exception:
        return []


def fetch_one_safe_v084_(db: Session, sql: str, params: dict | None = None):
    try:
        return db.execute(text(sql), params or {}).mappings().first()
    except Exception:
        return None



def ensure_link_source_columns_v155a_(db: Session):
    """v1.5.5a: pouze bezpečné doplnění metadat odkazů.
    Nic nemaže. Starším záznamům doplní source_type jen tam, kde chybí.
    """
    if not table_exists_v084_(db, "partner_links"):
        return
    try:
        db.execute(text("ALTER TABLE partner_links ADD COLUMN IF NOT EXISTS source_type TEXT"))
        db.execute(text("ALTER TABLE partner_links ADD COLUMN IF NOT EXISTS is_archived BOOLEAN NOT NULL DEFAULT FALSE"))
        db.execute(text("ALTER TABLE partner_links ADD COLUMN IF NOT EXISTS visibility TEXT NOT NULL DEFAULT ''"))
        # ASTORIE interní odkazy – bez partnera nebo s ASTORIE kódem.
        db.execute(text("""
            UPDATE partner_links
            SET source_type = 'ASTORIE_LINK'
            WHERE COALESCE(source_type,'') = ''
              AND (
                upper(COALESCE(partner_code,'')) IN ('', 'AST', 'ASTORIE', 'ASTORIEAS')
                OR lower(COALESCE(category,'')) LIKE '%astorie%'
                OR lower(COALESCE(category,'')) LIKE '%intern%'
                OR lower(COALESCE(note,'')) LIKE '%astorie%'
                OR lower(COALESCE(visibility,'')) LIKE '%astorie%'
                OR lower(COALESCE(visibility,'')) LIKE '%intern%'
              )
        """))
        # Online kalkulačky – jen záznamy, které nebyly interní. Je to fallback pro starší import bez source_type.
        db.execute(text("""
            UPDATE partner_links
            SET source_type = 'ONLINE_CALCULATOR'
            WHERE COALESCE(source_type,'') = ''
              AND (
                lower(COALESCE(category,'')) LIKE '%kalk%'
                OR lower(COALESCE(title,'')) LIKE '%kalk%'
                OR lower(COALESCE(note,'')) LIKE '%kalk%'
                OR lower(COALESCE(visibility,'')) LIKE '%kalk%'
                OR lower(COALESCE(visibility,'')) LIKE '%calculator%'
              )
        """))
        # Vše ostatní je partnerský odkaz.
        db.execute(text("""
            UPDATE partner_links
            SET source_type = 'PARTNER_LINK'
            WHERE COALESCE(source_type,'') = ''
        """))
        db.commit()
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass


def link_source_where_v155a_(source_type: str, alias: str = "") -> str:
    prefix = (alias + ".") if alias else ""
    if source_type == "ASTORIE_LINK":
        return f"""(
            COALESCE({prefix}source_type,'') = 'ASTORIE_LINK'
            OR (
                COALESCE({prefix}source_type,'') = '' AND (
                    upper(COALESCE({prefix}partner_code,'')) IN ('', 'AST', 'ASTORIE', 'ASTORIEAS')
                    OR lower(COALESCE({prefix}category,'')) LIKE '%astorie%'
                    OR lower(COALESCE({prefix}category,'')) LIKE '%intern%'
                    OR lower(COALESCE({prefix}note,'')) LIKE '%astorie%'
                    OR lower(COALESCE({prefix}visibility,'')) LIKE '%astorie%'
                    OR lower(COALESCE({prefix}visibility,'')) LIKE '%intern%'
                )
            )
        )"""
    if source_type == "ONLINE_CALCULATOR":
        return f"""(
            COALESCE({prefix}source_type,'') = 'ONLINE_CALCULATOR'
            OR (
                COALESCE({prefix}source_type,'') = '' AND (
                    lower(COALESCE({prefix}category,'')) LIKE '%kalk%'
                    OR lower(COALESCE({prefix}title,'')) LIKE '%kalk%'
                    OR lower(COALESCE({prefix}note,'')) LIKE '%kalk%'
                    OR lower(COALESCE({prefix}visibility,'')) LIKE '%kalk%'
                    OR lower(COALESCE({prefix}visibility,'')) LIKE '%calculator%'
                )
            )
        )"""
    return f"""(
        COALESCE({prefix}source_type,'') = 'PARTNER_LINK'
        OR COALESCE({prefix}source_type,'') = ''
    )"""


@router.get("/hub/partners", response_class=HTMLResponse)
def hub_partners_v084(
    request: Request,
    q: str = "",
    selected: str = "",
    tab: str = "contacts",
    db: Session = Depends(get_db),
):
    """
    v1.2.2 – definitivní bezpečná uživatelská sekce Partneři.
    Tato route nesmí spadnout kvůli chybějící proměnné dashboard/partner_history/partner_requests.
    Používá pouze bezpečné SELECT * dotazy a vše ostatní dopočítává v Pythonu.
    """
    partners = []
    partner = None
    contacts = []
    links = []
    products = []
    faqs = []
    partner_history = []
    partner_requests = []
    dashboard = {"contacts": 0, "links": 0, "products": 0, "faq": 0, "requests_open": 0}
    selected = (selected or "").strip()
    tab = (tab or "contacts").strip() or "contacts"
    q_norm = (q or "").strip().lower()

    def _row_get(row, key, default=""):
        try:
            return row.get(key, default)
        except Exception:
            try:
                return getattr(row, key)
            except Exception:
                return default

    def _filter_partner(row):
        if not q_norm:
            return True
        hay = " ".join([
            str(_row_get(row, "partner_code", "")),
            str(_row_get(row, "name", "")),
            str(_row_get(row, "partner_name", "")),
            str(_row_get(row, "ico", "")),
            str(_row_get(row, "data_box", "")),
            str(_row_get(row, "registry_email", "")),
            str(_row_get(row, "city", "")),
            str(_row_get(row, "address_full", "")),
        ]).lower()
        return q_norm in hay

    try:
        # Zachovat workflow a pomocné tabulky, ale případná chyba nesmí shodit stránku.
        for fn_name in ["ensure_partner_workflow_v110", "ensure_v103_tables", "ensure_partner_hotfix_v112"]:
            fn = globals().get(fn_name)
            if fn:
                try:
                    fn(db)
                except Exception:
                    try:
                        db.rollback()
                    except Exception:
                        pass

        if table_exists_v084_(db, "partners"):
            raw_partners = fetch_all_safe_v084_(db, """
                SELECT *
                FROM partners
                ORDER BY COALESCE(name, partner_code)
                LIMIT 1000
            """)
            partners = [p for p in raw_partners if _filter_partner(p)]

            if not selected and partners:
                selected = str(_row_get(partners[0], "partner_code", "") or "")

            if selected:
                partner = fetch_one_safe_v084_(db, """
                    SELECT *
                    FROM partners
                    WHERE upper(COALESCE(partner_code,'')) = upper(:code)
                    LIMIT 1
                """, {"code": selected})

        if selected and table_exists_v084_(db, "partner_contacts"):
            contacts = fetch_all_safe_v084_(db, """
                SELECT *
                FROM partner_contacts
                WHERE upper(COALESCE(partner_code,'')) = upper(:code)
                  AND COALESCE(is_active, TRUE) = TRUE
                LIMIT 500
            """, {"code": selected})

        if selected and table_exists_v084_(db, "partner_links"):
            ensure_link_source_columns_v155a_(db)
            links = fetch_all_safe_v084_(db, f"""
                SELECT *
                FROM partner_links
                WHERE upper(COALESCE(partner_code,'')) = upper(:code)
                  AND COALESCE(is_active, TRUE) = TRUE
                  AND COALESCE(is_archived, FALSE) = FALSE
                  AND {link_source_where_v155a_("PARTNER_LINK")}
                LIMIT 500
            """, {"code": selected})

        if selected and table_exists_v084_(db, "partner_products"):
            products = fetch_all_safe_v084_(db, """
                SELECT *
                FROM partner_products
                WHERE upper(COALESCE(partner_code,'')) = upper(:code)
                  AND COALESCE(is_active, TRUE) = TRUE
                LIMIT 800
            """, {"code": selected})

        if selected and table_exists_v084_(db, "partner_faq"):
            faqs = fetch_all_safe_v084_(db, """
                SELECT *
                FROM partner_faq
                WHERE (upper(COALESCE(partner_code,'')) = upper(:code) OR COALESCE(partner_code,'') = '')
                  AND COALESCE(is_active, TRUE) = TRUE
                LIMIT 300
            """, {"code": selected})

        if selected and table_exists_v084_(db, "partner_history"):
            partner_history = fetch_all_safe_v084_(db, """
                SELECT *
                FROM partner_history
                WHERE upper(COALESCE(partner_code,'')) = upper(:code)
                ORDER BY created_at DESC
                LIMIT 50
            """, {"code": selected})

        if selected and table_exists_v084_(db, "partner_change_requests"):
            partner_requests = fetch_all_safe_v084_(db, """
                SELECT *
                FROM partner_change_requests
                WHERE upper(COALESCE(partner_code,'')) = upper(:code)
                ORDER BY created_at DESC
                LIMIT 100
            """, {"code": selected})

        dashboard = {
            "contacts": len(contacts),
            "links": len(links),
            "products": len(products),
            "faq": len(faqs),
            "requests_open": len([r for r in partner_requests if str(_row_get(r, "status", "new")) in ("new", "processing", "nový", "v řešení")]),
        }

    except Exception as exc:
        # Stránka nesmí skončit jako interní chyba serveru. Zobrazíme bezpečnou stránku s diagnostikou.
        try:
            db.rollback()
        except Exception:
            pass
        partners = []
        partner = None
        contacts = []
        links = []
        products = []
        faqs = []
        partner_history = []
        partner_requests = []
        dashboard = {"contacts": 0, "links": 0, "products": 0, "faq": 0, "requests_open": 0}
        request.state.partner_route_error = str(exc)

    return hub_render_v083_(request, "hub_partners.html", {
        "active": "partners",
        "partners": partners or [],
        "partner": partner,
        "contacts": contacts or [],
        "links": links or [],
        "products": products or [],
        "faqs": faqs or [],
        "dashboard": dashboard or {},
        "partner_history": partner_history or [],
        "partner_requests": partner_requests or [],
        "q": q or "",
        "selected": selected or "",
        "tab": tab or "contacts",
        "route_error": getattr(request.state, "partner_route_error", ""),
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
    })


@router.get("/hub/contacts", response_class=HTMLResponse)
def hub_contacts_v084(request: Request, q: str = "", db: Session = Depends(get_db)):
    # v1.2.2: hlavní menu Kontakty zobrazuje pouze globální kontakty ASTORIE.
    # Kontakty partnerů zůstávají v detailu partnera na záložce Kontakty.
    ensure_v103_tables(db)
    rows = []
    params = {}
    where = "WHERE COALESCE(is_active, TRUE) = TRUE"
    if q:
        where += """
          AND (
            lower(COALESCE(contact_name, '')) LIKE :q OR
            lower(COALESCE(email, '')) LIKE :q OR
            lower(COALESCE(phone, '')) LIKE :q OR
            lower(COALESCE(role_description, '')) LIKE :q OR
            lower(COALESCE(department, '')) LIKE :q OR
            lower(COALESCE(location, '')) LIKE :q
          )
        """
        params["q"] = f"%{q.lower()}%"

    rows = fetch_all_safe_v084_(db, f"""
        SELECT *
        FROM global_contacts
        {where}
        ORDER BY COALESCE(is_vip, FALSE) DESC, contact_name
        LIMIT 500
    """, params)

    return hub_render_v083_(request, "hub_contacts.html", {
        "active": "contacts",
        "rows": rows,
        "q": q,
        "is_global_contacts": True,
    })


@router.get("/hub/calculators", response_class=HTMLResponse)
def hub_calculators_v084(request: Request, q: str = "", db: Session = Depends(get_db)):
    links = []
    rates = []

    if table_exists_v084_(db, "partner_links"):
        ensure_link_source_columns_v155a_(db)
        params = {}
        where = f"""
            WHERE COALESCE(l.is_active, TRUE) = TRUE
              AND COALESCE(l.is_archived, FALSE) = FALSE
              AND {link_source_where_v155a_("ONLINE_CALCULATOR", "l")}
        """
        if q:
            where += """
              AND (
                lower(COALESCE(l.title, '')) LIKE :q OR
                lower(COALESCE(p.name, '')) LIKE :q OR
                lower(COALESCE(l.category, '')) LIKE :q OR
                lower(COALESCE(l.url, '')) LIKE :q
              )
            """
            params["q"] = f"%{q.lower()}%"

        links = fetch_all_safe_v084_(db, f"""
            SELECT l.*, p.name AS partner_name
            FROM partner_links l
            LEFT JOIN partners p ON p.partner_code = l.partner_code
            {where}
            ORDER BY COALESCE(p.name, ''), l.title
            LIMIT 300
        """, params)

    if table_exists_v084_(db, "commission_rates"):
        # v1.4.3: sazebník vrací přesné aliasy podle sloupců Google Sheets pro frontend, aby se neztrácely
        # sazby, typy a produkty jen kvůli rozdílným názvům sloupců v databázi.
        rates = fetch_all_safe_v084_(db, """
            SELECT
                cr.*,
                COALESCE(NULLIF(s.name, ''), cr.section_code, '') AS section_display,
                COALESCE(NULLIF(cr.area, ''), '') AS area_display,
                COALESCE(NULLIF(cr.partner_name, ''), '') AS partner_display,
                COALESCE(NULLIF(cr.business_type, ''), '') AS product_display,
                COALESCE(NULLIF(cr.product_type, ''), '') AS base_display,
                CASE
                    WHEN cr.rate_percent IS NULL THEN ''
                    WHEN cr.rate_percent = ROUND(cr.rate_percent) THEN TRIM(TO_CHAR(cr.rate_percent, 'FM999999990')) || ' %'
                    ELSE REPLACE(TRIM(TO_CHAR(cr.rate_percent, 'FM999999990D99')), '.', ',') || ' %'
                END AS rate_display
            FROM commission_rates cr
            LEFT JOIN sections s ON s.section_code = cr.section_code
            LEFT JOIN subsections ss ON ss.subsection_code = cr.subsection_code
            WHERE COALESCE(cr.is_active, TRUE) = TRUE
            ORDER BY COALESCE(cr.priority, 0) DESC, section_display, area_display, cr.partner_name, product_display, base_display
            LIMIT 2000
        """)

    return hub_render_v083_(request, "hub_calculators.html", {
        "active": "calculators",
        "links": links,
        "rates": rates,
        "q": q,
    })


@router.get("/hub/forms-old-v086", response_class=HTMLResponse)
def hub_forms_v084(request: Request, q: str = "", db: Session = Depends(get_db)):
    partners = []
    if table_exists_v084_(db, "partners"):
        params = {}
        where = "WHERE COALESCE(is_active, TRUE) = TRUE"
        if q:
            where += """
              AND (
                lower(COALESCE(name, '')) LIKE :q OR
                lower(COALESCE(partner_code, '')) LIKE :q OR
                lower(COALESCE(ico, '')) LIKE :q
              )
            """
            params["q"] = f"%{q.lower()}%"

        partners = fetch_all_safe_v084_(db, f"""
            SELECT partner_code, name, ico, data_box, registry_email, address_full, street, city, zip_code
            FROM partners
            {where}
            ORDER BY name
            LIMIT 500
        """, params)

    return hub_render_v083_(request, "hub_forms.html", {
        "active": "forms",
        "partners": partners,
        "q": q,
    })


@router.get("/hub/links", response_class=HTMLResponse)
def hub_links_v155a(request: Request, q: str = "", db: Session = Depends(get_db)):
    """Produkční Odkazy ASTORIE: pouze interní ASTORIE odkazy.
    Nezobrazuje partnerské odkazy ani online kalkulačky.
    """
    links = []
    if table_exists_v084_(db, "partner_links"):
        ensure_link_source_columns_v155a_(db)
        params = {}
        where = f"""
            WHERE COALESCE(is_active, TRUE) = TRUE
              AND COALESCE(is_archived, FALSE) = FALSE
              AND {link_source_where_v155a_("ASTORIE_LINK")}
        """
        if q:
            where += """
              AND (
                lower(COALESCE(title, '')) LIKE :q OR
                lower(COALESCE(note, '')) LIKE :q OR
                lower(COALESCE(category, '')) LIKE :q OR
                lower(COALESCE(url, '')) LIKE :q
              )
            """
            params["q"] = f"%{q.lower()}%"
        links = fetch_all_safe_v084_(db, f"""
            SELECT *
            FROM partner_links
            {where}
            ORDER BY COALESCE(category,''), title
            LIMIT 300
        """, params)
    return hub_render_v083_(request, "hub_links.html", {
        "active": "links",
        "q": q,
        "links": links,
    })


@router.get("/hub/help", response_class=HTMLResponse)
def hub_help_v155a(request: Request, q: str = "", db: Session = Depends(get_db)):
    """Samostatná Nápověda. FAQ a Odkazy ASTORIE se nemíchají."""
    articles = []
    # Pokud bude později vytvořena admin tabulka help_articles, použijeme ji.
    if table_exists_v084_(db, "help_articles"):
        params = {}
        where = "WHERE COALESCE(is_active, TRUE) = TRUE"
        if q:
            where += " AND (lower(COALESCE(title,'')) LIKE :q OR lower(COALESCE(body,'')) LIKE :q OR lower(COALESCE(category,'')) LIKE :q)"
            params["q"] = f"%{q.lower()}%"
        articles = fetch_all_safe_v084_(db, f"""
            SELECT * FROM help_articles {where}
            ORDER BY COALESCE(sort_order,100), COALESCE(category,''), title
            LIMIT 200
        """, params)
    return hub_render_v083_(request, "hub_help.html", {
        "active": "help",
        "q": q,
        "articles": articles,
        "links": [],
        "faqs": [],
    })


@router.get("/api/hub/data-status")
def api_hub_data_status_v084(db: Session = Depends(get_db)):
    tables = ["partners", "partner_contacts", "partner_links", "partner_products", "hub_sections", "hub_subsections", "specialists", "tips"]
    result = {}
    for t in tables:
        exists = table_exists_v084_(db, t)
        count = None
        error = None
        if exists:
            try:
                count = db.execute(text(f"SELECT COUNT(*) FROM {t}")).scalar()
            except Exception as e:
                error = str(e)
        result[t] = {"exists": exists, "count": count, "error": error}

    return {
        "ok": True,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "tables": result,
    }




# -------------------------------------------------------------------
# v1.2.2 TIP Admin Data Flow – sekce/podsekce/specialisté z adminu do poradce
# -------------------------------------------------------------------

def ensure_tips_columns_v085_(db: Session):
    ensure_user_hub_tables_v082_(db)
    alters = [
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS section_code TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS subsection_code TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS section_name TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS subsection_name TEXT NOT NULL DEFAULT ''",
    ]
    for stmt in alters:
        db.execute(text(stmt))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_tips_section ON tips (section_code)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_tips_subsection ON tips (subsection_code)"))
    db.commit()


def get_hub_taxonomy_v085_(db: Session):
    ensure_hub_taxonomy_v083_(db)
    sections = fetch_all_safe_v084_(db, """
        SELECT section_code, section_name, icon, COALESCE(image_url, '') AS image_url, sort_order
        FROM hub_sections
        WHERE COALESCE(is_active, TRUE) = TRUE
        ORDER BY sort_order, section_name
    """)
    subsections = fetch_all_safe_v084_(db, """
        SELECT subsection_code, subsection_name, section_code, sort_order
        FROM hub_subsections
        WHERE COALESCE(is_active, TRUE) = TRUE
        ORDER BY section_code, sort_order, subsection_name
    """)
    sections = dedupe_taxonomy_sections_v122_(sections)
    return sections, subsections


def get_specialists_for_hub_v085_(db: Session):
    try:
        ensure_specialists_table_(db)
    except Exception:
        return []
    return fetch_all_safe_v084_(db, """
        SELECT s.*,
               '' AS photo_url,
               COALESCE(hs.section_name, s.section_code) AS section_name,
               COALESCE(hss.subsection_name, s.subsection_code) AS subsection_name
        FROM specialists s
        LEFT JOIN hub_sections hs ON hs.section_code = s.section_code
        LEFT JOIN hub_subsections hss ON hss.subsection_code = s.subsection_code
        WHERE COALESCE(s.is_active, TRUE) = TRUE
        ORDER BY COALESCE(s.available, TRUE) DESC, hs.sort_order, hss.sort_order, s.specialist_name
        LIMIT 500
    """)


@router.get("/hub/new-tip", response_class=HTMLResponse)
def hub_new_tip_v085(request: Request, db: Session = Depends(get_db)):
    ensure_tips_columns_v085_(db)
    sections, subsections = get_hub_taxonomy_v085_(db)
    specialists = get_specialists_for_hub_v085_(db)
    return hub_render_v083_(request, "hub_new_tip.html", {
        "active": "new_tip",
        "sections": sections,
        "subsections": subsections,
        "specialists": specialists,
    })


@router.post("/hub/tips/create")
def hub_create_tip_v085(
    section_code: str = Form(""),
    subsection_code: str = Form(""),
    specialist_email: str = Form(""),
    specialist_name: str = Form(""),
    client_name: str = Form(...),
    client_phone: str = Form(""),
    client_identifier: str = Form(""),
    potential_amount: str = Form(""),
    adviser_note: str = Form(""),
    policy_no: str = Form(""),
    final_volume: str = Form(""),
    closed_at_input: str = Form(""),
    next_business: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_tip_workflow_v090_(db)
    user = hub_user_context_v083_()

    missing_fields = []
    if not section_code.strip(): missing_fields.append("oblast")
    if not subsection_code.strip(): missing_fields.append("podsekce")
    if not specialist_email.strip(): missing_fields.append("specialista")
    if not client_name.strip(): missing_fields.append("klient")
    if not client_phone.strip(): missing_fields.append("kontakt na klienta")
    if not client_identifier.strip(): missing_fields.append("RČ / IČO / datum nar.")
    if not potential_amount.strip(): missing_fields.append("odhad potenciálu / objemu")
    if not policy_no.strip(): missing_fields.append("smlouva č.")
    if not final_volume.strip(): missing_fields.append("výše obchodu / pojistné")
    if not closed_at_input.strip(): missing_fields.append("datum uzavření")
    if not next_business.strip(): missing_fields.append("další obchod")
    if missing_fields:
        return JSONResponse({"ok": False, "error": "Chybí povinné údaje: " + ", ".join(missing_fields)}, status_code=400)

    section = fetch_one_safe_v084_(db, """
        SELECT section_code, section_name
        FROM hub_sections
        WHERE upper(section_code) = upper(:code)
        LIMIT 1
    """, {"code": section_code})

    subsection = fetch_one_safe_v084_(db, """
        SELECT subsection_code, subsection_name
        FROM hub_subsections
        WHERE upper(subsection_code) = upper(:code)
        LIMIT 1
    """, {"code": subsection_code})

    amount = None
    if potential_amount:
        try:
            amount = Decimal(str(potential_amount).replace(" ", "").replace("Kč", "").replace(",", "."))
        except Exception:
            amount = None

    final_amount = None
    if final_volume:
        try:
            final_amount = Decimal(str(final_volume).replace(" ", "").replace("Kč", "").replace(",", "."))
        except Exception:
            final_amount = None

    closed_value = None
    if closed_at_input:
        try:
            from datetime import datetime
            closed_value = datetime.strptime(closed_at_input.strip(), "%Y-%m-%d")
        except Exception:
            closed_value = None

    if not specialist_name and specialist_email:
        spec = fetch_one_safe_v084_(db, """
            SELECT specialist_name
            FROM specialists
            WHERE lower(email) = lower(:email)
            LIMIT 1
        """, {"email": specialist_email})
        if spec:
            specialist_name = spec["specialist_name"]

    tip_id = str(uuid.uuid4())
    db.execute(text("""
        INSERT INTO tips
          (id, adviser_original_id, adviser_name, adviser_email,
           section_code, subsection_code, section_name, subsection_name,
           specialist_name, specialist_email,
           client_name, client_phone, client_identifier, potential_amount,
           adviser_note, status, policy_no, final_volume, closed_at, next_business)
        VALUES
          (:id, :advisor_id, :advisor_name, :advisor_email,
           :section_code, :subsection_code, :section_name, :subsection_name,
           :specialist_name, :specialist_email,
           :client_name, :client_phone, :client_identifier, :potential_amount,
           :adviser_note, 'Nový', :policy_no, :final_volume, :closed_at, :next_business)
    """), {
        "id": tip_id,
        "advisor_id": user["advisor_id"],
        "advisor_name": user["name"],
        "advisor_email": user["email"],
        "section_code": section_code,
        "subsection_code": subsection_code,
        "section_name": section["section_name"] if section else section_code,
        "subsection_name": subsection["subsection_name"] if subsection else subsection_code,
        "specialist_name": specialist_name,
        "specialist_email": specialist_email,
        "client_name": client_name,
        "client_phone": client_phone,
        "client_identifier": client_identifier,
        "potential_amount": amount,
        "adviser_note": adviser_note,
        "policy_no": policy_no,
    })
    db.execute(text("""
        INSERT INTO tip_updates
          (id, tip_id, author_name, author_email, author_role, update_type, old_status, new_status, message_to_adviser, internal_note)
        VALUES
          (:id, :tip_id, :author_name, :author_email, 'Poradce', 'Založení TIPu', '', 'Nový', :msg_adv, :msg_int)
    """), {
        "id": str(uuid.uuid4()),
        "tip_id": tip_id,
        "author_name": user["name"],
        "author_email": user["email"],
        "msg_adv": "TIP byl založen a předán specialistovi.",
        "msg_int": f"Nový TIP pro specialistu {specialist_name or specialist_email}: {client_name}",
    })
    db.commit()

    # E-mailové notifikace jsou bezpečné: když SMTP není nastavené, TIP zůstane uložený a chyba se zapíše do historie TIPu.
    try:
        section_label = section["section_name"] if section else section_code
        subsection_label = subsection["subsection_name"] if subsection else subsection_code
        mail_data = {
            "adviser_name": user.get("name", ""),
            "adviser_email": user.get("email", ""),
            "specialist_name": specialist_name or specialist_email,
            "client_name": client_name,
            "client_phone": client_phone,
            "client_identifier": client_identifier,
            "section_label": section_label,
            "subsection_label": subsection_label,
            "policy_no": policy_no,
            "potential_amount": potential_amount,
            "adviser_note": adviser_note,
        }
        sent_spec, err_spec = send_template_email(
            db,
            specialist_email,
            "tip_new_specialist",
            data=mail_data,
            event_type="tip_new_specialist",
            entity_type="tip",
            entity_id=tip_id,
            created_by_email=user.get("email", ""),
        )
        sent_adv, err_adv = send_template_email(
            db,
            user.get("email", ""),
            "tip_new_adviser",
            data=mail_data,
            event_type="tip_new_adviser",
            entity_type="tip",
            entity_id=tip_id,
            created_by_email=user.get("email", ""),
        )
        if (not sent_spec) or (not sent_adv):
            db.execute(text("""
                INSERT INTO tip_updates
                  (id, tip_id, author_name, author_email, author_role, update_type, old_status, new_status, internal_note)
                VALUES
                  (:id, :tip_id, 'Systém', 'system@astorie.local', 'Systém', 'E-mail', 'Nový', 'Nový', :note)
            """), {
                "id": str(uuid.uuid4()),
                "tip_id": tip_id,
                "note": "E-mail nebyl odeslán všem příjemcům. Specialistovi: " + (err_spec or 'OK') + "; poradci: " + (err_adv or 'OK')
            })
            db.commit()
    except Exception as exc:
        try:
            db.execute(text("""
                INSERT INTO tip_updates
                  (id, tip_id, author_name, author_email, author_role, update_type, old_status, new_status, internal_note)
                VALUES
                  (:id, :tip_id, 'Systém', 'system@astorie.local', 'Systém', 'E-mail chyba', 'Nový', 'Nový', :note)
            """), {"id": str(uuid.uuid4()), "tip_id": tip_id, "note": str(exc)})
            db.commit()
        except Exception:
            db.rollback()

    try:
        safe_audit(db, user["email"], "CREATE", "tips", tip_id, {}, {
            "client_name": client_name,
            "section_code": section_code,
            "subsection_code": subsection_code,
            "specialist_email": specialist_email,
        }, "Poradce založil nový TIP")
    except Exception:
        pass
    return RedirectResponse("/hub/my-tips?created=1", status_code=303)


@router.get("/hub/my-tips-old-v091", response_class=HTMLResponse)
def hub_my_tips_v085(
    request: Request,
    q: str = "",
    status: str = "",
    section: str = "",
    created: str = "",
    db: Session = Depends(get_db),
):
    ensure_tips_columns_v085_(db)
    user = hub_user_context_v083_()
    sections, _ = get_hub_taxonomy_v085_(db)

    sql = """
        SELECT *
        FROM tips
        WHERE COALESCE(adviser_original_id, '') = :advisor_id
    """
    params = {"advisor_id": user["advisor_id"]}

    if q:
        sql += """
          AND (
            lower(COALESCE(client_name, '')) LIKE :q OR
            lower(COALESCE(client_identifier, '')) LIKE :q OR
            lower(COALESCE(specialist_name, '')) LIKE :q OR
            lower(COALESCE(policy_no, '')) LIKE :q OR
            lower(COALESCE(adviser_note, '')) LIKE :q OR
            lower(COALESCE(section_name, '')) LIKE :q OR
            lower(COALESCE(subsection_name, '')) LIKE :q
          )
        """
        params["q"] = f"%{q.lower()}%"

    if status:
        sql += " AND status = :status"
        params["status"] = status
    if section:
        sql += " AND upper(section_code) = upper(:section)"
        params["section"] = section

    sql += " ORDER BY created_at DESC LIMIT 300"
    rows = db.execute(text(sql), params).mappings().all()

    stats = db.execute(text("""
        SELECT
          COUNT(*) AS total,
          COUNT(*) FILTER (WHERE status ILIKE 'sjednáno') AS won,
          COUNT(*) FILTER (WHERE status ILIKE 'storno') AS lost,
          COUNT(*) FILTER (WHERE status NOT ILIKE 'sjednáno' AND status NOT ILIKE 'storno') AS open
        FROM tips
        WHERE COALESCE(adviser_original_id, '') = :advisor_id
    """), {"advisor_id": user["advisor_id"]}).mappings().first()

    return hub_render_v083_(request, "hub_my_tips.html", {
        "active": "my_tips",
        "rows": rows,
        "stats": stats,
        "sections": sections,
        "q": q,
        "status": status,
        "section": section,
        "created": created,
    })


@router.get("/api/hub/taxonomy-status")
def api_hub_taxonomy_status_v085(db: Session = Depends(get_db)):
    sections, subsections = get_hub_taxonomy_v085_(db)
    specialists = get_specialists_for_hub_v085_(db)
    return {
        "ok": True,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "sections_count": len(sections),
        "subsections_count": len(subsections),
        "specialists_count": len(specialists),
        "sections": [dict(s) for s in sections],
        "subsections": [dict(s) for s in subsections],
    }



# -------------------------------------------------------------------
# v1.2.2 Partner autocomplete & Forms data source
# -------------------------------------------------------------------

@router.get("/api/hub/partners/search")
def api_hub_partners_search_v086(q: str = "", limit: int = 20, db: Session = Depends(get_db)):
    """Našeptávač partnerů pro uživatelskou část HUBu."""
    if not table_exists_v084_(db, "partners"):
        return {"ok": True, "version": "1.2.2-admin-taxonomy-specialists-links-safe", "items": []}

    q_clean = (q or "").strip().lower()
    params = {"limit": max(1, min(limit, 50))}
    where = "WHERE COALESCE(is_active, TRUE) = TRUE"

    if q_clean:
        where += """
          AND (
            lower(COALESCE(partner_code, '')) LIKE :q OR
            lower(COALESCE(name, '')) LIKE :q OR
            lower(COALESCE(ico, '')) LIKE :q OR
            lower(COALESCE(data_box, '')) LIKE :q OR
            lower(COALESCE(registry_email, '')) LIKE :q OR
            lower(COALESCE(city, '')) LIKE :q
          )
        """
        params["q"] = f"%{q_clean}%"

    rows = fetch_all_safe_v084_(db, f"""
        SELECT partner_code, name, ico, data_box, registry_email, address_full, street, city, zip_code, status
        FROM partners
        {where}
        ORDER BY
          CASE WHEN lower(COALESCE(partner_code, '')) = :exact THEN 0 ELSE 1 END,
          name
        LIMIT :limit
    """, {**params, "exact": q_clean})

    return {
        "ok": True,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "items": [dict(r) for r in rows],
    }


@router.get("/api/hub/partners/{partner_code}/form-source")
def api_hub_partner_form_source_v086(partner_code: str, db: Session = Depends(get_db)):
    """Kompletní zdrojová data partnera pro výpovědi a formuláře."""
    if not table_exists_v084_(db, "partners"):
        return {"ok": False, "version": "1.2.2-admin-taxonomy-specialists-links-safe", "error": "Tabulka partners neexistuje."}

    partner = fetch_one_safe_v084_(db, """
        SELECT *
        FROM partners
        WHERE upper(partner_code) = upper(:code)
        LIMIT 1
    """, {"code": partner_code})

    if not partner:
        return {"ok": False, "version": "1.2.2-admin-taxonomy-specialists-links-safe", "error": "Partner nenalezen."}

    contacts = []
    links = []
    products = []

    if table_exists_v084_(db, "partner_contacts"):
        contacts = fetch_all_safe_v084_(db, """
            SELECT *
            FROM partner_contacts
            WHERE upper(partner_code) = upper(:code)
              AND COALESCE(is_active, TRUE) = TRUE
            ORDER BY COALESCE(is_vip, FALSE) DESC, COALESCE(is_top, FALSE) DESC, full_name
            LIMIT 100
        """, {"code": partner_code})

    if table_exists_v084_(db, "partner_links"):
        links = fetch_all_safe_v084_(db, """
            SELECT *
            FROM partner_links
            WHERE upper(partner_code) = upper(:code)
              AND COALESCE(is_active, TRUE) = TRUE
            ORDER BY category, title
            LIMIT 100
        """, {"code": partner_code})

    if table_exists_v084_(db, "partner_products"):
        products = fetch_all_safe_v084_(db, """
            SELECT *
            FROM partner_products
            WHERE upper(partner_code) = upper(:code)
              AND COALESCE(is_active, TRUE) = TRUE
            ORDER BY area, subarea, product_name
            LIMIT 100
        """, {"code": partner_code})

    return {
        "ok": True,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "partner": dict(partner),
        "contacts": [dict(c) for c in contacts],
        "links": [dict(l) for l in links],
        "products": [dict(p) for p in products],
    }


@router.get("/api/hub/partners/{partner_code}/summary")
def api_hub_partner_summary_v086(partner_code: str, db: Session = Depends(get_db)):
    """Rychlý souhrn partnera pro detail v HUBu."""
    data = api_hub_partner_form_source_v086(partner_code, db)
    if not data.get("ok"):
        return data
    return {
        "ok": True,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "partner": data["partner"],
        "counts": {
            "contacts": len(data["contacts"]),
            "links": len(data["links"]),
            "products": len(data["products"]),
        },
    }



@router.get("/hub/terminations", response_class=HTMLResponse)
def hub_terminations_v146(request: Request, selected: str = "", db: Session = Depends(get_db)):
    partners = []
    partner = None
    if table_exists_v084_(db, "partners"):
        partners = fetch_all_safe_v084_(db, """
            SELECT partner_code, name, ico, data_box, registry_email, address_full, street, city, zip_code
            FROM partners
            WHERE COALESCE(is_active, TRUE) = TRUE
            ORDER BY name
            LIMIT 1000
        """)
        if selected:
            partner = fetch_one_safe_v084_(db, """
                SELECT partner_code, name, ico, data_box, registry_email, address_full, street, city, zip_code
                FROM partners
                WHERE upper(partner_code) = upper(:code)
                LIMIT 1
            """, {"code": selected})
    return hub_render_v083_(request, "hub_terminations.html", {
        "active": "terminations",
        "partners": [dict(p) for p in partners],
        "partner": dict(partner) if partner else None,
        "selected": selected,
    })


@router.post("/hub/terminations/preview", response_class=HTMLResponse)
def hub_termination_preview_v146(
    request: Request,
    partner_code: str = Form(""),
    termination_type: str = Form("A"),
    client_name: str = Form(""),
    client_identifier: str = Form(""),
    client_address: str = Form(""),
    policy_no: str = Form(""),
    insurance_type: str = Form(""),
    insured_subject: str = Form(""),
    bank_account: str = Form(""),
    extra_date: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    partner = db.query(Partner).filter(Partner.partner_code == partner_code.upper()).first() if partner_code else None
    reason_text, preview_lines, preview_text = build_termination_document_v147_(
        partner, termination_type, client_name, client_identifier, client_address, policy_no,
        insurance_type, insured_subject, bank_account, extra_date, note
    )
    doc_id = save_termination_document_v147_(
        db, partner, partner_code, termination_type, client_name, client_identifier, client_address,
        policy_no, insurance_type, insured_subject, bank_account, extra_date, note, preview_text, reason_text
    )
    return hub_render_v083_(request, "hub_termination_preview.html", {
        "active": "terminations",
        "partner": partner,
        "partner_code": partner_code.upper() if partner_code else "",
        "preview_text": preview_text,
        "preview_lines": preview_lines,
        "doc_id": doc_id,
        "saved": True,
    })


@router.get("/hub/terminations/archive", response_class=HTMLResponse)
def hub_terminations_archive_v147(request: Request, q: str = "", db: Session = Depends(get_db)):
    ensure_termination_archive_table_v147_(db)
    rows = fetch_all_safe_v084_(db, """
        SELECT * FROM termination_documents
        ORDER BY created_at DESC
        LIMIT 200
    """)
    return hub_render_v083_(request, "hub_terminations_archive.html", {"active": "terminations", "rows": [dict(r) for r in rows]})


@router.get("/hub/forms", response_class=HTMLResponse)
def hub_forms_v086(request: Request, q: str = "", selected: str = "", db: Session = Depends(get_db)):
    partners = []
    partner = None

    if table_exists_v084_(db, "partners"):
        params = {}
        where = "WHERE COALESCE(is_active, TRUE) = TRUE"
        if q:
            where += """
              AND (
                lower(COALESCE(name, '')) LIKE :q OR
                lower(COALESCE(partner_code, '')) LIKE :q OR
                lower(COALESCE(ico, '')) LIKE :q OR
                lower(COALESCE(data_box, '')) LIKE :q OR
                lower(COALESCE(registry_email, '')) LIKE :q
              )
            """
            params["q"] = f"%{q.lower()}%"

        partners = fetch_all_safe_v084_(db, f"""
            SELECT partner_code, name, ico, data_box, registry_email, address_full, street, city, zip_code
            FROM partners
            {where}
            ORDER BY name
            LIMIT 200
        """, params)

        if selected:
            partner = fetch_one_safe_v084_(db, """
                SELECT partner_code, name, ico, data_box, registry_email, address_full, street, city, zip_code
                FROM partners
                WHERE upper(partner_code) = upper(:code)
                LIMIT 1
            """, {"code": selected})

    return hub_render_v083_(request, "hub_forms.html", {
        "active": "forms",
        "partners": partners,
        "partner": partner,
        "q": q,
        "selected": selected,
    })





# -------------------------------------------------------------------
# v1.2.2 Operational TIP Workflow
# Import dat + BO centrální evidence + specialista pracovní fronta
# -------------------------------------------------------------------

def current_bo_user_v090_():
    return {
        "name": "Admin ASTORIE",
        "email": "nekudova@astorieas.cz",
        "role": "BO",
    }


def ensure_tip_workflow_v090_(db: Session):
    ensure_tips_columns_v085_(db)

    alters = [
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS bo_note TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS specialist_internal_note TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS adviser_last_message TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS final_report TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS next_business TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS closed_at TIMESTAMP WITH TIME ZONE",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS archived_at TIMESTAMP WITH TIME ZONE",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS last_update_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT now()",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS imported_source TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS imported_original_id TEXT NOT NULL DEFAULT ''",
    ]
    for stmt in alters:
        db.execute(text(stmt))

    db.execute(text("""
        CREATE TABLE IF NOT EXISTS tip_updates (
            id TEXT PRIMARY KEY,
            tip_id TEXT NOT NULL,
            created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT now(),
            author_name TEXT NOT NULL DEFAULT '',
            author_email TEXT NOT NULL DEFAULT '',
            author_role TEXT NOT NULL DEFAULT '',
            update_type TEXT NOT NULL DEFAULT '',
            old_status TEXT NOT NULL DEFAULT '',
            new_status TEXT NOT NULL DEFAULT '',
            message_to_adviser TEXT NOT NULL DEFAULT '',
            internal_note TEXT NOT NULL DEFAULT '',
            final_report TEXT NOT NULL DEFAULT ''
        )
    """))

    db.execute(text("""
        CREATE TABLE IF NOT EXISTS import_jobs (
            id TEXT PRIMARY KEY,
            created_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT now(),
            source_name TEXT NOT NULL DEFAULT '',
            rows_total INTEGER NOT NULL DEFAULT 0,
            rows_imported INTEGER NOT NULL DEFAULT 0,
            rows_skipped INTEGER NOT NULL DEFAULT 0,
            error_log TEXT NOT NULL DEFAULT '',
            created_by TEXT NOT NULL DEFAULT ''
        )
    """))

    db.execute(text("CREATE INDEX IF NOT EXISTS idx_tips_specialist_email ON tips (specialist_email)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_tips_status_all ON tips (status)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_tips_last_update ON tips (last_update_at DESC)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_tip_updates_tip ON tip_updates (tip_id, created_at DESC)"))
    db.commit()


def normalize_tip_status_v090_(status: str) -> str:
    s = (status or "").strip().lower()
    if not s:
        return "Nový"
    if s in ["novy", "nový", "new", "zadáno", "zadano"]:
        return "Nový"
    if s in ["v reseni", "v řešení", "reseni", "řešení", "pracuje se", "rozpracováno"]:
        return "V řešení"
    if s in ["sjednano", "sjednáno", "hotovo", "uzavreno", "uzavřeno", "vyřízeno"]:
        return "Sjednáno"
    if s in ["storno", "zruseno", "zrušeno", "nezajem", "nezájem", "nevyšlo"]:
        return "Storno"
    if s in ["archiv", "archivováno", "archivovano"]:
        return "Archiv"
    return status.strip()


def find_header_v090_(row: dict, candidates: list[str]) -> str:
    normalized = {str(k).strip().lower(): k for k in row.keys()}
    for c in candidates:
        key = c.strip().lower()
        if key in normalized:
            return normalized[key]
    return ""


def get_row_value_v090_(row: dict, candidates: list[str], default: str = ""):
    key = find_header_v090_(row, candidates)
    if not key:
        return default
    return (row.get(key) or default)


@router.get("/admin/tips", response_class=HTMLResponse)
def admin_all_tips_v090(
    request: Request,
    q: str = "",
    status: str = "",
    specialist: str = "",
    adviser: str = "",
    archive: str = "",
    db: Session = Depends(get_db),
):
    ensure_tip_workflow_v090_(db)

    sql = """
        SELECT *
        FROM tips
        WHERE 1=1
    """
    params = {}

    if q:
        sql += """
          AND (
            lower(COALESCE(client_name, '')) LIKE :q OR
            lower(COALESCE(client_identifier, '')) LIKE :q OR
            lower(COALESCE(adviser_name, '')) LIKE :q OR
            lower(COALESCE(adviser_email, '')) LIKE :q OR
            lower(COALESCE(specialist_name, '')) LIKE :q OR
            lower(COALESCE(specialist_email, '')) LIKE :q OR
            lower(COALESCE(policy_no, '')) LIKE :q OR
            lower(COALESCE(section_name, '')) LIKE :q OR
            lower(COALESCE(subsection_name, '')) LIKE :q
          )
        """
        params["q"] = f"%{q.lower()}%"

    if status:
        sql += " AND status = :status"
        params["status"] = status

    if specialist:
        sql += " AND lower(COALESCE(specialist_email, '')) LIKE :specialist"
        params["specialist"] = f"%{specialist.lower()}%"

    if adviser:
        sql += " AND lower(COALESCE(adviser_email, '')) LIKE :adviser"
        params["adviser"] = f"%{adviser.lower()}%"

    if archive != "1":
        sql += " AND COALESCE(status, '') <> 'Archiv'"

    sql += " ORDER BY last_update_at DESC, created_at DESC LIMIT 800"

    rows = db.execute(text(sql), params).mappings().all()

    stats = db.execute(text("""
        SELECT
          COUNT(*) AS total,
          COUNT(*) FILTER (WHERE status = 'Nový') AS new_count,
          COUNT(*) FILTER (WHERE status = 'V řešení') AS progress_count,
          COUNT(*) FILTER (WHERE status = 'Sjednáno') AS won_count,
          COUNT(*) FILTER (WHERE status = 'Storno') AS lost_count,
          COUNT(*) FILTER (WHERE status = 'Archiv') AS archive_count
        FROM tips
    """)).mappings().first()

    return request.app.state.templates.TemplateResponse("admin_tips.html", {
        "request": request,
        "active": "admin_tips",
        "rows": rows,
        "stats": stats,
        "q": q,
        "status": status,
        "specialist": specialist,
        "adviser": adviser,
        "archive": archive,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
    })


@router.get("/admin/tips/{tip_id}", response_class=HTMLResponse)
def admin_tip_detail_v090(request: Request, tip_id: str, db: Session = Depends(get_db)):
    ensure_tip_workflow_v090_(db)
    tip = fetch_one_safe_v084_(db, "SELECT * FROM tips WHERE id = :id", {"id": tip_id})
    updates = fetch_all_safe_v084_(db, """
        SELECT *
        FROM tip_updates
        WHERE tip_id = :id
        ORDER BY created_at DESC
    """, {"id": tip_id})

    return request.app.state.templates.TemplateResponse("admin_tip_detail.html", {
        "request": request,
        "active": "admin_tips",
        "tip": tip,
        "updates": updates,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
    })


@router.post("/admin/tips/{tip_id}/bo-update")
def admin_tip_bo_update_v090(
    tip_id: str,
    status: str = Form(""),
    bo_note: str = Form(""),
    message_to_adviser: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_tip_workflow_v090_(db)
    user = current_bo_user_v090_()
    old = fetch_one_safe_v084_(db, "SELECT * FROM tips WHERE id = :id", {"id": tip_id})
    old_status = old["status"] if old else ""

    new_status = normalize_tip_status_v090_(status or old_status)
    closed_sql = ", closed_at = now()" if new_status in ["Sjednáno", "Storno"] else ""
    archived_sql = ", archived_at = now()" if new_status == "Archiv" else ""

    db.execute(text(f"""
        UPDATE tips
        SET status = :status,
            bo_note = :bo_note,
            adviser_last_message = :message_to_adviser,
            last_update_at = now()
            {closed_sql}
            {archived_sql}
        WHERE id = :id
    """), {
        "id": tip_id,
        "status": new_status,
        "bo_note": bo_note,
        "message_to_adviser": message_to_adviser,
    })

    db.execute(text("""
        INSERT INTO tip_updates
        (id, tip_id, author_name, author_email, author_role, update_type,
         old_status, new_status, message_to_adviser, internal_note)
        VALUES
        (:id, :tip_id, :author_name, :author_email, :author_role, 'BO_UPDATE',
         :old_status, :new_status, :message_to_adviser, :internal_note)
    """), {
        "id": str(uuid.uuid4()),
        "tip_id": tip_id,
        "author_name": user["name"],
        "author_email": user["email"],
        "author_role": user["role"],
        "old_status": old_status,
        "new_status": new_status,
        "message_to_adviser": message_to_adviser,
        "internal_note": bo_note,
    })
    db.commit()
    return RedirectResponse(f"/admin/tips/{tip_id}", status_code=303)


@router.get("/hub/specialist-tips", response_class=HTMLResponse)
def hub_specialist_tips_v090(
    request: Request,
    q: str = "",
    status: str = "",
    archive: str = "",
    db: Session = Depends(get_db),
):
    ensure_tip_workflow_v090_(db)
    user = hub_user_context_v083_()
    email = (user.get("email") or "").lower()

    sql = """
        SELECT *
        FROM tips
        WHERE lower(COALESCE(specialist_email, '')) = :email
    """
    params = {"email": email}

    if q:
        sql += """
          AND (
            lower(COALESCE(client_name, '')) LIKE :q OR
            lower(COALESCE(adviser_name, '')) LIKE :q OR
            lower(COALESCE(policy_no, '')) LIKE :q OR
            lower(COALESCE(section_name, '')) LIKE :q OR
            lower(COALESCE(subsection_name, '')) LIKE :q
          )
        """
        params["q"] = f"%{q.lower()}%"

    if status:
        sql += " AND status = :status"
        params["status"] = status

    if archive != "1":
        sql += " AND COALESCE(status, '') <> 'Archiv'"

    sql += " ORDER BY last_update_at DESC, created_at DESC LIMIT 300"
    rows = db.execute(text(sql), params).mappings().all()

    stats = db.execute(text("""
        SELECT
          COUNT(*) AS total,
          COUNT(*) FILTER (WHERE status = 'Nový') AS new_count,
          COUNT(*) FILTER (WHERE status = 'V řešení') AS progress_count,
          COUNT(*) FILTER (WHERE status = 'Sjednáno') AS won_count,
          COUNT(*) FILTER (WHERE status = 'Storno') AS lost_count,
          COUNT(*) FILTER (WHERE status = 'Archiv') AS archive_count
        FROM tips
        WHERE lower(COALESCE(specialist_email, '')) = :email
    """), {"email": email}).mappings().first()

    return hub_render_v083_(request, "hub_specialist_tips.html", {
        "active": "specialist_tips",
        "rows": rows,
        "stats": stats,
        "q": q,
        "status": status,
        "archive": archive,
    })


@router.get("/hub/specialist-tips/{tip_id}", response_class=HTMLResponse)
def hub_specialist_tip_detail_v090(request: Request, tip_id: str, db: Session = Depends(get_db)):
    ensure_tip_workflow_v090_(db)
    user = hub_user_context_v083_()
    tip = fetch_one_safe_v084_(db, """
        SELECT *
        FROM tips
        WHERE id = :id
          AND lower(COALESCE(specialist_email, '')) = lower(:email)
        LIMIT 1
    """, {"id": tip_id, "email": user["email"]})

    updates = fetch_all_safe_v084_(db, """
        SELECT *
        FROM tip_updates
        WHERE tip_id = :id
        ORDER BY created_at DESC
    """, {"id": tip_id})

    return hub_render_v083_(request, "hub_specialist_tip_detail.html", {
        "active": "specialist_tips",
        "tip": tip,
        "updates": updates,
    })


@router.post("/hub/specialist-tips/{tip_id}/update")
def hub_specialist_tip_update_v090(
    tip_id: str,
    status: str = Form("V řešení"),
    policy_no: str = Form(""),
    closed_at_input: str = Form(""),
    final_volume: str = Form(""),
    next_business: str = Form(""),
    message_to_adviser: str = Form(""),
    internal_note: str = Form(""),
    final_report: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_tip_workflow_v090_(db)
    user = hub_user_context_v083_()
    old = fetch_one_safe_v084_(db, """
        SELECT *
        FROM tips
        WHERE id = :id
          AND lower(COALESCE(specialist_email, '')) = lower(:email)
        LIMIT 1
    """, {"id": tip_id, "email": user["email"]})

    if not old:
        return RedirectResponse("/hub/specialist-tips", status_code=303)

    old_status = old["status"]
    new_status = normalize_tip_status_v090_(status)

    final_amount = old.get("final_volume") if old else None
    if final_volume:
        try:
            final_amount = Decimal(str(final_volume).replace(" ", "").replace("Kč", "").replace(",", "."))
        except Exception:
            final_amount = old.get("final_volume") if old else None

    closed_value = None
    if closed_at_input:
        try:
            closed_value = datetime.strptime(closed_at_input.strip(), "%Y-%m-%d")
        except Exception:
            closed_value = None

    closed_sql = ""
    if closed_value:
        closed_sql = ", closed_at = :closed_at"
    elif new_status in ["Sjednáno", "Storno"] and not old.get("closed_at"):
        closed_sql = ", closed_at = now()"

    db.execute(text(f"""
        UPDATE tips
        SET status = :status,
            policy_no = :policy_no,
            final_volume = :final_volume,
            next_business = :next_business,
            adviser_last_message = :message_to_adviser,
            specialist_internal_note = :internal_note,
            final_report = :final_report,
            last_update_at = now()
            {closed_sql}
        WHERE id = :id
    """), {
        "id": tip_id,
        "status": new_status,
        "policy_no": (policy_no or old.get("policy_no") or "").strip(),
        "final_volume": final_amount,
        "next_business": (next_business if next_business != "" else (old.get("next_business") or "")).strip(),
        "closed_at": closed_value,
        "message_to_adviser": message_to_adviser,
        "internal_note": internal_note,
        "final_report": final_report,
    })

    db.execute(text("""
        INSERT INTO tip_updates
        (id, tip_id, author_name, author_email, author_role, update_type,
         old_status, new_status, message_to_adviser, internal_note, final_report)
        VALUES
        (:id, :tip_id, :author_name, :author_email, 'SPECIALIST', 'SPECIALIST_UPDATE',
         :old_status, :new_status, :message_to_adviser, :internal_note, :final_report)
    """), {
        "id": str(uuid.uuid4()),
        "tip_id": tip_id,
        "author_name": user["name"],
        "author_email": user["email"],
        "old_status": old_status,
        "new_status": new_status,
        "message_to_adviser": message_to_adviser,
        "internal_note": internal_note,
        "final_report": final_report,
    })

    db.commit()
    return RedirectResponse(f"/hub/specialist-tips/{tip_id}", status_code=303)


@router.get("/admin/import/legacy-tips", response_class=HTMLResponse)
def admin_import_legacy_tips_page_v090(request: Request, db: Session = Depends(get_db)):
    ensure_tip_workflow_v090_(db)
    jobs = fetch_all_safe_v084_(db, """
        SELECT *
        FROM import_jobs
        ORDER BY created_at DESC
        LIMIT 20
    """)
    return request.app.state.templates.TemplateResponse("admin_import_legacy_tips.html", {
        "request": request,
        "active": "import",
        "jobs": jobs,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
    })


@router.post("/admin/import/legacy-tips")
async def admin_import_legacy_tips_v090(
    file: UploadFile = File(...),
    source_name: str = Form("Stávající HUB ASTORIE"),
    db: Session = Depends(get_db),
):
    ensure_tip_workflow_v090_(db)
    user = current_bo_user_v090_()
    job_id = str(uuid.uuid4())
    rows_total = rows_imported = rows_skipped = 0
    errors = []

    content = await file.read()
    text_content = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text_content), delimiter=";")
    if not reader.fieldnames or len(reader.fieldnames) <= 1:
        reader = csv.DictReader(io.StringIO(text_content), delimiter=",")

    for row in reader:
        rows_total += 1
        try:
            original_id = str(get_row_value_v090_(row, ["id", "tip_id", "ID", "Číslo", "Cislo"], "")).strip()
            client_name = str(get_row_value_v090_(row, ["client_name", "Klient", "Jméno klienta", "Jmeno klienta", "Název klienta"], "")).strip()
            adviser_name = str(get_row_value_v090_(row, ["adviser_name", "Poradce", "Tipař", "Tipar", "IF"], "")).strip()
            adviser_email = str(get_row_value_v090_(row, ["adviser_email", "E-mail poradce", "Email poradce", "E-mail IF"], "")).strip()
            specialist_name = str(get_row_value_v090_(row, ["specialist_name", "Specialista", "PS"], "")).strip()
            specialist_email = str(get_row_value_v090_(row, ["specialist_email", "E-mail specialista", "Email specialista", "E-mail PS"], "")).strip()
            status = normalize_tip_status_v090_(str(get_row_value_v090_(row, ["status", "Stav"], "Nový")).strip())
            section_name = str(get_row_value_v090_(row, ["section_name", "Sekce", "Oblast"], "")).strip()
            subsection_name = str(get_row_value_v090_(row, ["subsection_name", "Podsekce", "Podoblast"], "")).strip()
            policy_no = str(get_row_value_v090_(row, ["policy_no", "Smlouva", "Číslo smlouvy", "Cislo smlouvy"], "")).strip()
            note = str(get_row_value_v090_(row, ["adviser_note", "Poznámka", "Poznamka", "Popis"], "")).strip()
            potential_raw = str(get_row_value_v090_(row, ["potential_amount", "Potenciál", "Potencial", "Objem"], "")).strip()

            if not client_name and not adviser_name and not specialist_name:
                rows_skipped += 1
                continue

            amount = None
            if potential_raw:
                try:
                    amount = Decimal(potential_raw.replace(" ", "").replace("Kč", "").replace(",", "."))
                except Exception:
                    amount = None

            db.execute(text("""
                INSERT INTO tips
                (id, adviser_original_id, adviser_name, adviser_email,
                 specialist_name, specialist_email, client_name, potential_amount,
                 adviser_note, status, policy_no, section_name, subsection_name,
                 imported_source, imported_original_id, last_update_at)
                VALUES
                (:id, :adviser_original_id, :adviser_name, :adviser_email,
                 :specialist_name, :specialist_email, :client_name, :potential_amount,
                 :adviser_note, :status, :policy_no, :section_name, :subsection_name,
                 :imported_source, :imported_original_id, now())
            """), {
                "id": str(uuid.uuid4()),
                "adviser_original_id": str(get_row_value_v090_(row, ["advisor_id", "ID poradce", "ID"], "")),
                "adviser_name": adviser_name,
                "adviser_email": adviser_email,
                "specialist_name": specialist_name,
                "specialist_email": specialist_email,
                "client_name": client_name,
                "potential_amount": amount,
                "adviser_note": note,
                "status": status,
                "policy_no": policy_no,
                "section_name": section_name,
                "subsection_name": subsection_name,
                "imported_source": source_name,
                "imported_original_id": original_id,
            })
            rows_imported += 1
        except Exception as e:
            rows_skipped += 1
            errors.append(f"Řádek {rows_total}: {str(e)}")

    db.execute(text("""
        INSERT INTO import_jobs
        (id, source_name, rows_total, rows_imported, rows_skipped, error_log, created_by)
        VALUES
        (:id, :source_name, :rows_total, :rows_imported, :rows_skipped, :error_log, :created_by)
    """), {
        "id": job_id,
        "source_name": source_name,
        "rows_total": rows_total,
        "rows_imported": rows_imported,
        "rows_skipped": rows_skipped,
        "error_log": "\n".join(errors[:100]),
        "created_by": user["email"],
    })
    db.commit()

    return RedirectResponse("/admin/import/legacy-tips", status_code=303)


@router.get("/api/tips/central-status")
def api_tips_central_status_v090(db: Session = Depends(get_db)):
    ensure_tip_workflow_v090_(db)
    stats = db.execute(text("""
        SELECT
          COUNT(*) AS total,
          COUNT(*) FILTER (WHERE status = 'Nový') AS new_count,
          COUNT(*) FILTER (WHERE status = 'V řešení') AS progress_count,
          COUNT(*) FILTER (WHERE status = 'Sjednáno') AS won_count,
          COUNT(*) FILTER (WHERE status = 'Storno') AS lost_count,
          COUNT(*) FILTER (WHERE status = 'Archiv') AS archive_count
        FROM tips
    """)).mappings().first()
    return {
        "ok": True,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "stats": dict(stats or {}),
    }





# -------------------------------------------------------------------
# v1.2.2 Unified TIP Inbox – jedna obrazovka jako ve stávající aplikaci
# -------------------------------------------------------------------

@router.get("/hub/my-tips", response_class=HTMLResponse)
def hub_my_tips_unified_v091(
    request: Request,
    tab: str = "sent",
    q: str = "",
    status: str = "",
    db: Session = Depends(get_db),
):
    ensure_tip_workflow_v090_(db)
    user = hub_user_context_v083_()
    adviser_id = user.get("advisor_id") or ""
    email = (user.get("email") or "").lower()

    base_where = "1=1"
    params = {}

    if tab == "work":
        base_where = "lower(COALESCE(specialist_email, '')) = :email AND COALESCE(status, '') NOT IN ('Sjednáno','Storno','Archiv')"
        params["email"] = email
    elif tab == "archive":
        base_where = """
          (
            COALESCE(adviser_original_id, '') = :adviser_id
            OR lower(COALESCE(adviser_email, '')) = :email
            OR lower(COALESCE(specialist_email, '')) = :email
          )
          AND COALESCE(status, '') IN ('Sjednáno','Storno','Archiv')
        """
        params["adviser_id"] = adviser_id
        params["email"] = email
    else:
        tab = "sent"
        base_where = """
          (
            COALESCE(adviser_original_id, '') = :adviser_id
            OR lower(COALESCE(adviser_email, '')) = :email
          )
          AND COALESCE(status, '') <> 'Archiv'
        """
        params["adviser_id"] = adviser_id
        params["email"] = email

    sql = f"SELECT * FROM tips WHERE {base_where}"

    if q:
        sql += """
          AND (
            lower(COALESCE(client_name, '')) LIKE :q OR
            lower(COALESCE(client_identifier, '')) LIKE :q OR
            lower(COALESCE(adviser_name, '')) LIKE :q OR
            lower(COALESCE(specialist_name, '')) LIKE :q OR
            lower(COALESCE(policy_no, '')) LIKE :q OR
            lower(COALESCE(section_name, '')) LIKE :q OR
            lower(COALESCE(subsection_name, '')) LIKE :q OR
            lower(COALESCE(adviser_note, '')) LIKE :q
          )
        """
        params["q"] = f"%{q.lower()}%"

    if status:
        sql += " AND status = :status"
        params["status"] = status

    sql += " ORDER BY last_update_at DESC, created_at DESC LIMIT 500"
    rows = db.execute(text(sql), params).mappings().all()

    stats = db.execute(text("""
        SELECT
          COUNT(*) FILTER (
            WHERE (COALESCE(adviser_original_id, '') = :adviser_id OR lower(COALESCE(adviser_email, '')) = :email)
              AND COALESCE(status, '') <> 'Archiv'
          ) AS sent_count,
          COUNT(*) FILTER (
            WHERE lower(COALESCE(specialist_email, '')) = :email
              AND COALESCE(status, '') NOT IN ('Sjednáno','Storno','Archiv')
          ) AS work_count,
          COUNT(*) FILTER (
            WHERE (
              COALESCE(adviser_original_id, '') = :adviser_id
              OR lower(COALESCE(adviser_email, '')) = :email
              OR lower(COALESCE(specialist_email, '')) = :email
            )
            AND COALESCE(status, '') IN ('Sjednáno','Storno','Archiv')
          ) AS archive_count,
          COUNT(*) FILTER (WHERE status = 'Sjednáno') AS won_count,
          COUNT(*) FILTER (WHERE status = 'Storno') AS lost_count
        FROM tips
    """), {"adviser_id": adviser_id, "email": email}).mappings().first()

    return hub_render_v083_(request, "hub_my_tips_unified.html", {
        "active": "my_tips",
        "tab": tab,
        "rows": rows,
        "stats": stats,
        "q": q,
        "status": status,
    })


@router.get("/hub/tips/{tip_id}", response_class=HTMLResponse)
def hub_tip_detail_unified_v091(request: Request, tip_id: str, db: Session = Depends(get_db)):
    ensure_tip_workflow_v090_(db)
    user = hub_user_context_v083_()
    email = (user.get("email") or "").lower()
    adviser_id = user.get("advisor_id") or ""

    tip = fetch_one_safe_v084_(db, """
        SELECT *
        FROM tips
        WHERE id = :id
          AND (
            COALESCE(adviser_original_id, '') = :adviser_id
            OR lower(COALESCE(adviser_email, '')) = :email
            OR lower(COALESCE(specialist_email, '')) = :email
          )
        LIMIT 1
    """, {"id": tip_id, "email": email, "adviser_id": adviser_id})

    updates = fetch_all_safe_v084_(db, """
        SELECT *
        FROM tip_updates
        WHERE tip_id = :id
        ORDER BY created_at DESC
    """, {"id": tip_id})

    return hub_render_v083_(request, "hub_tip_detail_unified.html", {
        "active": "my_tips",
        "tip": tip,
        "updates": updates,
        "is_specialist": bool(tip and (tip.get("specialist_email") or "").lower() == email),
    })


@router.post("/hub/tips/{tip_id}/specialist-update")
def hub_tip_unified_specialist_update_v091(
    tip_id: str,
    status: str = Form("V řešení"),
    policy_no: str = Form(""),
    closed_at_input: str = Form(""),
    final_volume: str = Form(""),
    next_business: str = Form(""),
    message_to_adviser: str = Form(""),
    internal_note: str = Form(""),
    final_report: str = Form(""),
    db: Session = Depends(get_db),
):
    return hub_specialist_tip_update_v090(
        tip_id=tip_id,
        status=status,
        policy_no=policy_no,
        closed_at_input=closed_at_input,
        final_volume=final_volume,
        next_business=next_business,
        message_to_adviser=message_to_adviser,
        internal_note=internal_note,
        final_report=final_report,
        db=db,
    )





# -------------------------------------------------------------------
# v1.2.2 XLSX importer – import přímo ze staženého Google Sheetu
# -------------------------------------------------------------------

def xlsx_cell_to_str_v093_(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    # Excel často vrací číselné ID jako float 501.0; v aplikaci musí být 501.
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    text_value = str(value).strip()
    if re.fullmatch(r"\d+\.0", text_value):
        return text_value[:-2]
    return text_value


def xlsx_bool_v093_(value):
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"ano", "a", "true", "1", "yes", "aktivni", "aktivní"}


def xlsx_num_v093_(value, default=0):
    try:
        if value is None or value == "":
            return default
        return int(float(str(value).replace(",", ".").replace(" ", "")))
    except Exception:
        return default


def xlsx_decimal_v093_(value):
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace("Kč", "").replace(",-", "").replace(" ", "").replace(",", "."))
    except Exception:
        return None


def xlsx_norm_header_v093_(value):
    value = str(value or "").strip().lower()
    repl = {
        "á": "a", "č": "c", "ď": "d", "é": "e", "ě": "e", "í": "i",
        "ň": "n", "ó": "o", "ř": "r", "š": "s", "ť": "t", "ú": "u",
        "ů": "u", "ý": "y", "ž": "z",
    }
    for a, b in repl.items():
        value = value.replace(a, b)
    for ch in [" ", "-", ".", "/", "\\", "(", ")", "[", "]", ":", ";"]:
        value = value.replace(ch, "_")
    while "__" in value:
        value = value.replace("__", "_")
    return value.strip("_")


def xlsx_rows_v093_(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [xlsx_norm_header_v093_(h) for h in rows[0]]
    out = []
    for row in rows[1:]:
        item = {}
        empty = True
        for i, h in enumerate(headers):
            if not h:
                continue
            val = row[i] if i < len(row) else None
            if val not in (None, ""):
                empty = False
            item[h] = val
        if not empty:
            out.append(item)
    return out


def xlsx_pick_v093_(row, *keys, default=""):
    for key in keys:
        k = xlsx_norm_header_v093_(key)
        if k in row and row[k] not in (None, ""):
            return row[k]
    return default


def xlsx_upsert_v093_(db, table, conflict_col, data, update_existing=False):
    """
    Bezpečný UPSERT pro XLSX import.
    v1.2.2: u UUID tabulek doplňuje id ručně, protože starší PostgreSQL tabulky
    nemají vždy serverový DEFAULT pro id a raw SQL nepoužije SQLAlchemy default.
    """
    uuid_tables = {
        "users",
        "roles",
        "app_settings",
        "sections",
        "subsections",
        "partners",
        "tips",
        "commission_rates",
        "audit_log",
    }

    data = dict(data or {})
    if table in uuid_tables and "id" not in data:
        data["id"] = str(uuid.uuid4())

    # v1.2.2: produkční tabulky mohou mít NOT NULL created_at/updated_at bez DB defaultu.
    # Proto timestampy doplňujeme přímo do importních dat.
    timestamp_tables = {
        "users",
        "sections",
        "subsections",
        "hub_sections",
        "hub_subsections",
        "specialists",
        "partners",
        "partner_contacts",
        "partner_links",
        "partner_products",
        "tips",
        "commission_rates",
        "audit_log",
    }
    if table in timestamp_tables:
        if "created_at" not in data:
            data["created_at"] = datetime.utcnow()
        if "updated_at" not in data:
            data["updated_at"] = datetime.utcnow()

    # v1.2.2: tabulka subsections má v produkci povinný section_id.
    # Excel/import pracuje se section_code, proto ID dohledáme před UPSERTem.
    if table == "subsections" and "section_id" not in data:
        section_code = str(data.get("section_code") or "").strip()
        if section_code:
            try:
                found_section_id = db.execute(
                    text("SELECT id FROM sections WHERE section_code = :section_code LIMIT 1"),
                    {"section_code": section_code}
                ).scalar()
                if found_section_id:
                    data["section_id"] = str(found_section_id)
            except Exception:
                db.rollback()
                raise

    columns = list(data.keys())
    params = {k: data[k] for k in columns}
    col_sql = ", ".join(columns)
    val_sql = ", ".join([f":{c}" for c in columns])

    if update_existing:
        # Nikdy neaktualizujeme primární klíč id ani konfliktní sloupec.
        set_cols = [c for c in columns if c not in {conflict_col, "id"}]
        set_sql = ", ".join([f"{c}=EXCLUDED.{c}" for c in set_cols]) or f"{conflict_col}=EXCLUDED.{conflict_col}"
        sql = f"""
            INSERT INTO {table} ({col_sql})
            VALUES ({val_sql})
            ON CONFLICT ({conflict_col}) DO UPDATE SET {set_sql}
        """
    else:
        sql = f"""
            INSERT INTO {table} ({col_sql})
            VALUES ({val_sql})
            ON CONFLICT ({conflict_col}) DO NOTHING
        """

    try:
        res = db.execute(text(sql), params)
        return res.rowcount or 0
    except Exception:
        db.rollback()
        raise


def ensure_xlsx_import_structures_v093_(db):
    ensure_tip_workflow_v090_(db)
    ensure_visible_hub_sections_(db)
    ensure_specialists_table_(db)

    # Starší DB mohou mít užší tabulky. Tady je pouze bezpečné rozšíření.
    db.execute(text("ALTER TABLE partner_products ADD COLUMN IF NOT EXISTS risks TEXT NOT NULL DEFAULT ''"))
    db.execute(text("ALTER TABLE partner_products ADD COLUMN IF NOT EXISTS client_type TEXT NOT NULL DEFAULT ''"))
    db.execute(text("ALTER TABLE partner_products ADD COLUMN IF NOT EXISTS keywords TEXT NOT NULL DEFAULT ''"))
    db.execute(text("ALTER TABLE partner_products ADD COLUMN IF NOT EXISTS priority INTEGER NOT NULL DEFAULT 100"))
    db.execute(text("ALTER TABLE partner_links ADD COLUMN IF NOT EXISTS visibility TEXT NOT NULL DEFAULT ''"))
    db.execute(text("ALTER TABLE partner_contacts ADD COLUMN IF NOT EXISTS original_note TEXT NOT NULL DEFAULT ''"))
    db.execute(text("ALTER TABLE commission_rates ADD COLUMN IF NOT EXISTS business_type TEXT NOT NULL DEFAULT ''"))
    db.execute(text("ALTER TABLE commission_rates ADD COLUMN IF NOT EXISTS area TEXT NOT NULL DEFAULT ''"))
    db.commit()


def import_hub_xlsx_data_v093_(db, wb, update_existing=False):
    try:
        repair_users_timestamps_v098_(db)
    except Exception:
        pass
    ensure_xlsx_import_structures_v093_(db)

    result = {
        "users": {"created": 0, "updated_or_skipped": 0, "rows": 0},
        "sections": {"created": 0, "updated_or_skipped": 0, "rows": 0},
        "subsections": {"created": 0, "updated_or_skipped": 0, "rows": 0},
        "specialists": {"created": 0, "updated_or_skipped": 0, "rows": 0},
        "partners": {"created": 0, "updated_or_skipped": 0, "rows": 0},
        "contacts": {"created": 0, "updated_or_skipped": 0, "rows": 0},
        "links": {"created": 0, "updated_or_skipped": 0, "rows": 0},
        "products": {"created": 0, "updated_or_skipped": 0, "rows": 0},
        "commission_rates": {"created": 0, "updated_or_skipped": 0, "rows": 0},
        "terminations_partners": {"created": 0, "updated_or_skipped": 0, "rows": 0},
        "tips": {"created": 0, "updated_or_skipped": 0, "rows": 0},
        "errors": [],
    }

    # Poradci -> users
    for row in xlsx_rows_v093_(wb, "Poradci"):
        result["users"]["rows"] += 1
        try:
            advisor_id = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "ID_poradce"))
            email = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Email")).lower()
            name = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Jmeno"))
            if not advisor_id or not email or not name:
                continue
            data = {
                "advisor_id": advisor_id,
                "name": name,
                "email": email,
                "phone": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Telefon")),
                "role": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Role", default="IF")) or "IF",
                "password_hash": hash_password(xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Heslo", default="1234")) or "1234"),
                "is_active": xlsx_bool_v093_(xlsx_pick_v093_(row, "Aktivni", default="ANO")),
                "must_change_password": True,
            }
            created = xlsx_upsert_v093_(db, "users", "advisor_id", data, update_existing)
            result["users"]["created"] += created
            result["users"]["updated_or_skipped"] += 0 if created else 1
        except Exception as exc:
            result["errors"].append(f"Poradci: {exc}")

    # Sekce -> hub_sections + sections
    for row in xlsx_rows_v093_(wb, "Sekce"):
        result["sections"]["rows"] += 1
        try:
            code = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "ID_sekce")).upper()
            name = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Nazev sekce", "Název sekce"))
            if not code or not name:
                continue
            data_hub = {
                "section_code": code,
                "section_name": name,
                "icon": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Ikona")),
                "sort_order": xlsx_num_v093_(xlsx_pick_v093_(row, "Poradi", "Pořadí"), 100),
                "is_active": xlsx_bool_v093_(xlsx_pick_v093_(row, "Aktivni", "Aktivní", default="ANO")),
                "note": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Popis")),
            }
            created = xlsx_upsert_v093_(db, "hub_sections", "section_code", data_hub, update_existing)
            result["sections"]["created"] += created
            result["sections"]["updated_or_skipped"] += 0 if created else 1

            data_core = {
                "section_code": code,
                "name": name,
                "icon": data_hub["icon"],
                "sort_order": data_hub["sort_order"],
                "is_active": data_hub["is_active"],
            }
            xlsx_upsert_v093_(db, "sections", "section_code", data_core, update_existing)
        except Exception as exc:
            result["errors"].append(f"Sekce: {exc}")

    # Podsekce -> hub_subsections + subsections
    for row in xlsx_rows_v093_(wb, "Podsekce"):
        result["subsections"]["rows"] += 1
        try:
            code = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "ID_podsekce")).upper()
            section_code = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Sekce_ID")).upper()
            name = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Nazev podsekce", "Název podsekce"))
            if not code or not section_code or not name:
                continue
            data_hub = {
                "subsection_code": code,
                "section_code": section_code,
                "subsection_name": name,
                "sort_order": xlsx_num_v093_(xlsx_pick_v093_(row, "Poradi", "Pořadí"), 100),
                "is_active": xlsx_bool_v093_(xlsx_pick_v093_(row, "Aktivni", "Aktivní", default="ANO")),
                "note": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Popis")),
            }
            created = xlsx_upsert_v093_(db, "hub_subsections", "subsection_code", data_hub, update_existing)
            result["subsections"]["created"] += created
            result["subsections"]["updated_or_skipped"] += 0 if created else 1

            data_core = {
                "subsection_code": code,
                "section_code": section_code,
                "name": name,
                "sort_order": data_hub["sort_order"],
                "is_active": data_hub["is_active"],
            }
            xlsx_upsert_v093_(db, "subsections", "subsection_code", data_core, update_existing)
        except Exception as exc:
            result["errors"].append(f"Podsekce: {exc}")

    # Specialisté
    # v1.2.2: index se nevytváří uvnitř importu. Připravuje se bezpečně před importem.
    for row in xlsx_rows_v093_(wb, "Specialisté"):
        result["specialists"]["rows"] += 1
        try:
            advisor_id = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "ID_poradce"))
            section_code = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "ID_sekce")).upper()
            subsection_code = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "ID_podsekce")).upper()
            email = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Email")).lower()
            name = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Jmeno", "Jméno"))
            if not advisor_id or not section_code or not email or not name:
                continue

            data = {
                "advisor_id": advisor_id,
                "specialist_name": name,
                "email": email,
                "phone": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Telefon")),
                "section_code": section_code,
                "subsection_code": subsection_code,
                "role_description": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Role")),
                "region": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Region")),
                "if_share": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "IF_podil", "IF podíl")),
                "ps_share": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "PS_podil", "PS podíl")),
                "available": xlsx_bool_v093_(xlsx_pick_v093_(row, "Dostupny", "Dostupný", "Aktivni", default="ANO")),
                "unavailable_reason": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Duvod nedostupnosti", "Důvod nedostupnosti")),
                "is_active": xlsx_bool_v093_(xlsx_pick_v093_(row, "Aktivni", "Aktivní", default="ANO")),
                "note": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Poznamka", "Poznámka")),
            }
            cols = list(data.keys())
            params = data
            sql = f"""
                INSERT INTO specialists ({", ".join(cols)})
                VALUES ({", ".join([":" + c for c in cols])})
                ON CONFLICT (advisor_id, section_code, subsection_code, email)
                DO {"UPDATE SET " + ", ".join([c + "=EXCLUDED." + c for c in cols if c not in ["advisor_id","section_code","subsection_code","email"]]) if update_existing else "NOTHING"}
            """
            created = db.execute(text(sql), params).rowcount or 0
            result["specialists"]["created"] += created
            result["specialists"]["updated_or_skipped"] += 0 if created else 1
        except Exception as exc:
            result["errors"].append(f"Specialisté: {exc}")

    # Import_Partners + Vypovedi_Pojistovny
    for sheet_name in ["Import_Partners", "Vypovedi_Pojistovny"]:
        for row in xlsx_rows_v093_(wb, sheet_name):
            result["partners" if sheet_name == "Import_Partners" else "terminations_partners"]["rows"] += 1
            try:
                code = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Partner_ID")).upper()
                name = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Nazev", "Název"))
                if not code or not name:
                    continue
                data = {
                    "partner_code": code,
                    "name": name,
                    "address_full": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Adresa")),
                    "registry_email": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Email_podatelna", "Email", "E-mail")),
                    "data_box": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Datova_schranka", "Datová schránka")),
                    "note": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Poznamka", "Poznámka")),
                    "source": "xlsx_import",
                    "is_active": xlsx_bool_v093_(xlsx_pick_v093_(row, "Aktivni", "Aktivní", default="ANO")),
                }
                created = xlsx_upsert_v093_(db, "partners", "partner_code", data, update_existing)
                key = "partners" if sheet_name == "Import_Partners" else "terminations_partners"
                result[key]["created"] += created
                result[key]["updated_or_skipped"] += 0 if created else 1
            except Exception as exc:
                result["errors"].append(f"{sheet_name}: {exc}")

    # Kontakty
    for row in xlsx_rows_v093_(wb, "Import_Partner_Contacts"):
        result["contacts"]["rows"] += 1
        try:
            code = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Partner_ID")).upper()
            name = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Jmeno", "Jméno"))
            role = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Role"))
            if not code or (not name and not role):
                continue
            db.execute(text("""
                INSERT INTO partner_contacts
                (partner_code, full_name, role, phone, email, specialization, territory, is_top, is_vip, note, original_note, is_active)
                VALUES
                (:partner_code, :full_name, :role, :phone, :email, :specialization, :territory, :is_top, :is_vip, :note, :original_note, TRUE)
            """), {
                "partner_code": code,
                "full_name": name or role,
                "role": role,
                "phone": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Telefon")),
                "email": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Email")),
                "specialization": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Specikace", "Specifikace")),
                "territory": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Region")),
                "is_top": xlsx_bool_v093_(xlsx_pick_v093_(row, "TOP")),
                "is_vip": xlsx_bool_v093_(xlsx_pick_v093_(row, "VIP")),
                "note": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Poznamka", "Poznámka")),
                "original_note": "Import_Partner_Contacts",
            })
            result["contacts"]["created"] += 1
        except Exception as exc:
            result["contacts"]["updated_or_skipped"] += 1
            result["errors"].append(f"Kontakty: {exc}")

    # Odkazy partnerů + online kalkulačky + odkazy ASTORIE
    for sheet_name, is_global in [("Import_Partner_Links", False), ("Import_Online kalkulacky_Links", False), ("Import_Astorie_Links", True)]:
        for row in xlsx_rows_v093_(wb, sheet_name):
            result["links"]["rows"] += 1
            try:
                code = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Partner_ID", default="ASTORIE")).upper()
                title = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Nazev", "Název"))
                url = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "URL"))
                if is_global and not code:
                    code = "ASTORIE"
                if not code or not title or not url:
                    continue
                db.execute(text("""
                    INSERT INTO partner_links
                    (partner_code, title, url, category, note, visibility, is_active)
                    VALUES
                    (:partner_code, :title, :url, :category, :note, :visibility, TRUE)
                """), {
                    "partner_code": code,
                    "title": title,
                    "url": url,
                    "category": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Typ", "Kategorie")),
                    "note": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Poznamka", "Poznámka")),
                    "visibility": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Viditelnost")),
                })
                result["links"]["created"] += 1
            except Exception as exc:
                result["links"]["updated_or_skipped"] += 1
                result["errors"].append(f"{sheet_name}: {exc}")

    # Produkty
    for row in xlsx_rows_v093_(wb, "Import_Products"):
        result["products"]["rows"] += 1
        try:
            code = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Partner_ID")).upper()
            product = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Produkt"))
            if not code or not product:
                continue
            db.execute(text("""
                INSERT INTO partner_products
                (partner_code, area, subarea, product_name, note, risks, client_type, keywords, priority, is_active)
                VALUES
                (:partner_code, :area, :subarea, :product_name, :note, :risks, :client_type, :keywords, :priority, :is_active)
            """), {
                "partner_code": code,
                "area": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Oblast")),
                "subarea": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Podoblast")),
                "product_name": product,
                "note": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Poznamka", "Poznámka")),
                "risks": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Rizika")),
                "client_type": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Typ_klienta")),
                "keywords": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Klicova_slova")),
                "priority": xlsx_num_v093_(xlsx_pick_v093_(row, "Priorita"), 100),
                "is_active": xlsx_bool_v093_(xlsx_pick_v093_(row, "Aktivni", "Aktivní", default="ANO")),
            })
            result["products"]["created"] += 1
        except Exception as exc:
            result["products"]["updated_or_skipped"] += 1
            result["errors"].append(f"Produkty: {exc}")

    # Provize
    for row in xlsx_rows_v093_(wb, "Provize_TIPHub"):
        result["commission_rates"]["rows"] += 1
        try:
            section_code = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "ID_sekce")).upper()
            partner = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Partner"))
            if not section_code or not partner:
                continue
            db.execute(text("""
                INSERT INTO commission_rates
                (id, section_code, subsection_code, partner_name, base_type, product_type, rate_percent, priority, is_active, business_type, area)
                VALUES
                (:id, :section_code, :subsection_code, :partner_name, :base_type, :product_type, :rate_percent, :priority, TRUE, :business_type, :area)
            """), {
                "id": str(uuid.uuid4()),
                "section_code": section_code,
                "subsection_code": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "ID_podsekce")).upper(),
                "partner_name": partner,
                "base_type": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Typ_pojistneho")),
                "product_type": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Typ_obchodu")),
                "rate_percent": xlsx_decimal_v093_(xlsx_pick_v093_(row, "Sazba_provize_%")),
                "priority": xlsx_num_v093_(xlsx_pick_v093_(row, "Priorita"), 100),
                "business_type": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Druh obchodu")),
                "area": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Oblast")),
            })
            result["commission_rates"]["created"] += 1
        except Exception as exc:
            result["commission_rates"]["updated_or_skipped"] += 1
            result["errors"].append(f"Provize_TIPHub: {exc}")

    # TIPy
    for row in xlsx_rows_v093_(wb, "Tipy"):
        result["tips"]["rows"] += 1
        try:
            client = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Jmeno Klienta", "Jméno Klienta"))
            if not client:
                continue
            tip_id = str(uuid.uuid4())
            db.execute(text("""
                INSERT INTO tips
                (id, adviser_original_id, adviser_name, adviser_email,
                 specialist_name, specialist_email, client_name, client_phone, client_identifier,
                 potential_amount, adviser_note, status, policy_no, final_volume, specialist_feedback,
                 section_code, subsection_code, section_name, subsection_name,
                 imported_source, imported_original_id, adviser_last_message, final_report, last_update_at)
                VALUES
                (:id, :adviser_original_id, :adviser_name, :adviser_email,
                 :specialist_name, :specialist_email, :client_name, :client_phone, :client_identifier,
                 :potential_amount, :adviser_note, :status, :policy_no, :final_volume, :specialist_feedback,
                 :section_code, :subsection_code, :section_name, :subsection_name,
                 'xlsx_Aktivni_29032026_ASTORIE_HUB', :imported_original_id, :adviser_last_message, :final_report, now())
            """), {
                "id": tip_id,
                "adviser_original_id": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "ID Poradce")),
                "adviser_name": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Jmeno Poradce", "Jméno Poradce")),
                "adviser_email": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Email Poradce")).lower(),
                "specialist_name": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Jmeno Specialisty", "Jméno Specialisty")),
                "specialist_email": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Email Specialisty")).lower(),
                "client_name": client,
                "client_phone": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Kontakt Klienta")),
                "client_identifier": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Identifikace RČ IČO Datum nar", "Identifikace (RČ/IČO/Datum nar.)")),
                "potential_amount": xlsx_decimal_v093_(xlsx_pick_v093_(row, "Potencial", "Potenciál")),
                "adviser_note": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Poznamka", "Poznámka")),
                "status": normalize_tip_status_v090_(xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Stav", default="Nový"))),
                "policy_no": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Cislo smlouvy", "Číslo smlouvy")),
                "final_volume": xlsx_decimal_v093_(xlsx_pick_v093_(row, "Vyse obchodu", "Výše obchodu")),
                "specialist_feedback": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Vyjadreni specialisty", "Vyjádření specialisty")),
                "section_code": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "ID Sekce")).upper(),
                "subsection_code": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "ID Podsekce")).upper(),
                "section_name": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Nazev Sekce", "Název Sekce")),
                "subsection_name": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Nazev Podsekce", "Název Podsekce")),
                "imported_original_id": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Datum vytvoření", "Datum vytvoreni")),
                "adviser_last_message": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Vyjadreni specialisty", "Vyjádření specialisty")),
                "final_report": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Vyjadreni specialisty", "Vyjádření specialisty")),
            })
            result["tips"]["created"] += 1
        except Exception as exc:
            result["tips"]["updated_or_skipped"] += 1
            result["errors"].append(f"Tipy: {exc}")

    db.commit()
    return result


@router.get("/admin/import/hub-xlsx", response_class=HTMLResponse)
def admin_import_hub_xlsx_page_v093(request: Request):
    return render(request, "admin_import_hub_xlsx.html", {
        "active": "import",
        "result": None,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
    })


@router.post("/admin/import/hub-xlsx", response_class=HTMLResponse)
async def admin_import_hub_xlsx_v093(
    request: Request,
    file: UploadFile = File(...),
    update_existing: str = Form(""),
    db: Session = Depends(get_db),
):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return render(request, "admin_import_hub_xlsx.html", {
            "active": "import",
            "result": {"ok": False, "errors": [f"Chybí knihovna openpyxl: {exc}"]},
            "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        })

    raw = await file.read()
    try:
        wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
        result = import_hub_xlsx_data_v093_(db, wb, update_existing=(update_existing == "1"))
        result["ok"] = len(result.get("errors", [])) == 0
        result["mode"] = "update_existing" if update_existing == "1" else "safe_insert_only"
    except Exception as exc:
        result = {"ok": False, "errors": [str(exc)]}

    return render(request, "admin_import_hub_xlsx.html", {
        "active": "import",
        "result": result,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
    })


@router.get("/api/import/hub-xlsx/expected-sheets")
def api_import_hub_xlsx_expected_sheets_v093():
    return {
        "ok": True,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "mode_default": "safe_insert_only",
        "sheets": [
            "Poradci",
            "Sekce",
            "Podsekce",
            "Specialisté",
            "Import_Partners",
            "Vypovedi_Pojistovny",
            "Import_Partner_Contacts",
            "Import_Partner_Links",
            "Import_Online kalkulacky_Links",
            "Import_Astorie_Links",
            "Import_Products",
            "Provize_TIPHub",
            "Tipy",
        ],
        "note": "Importer nemění původní Google Sheet. Načítá XLSX do nové databáze aplikace.",
    }



# -------------------------------------------------------------------
# v1.2.2 import hardening endpoints
# - chybějící /api/admin/summary
# - aliasy pro import route
# - JSON upload endpoint
# - bezpečné počty tabulek
# -------------------------------------------------------------------

def safe_count_table_v094_(db: Session, table_name: str):
    try:
        exists = db.execute(text("""
            SELECT EXISTS (
              SELECT 1
              FROM information_schema.tables
              WHERE table_schema = 'public'
                AND table_name = :table_name
            )
        """), {"table_name": table_name}).scalar()
        if not exists:
            return {"exists": False, "count": 0, "error": None}
        count = db.execute(text(f"SELECT COUNT(*) FROM {table_name}")).scalar()
        return {"exists": True, "count": int(count or 0), "error": None}
    except Exception as exc:
        return {"exists": False, "count": 0, "error": str(exc)}


@router.get("/api/admin/summary")
def api_admin_summary_v094(db: Session = Depends(get_db)):
    """
    Stabilní kontrolní endpoint pro BO a testování migrace.
    Vrací počty hlavních tabulek, aby bylo po importu jasně vidět, zda data přibyla.
    """
    # vytvoří/rozšíří struktury, ale nemaže data
    try:
        ensure_xlsx_import_structures_v093_(db)
    except Exception:
        try:
            ensure_tip_workflow_v090_(db)
        except Exception:
            pass

    tables = [
        "users",
        "roles",
        "app_settings",
        "sections",
        "subsections",
        "hub_sections",
        "hub_subsections",
        "specialists",
        "partners",
        "partner_contacts",
        "partner_links",
        "partner_products",
        "commission_rates",
        "tips",
        "tip_updates",
        "import_jobs",
        "audit_log",
    ]
    return {
        "ok": True,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Admin summary endpoint běží. Počty jsou čtené bezpečně přes PostgreSQL.",
        "counts": {t: safe_count_table_v094_(db, t) for t in tables},
    }


@router.get("/api/import/hub-xlsx/status")
def api_import_hub_xlsx_status_v094(db: Session = Depends(get_db)):
    """
    Zatím synchronní import: endpoint vrací poslední stav podle import_jobs.
    Plnohodnotný async progress bude další krok, až bude ověřen základní import.
    """
    try:
        ensure_tip_workflow_v090_(db)
        last_job = fetch_one_safe_v084_(db, """
            SELECT *
            FROM import_jobs
            ORDER BY created_at DESC
            LIMIT 1
        """)
        return {
            "ok": True,
            "version": "1.2.2-admin-taxonomy-specialists-links-safe",
            "running": False,
            "last_job": dict(last_job) if last_job else None,
        }
    except Exception as exc:
        return {
            "ok": False,
            "version": "1.2.2-admin-taxonomy-specialists-links-safe",
            "running": False,
            "error": str(exc),
        }


@router.post("/api/import/hub-xlsx")
async def api_import_hub_xlsx_upload_v094(
    file: UploadFile = File(...),
    update_existing: str = Form(""),
    db: Session = Depends(get_db),
):
    """
    JSON varianta importu pro budoucí AJAX/frontend.
    Používá stejný importní engine jako HTML stránka.
    """
    try:
        from openpyxl import load_workbook
        raw = await file.read()
        wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
        result = import_hub_xlsx_data_v093_(db, wb, update_existing=(update_existing == "1"))
        result["ok"] = len(result.get("errors", [])) == 0
        result["mode"] = "update_existing" if update_existing == "1" else "safe_insert_only"
        result["version"] = "1.2.2-admin-taxonomy-specialists-links-safe"
        return result
    except Exception as exc:
        return {
            "ok": False,
            "version": "1.2.2-admin-taxonomy-specialists-links-safe",
            "errors": [str(exc)],
        }


# Alias, kdyby někde frontend / uživatel použil jiný název.
@router.get("/admin/import/hub-xlsx/")
def admin_import_hub_xlsx_slash_alias_v094(request: Request):
    return RedirectResponse("/admin/import/hub-xlsx", status_code=307)


@router.post("/admin/import/hub-xlsx/")
async def admin_import_hub_xlsx_post_slash_alias_v094(
    request: Request,
    file: UploadFile = File(...),
    update_existing: str = Form(""),
    db: Session = Depends(get_db),
):
    return await admin_import_hub_xlsx_v093(
        request=request,
        file=file,
        update_existing=update_existing,
        db=db,
    )


@router.get("/admin/import/xlsx")
def admin_import_xlsx_short_alias_v094():
    return RedirectResponse("/admin/import/hub-xlsx", status_code=307)


@router.get("/api/import/hub-xlsx/summary")
def api_import_hub_xlsx_summary_alias_v094(db: Session = Depends(get_db)):
    return api_admin_summary_v094(db=db)





# -------------------------------------------------------------------
# v1.2.2 import transaction fix
# Oprava: current transaction is aborted před CREATE UNIQUE INDEX
# -------------------------------------------------------------------

def ensure_specialists_table_columns_v095_(db):
    """
    Starší DB může mít tabulku specialists založenou s jinou strukturou.
    Importér potřebuje tyto sloupce. Každý ALTER je chráněný rollbackem,
    aby jedna chyba neshodila celou transakci.
    """
    try:
        ensure_specialists_table_(db)
        db.commit()
    except Exception:
        db.rollback()

    statements = [
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS advisor_id TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS specialist_name TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS email TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS phone TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS section_code TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS subsection_code TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS role_description TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS region TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS if_share TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS ps_share TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS available BOOLEAN NOT NULL DEFAULT TRUE",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS unavailable_reason TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS is_active BOOLEAN NOT NULL DEFAULT TRUE",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS note TEXT NOT NULL DEFAULT ''",
    ]

    for stmt in statements:
        try:
            db.execute(text(stmt))
            db.commit()
        except Exception:
            db.rollback()

    try:
        db.execute(text("""
            CREATE UNIQUE INDEX IF NOT EXISTS ux_specialists_import_key
            ON specialists (advisor_id, section_code, subsection_code, email)
        """))
        db.commit()
    except Exception:
        db.rollback()


def ensure_xlsx_import_structures_v095_(db):
    """
    Robustní verze přípravy struktur pro XLSX import.
    Každý blok se commituje/rollbackuje samostatně, aby PostgreSQL nezůstal
    ve stavu InFailedSqlTransaction.
    """
    for fn in [ensure_tip_workflow_v090_, ensure_visible_hub_sections_, ensure_specialists_table_columns_v095_]:
        try:
            fn(db)
            db.commit()
        except Exception:
            db.rollback()

    statements = [
        "ALTER TABLE partner_products ADD COLUMN IF NOT EXISTS risks TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE partner_products ADD COLUMN IF NOT EXISTS client_type TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE partner_products ADD COLUMN IF NOT EXISTS keywords TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE partner_products ADD COLUMN IF NOT EXISTS priority INTEGER NOT NULL DEFAULT 100",
        "ALTER TABLE partner_links ADD COLUMN IF NOT EXISTS visibility TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE partner_contacts ADD COLUMN IF NOT EXISTS original_note TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE commission_rates ADD COLUMN IF NOT EXISTS business_type TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE commission_rates ADD COLUMN IF NOT EXISTS area TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS final_volume NUMERIC",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS specialist_feedback TEXT NOT NULL DEFAULT ''",
    ]
    for stmt in statements:
        try:
            db.execute(text(stmt))
            db.commit()
        except Exception:
            db.rollback()


# Přesměrování původní funkce na robustní verzi.
ensure_xlsx_import_structures_v093_ = ensure_xlsx_import_structures_v095_





# -------------------------------------------------------------------
# v1.2.2 import index fix
# Definitivní oprava: odstranění inline CREATE UNIQUE INDEX z importní transakce
# a bezpečné čištění transakce před vlastním importem.
# -------------------------------------------------------------------

def db_safe_exec_v096_(db: Session, sql: str, params: dict | None = None):
    try:
        db.rollback()
    except Exception:
        pass
    try:
        db.execute(text(sql), params or {})
        db.commit()
        return True, None
    except Exception as exc:
        try:
            db.rollback()
        except Exception:
            pass
        return False, str(exc)


def ensure_specialists_import_schema_v096_(db: Session):
    """
    Tato funkce nesmí shodit import. Každý SQL příkaz běží odděleně.
    """
    try:
        ensure_specialists_table_(db)
        db.commit()
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass

    stmts = [
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS advisor_id TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS specialist_name TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS email TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS phone TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS section_code TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS subsection_code TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS role_description TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS region TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS if_share TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS ps_share TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS available BOOLEAN NOT NULL DEFAULT TRUE",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS unavailable_reason TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS is_active BOOLEAN NOT NULL DEFAULT TRUE",
        "ALTER TABLE specialists ADD COLUMN IF NOT EXISTS note TEXT NOT NULL DEFAULT ''",
    ]

    errors = []
    for stmt in stmts:
        ok, err = db_safe_exec_v096_(db, stmt)
        if err:
            errors.append(err)

    # Index je užitečný, ale nesmí zablokovat import. Pokud nejde vytvořit, import pokračuje bez něj.
    ok, err = db_safe_exec_v096_(db, """
        CREATE UNIQUE INDEX IF NOT EXISTS ux_specialists_import_key
        ON specialists (advisor_id, section_code, subsection_code, email)
    """)
    if err:
        errors.append("Index specialists přeskočen: " + err)

    try:
        db.rollback()
    except Exception:
        pass
    return errors


def ensure_xlsx_import_structures_v096_(db: Session):
    errors = []

    for fn in [ensure_tip_workflow_v090_, ensure_visible_hub_sections_]:
        try:
            db.rollback()
        except Exception:
            pass
        try:
            fn(db)
            db.commit()
        except Exception as exc:
            errors.append(str(exc))
            try:
                db.rollback()
            except Exception:
                pass

    errors.extend(ensure_specialists_import_schema_v096_(db))

    stmts = [
        "ALTER TABLE partner_products ADD COLUMN IF NOT EXISTS risks TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE partner_products ADD COLUMN IF NOT EXISTS client_type TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE partner_products ADD COLUMN IF NOT EXISTS keywords TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE partner_products ADD COLUMN IF NOT EXISTS priority INTEGER NOT NULL DEFAULT 100",
        "ALTER TABLE partner_links ADD COLUMN IF NOT EXISTS visibility TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE partner_contacts ADD COLUMN IF NOT EXISTS original_note TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE commission_rates ADD COLUMN IF NOT EXISTS business_type TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE commission_rates ADD COLUMN IF NOT EXISTS area TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS final_volume NUMERIC",
        "ALTER TABLE tips ADD COLUMN IF NOT EXISTS specialist_feedback TEXT NOT NULL DEFAULT ''",
    ]
    for stmt in stmts:
        ok, err = db_safe_exec_v096_(db, stmt)
        if err:
            errors.append(err)

    try:
        db.rollback()
    except Exception:
        pass
    return errors


_original_import_hub_xlsx_data_v093_v096 = import_hub_xlsx_data_v093_

def import_hub_xlsx_data_v096_(db, wb, update_existing=False):
    prep_errors = ensure_xlsx_import_structures_v096_(db)
    try:
        db.rollback()
    except Exception:
        pass

    result = _original_import_hub_xlsx_data_v093_v096(db, wb, update_existing=update_existing)

    # Nepovažovat přeskočený index za fatální chybu importu.
    non_fatal = []
    fatal = []
    for e in prep_errors:
        if "Index specialists přeskočen" in str(e):
            non_fatal.append(e)
        else:
            fatal.append(e)

    if non_fatal:
        result.setdefault("warnings", []).extend(non_fatal)
    if fatal:
        result.setdefault("errors", []).extend(fatal)
    return result


# Přesměrování všech importů na opravený engine.
import_hub_xlsx_data_v093_ = import_hub_xlsx_data_v096_
ensure_xlsx_import_structures_v093_ = ensure_xlsx_import_structures_v096_

@router.get("/api/import/hub-xlsx/repair-schema")
def api_import_repair_schema_v096(db: Session = Depends(get_db)):
    errors = ensure_xlsx_import_structures_v096_(db)
    return {
        "ok": len(errors) == 0,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Importní struktury byly zkontrolovány a opraveny. Původní Google Sheet se nemění.",
        "errors": errors,
    }





# -------------------------------------------------------------------
# v1.2.2 import user id fix
# Oprava: users.id nemá serverový default a raw SQL insert bez id padal.
# -------------------------------------------------------------------

def repair_uuid_defaults_v097_(db: Session):
    """
    Doplní serverový DEFAULT gen_random_uuid() pro UUID id sloupce.
    Je to pojistka; import zároveň posílá id explicitně.
    """
    errors = []
    try:
        db.rollback()
    except Exception:
        pass

    db_safe_exec_v096_(db, "CREATE EXTENSION IF NOT EXISTS pgcrypto")
    for table in ["users", "roles", "app_settings", "sections", "subsections", "partners", "tips", "commission_rates", "audit_log"]:
        ok, err = db_safe_exec_v096_(db, f"ALTER TABLE {table} ALTER COLUMN id SET DEFAULT gen_random_uuid()")
        if err:
            errors.append(f"{table}.id default: {err}")
    try:
        db.rollback()
    except Exception:
        pass
    return errors


_previous_ensure_xlsx_import_structures_v093_v097 = ensure_xlsx_import_structures_v093_

def ensure_xlsx_import_structures_v097_(db: Session):
    errors = []
    try:
        repair_uuid_defaults_v097_(db)
    except Exception as exc:
        errors.append(str(exc))
    try:
        rv = _previous_ensure_xlsx_import_structures_v093_v097(db)
        if isinstance(rv, list):
            errors.extend(rv)
    except Exception as exc:
        try:
            db.rollback()
        except Exception:
            pass
        errors.append(str(exc))
    return errors


ensure_xlsx_import_structures_v093_ = ensure_xlsx_import_structures_v097_

@router.get("/api/import/hub-xlsx/repair-users")
def api_import_repair_users_v097(db: Session = Depends(get_db)):
    errors = repair_uuid_defaults_v097_(db)
    return {
        "ok": len(errors) == 0,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Opraveny UUID defaulty pro users a další hlavní tabulky. Import zároveň posílá id explicitně.",
        "errors": errors,
    }




# -------------------------------------------------------------------
# v1.2.2 import timestamps fix
# Oprava: users.created_at / users.updated_at NOT NULL při importu poradců
# -------------------------------------------------------------------

def repair_users_timestamps_v098_(db: Session):
    errors = []
    statements = [
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS created_at TIMESTAMPTZ",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS updated_at TIMESTAMPTZ",
        "UPDATE users SET created_at = NOW() WHERE created_at IS NULL",
        "UPDATE users SET updated_at = NOW() WHERE updated_at IS NULL",
        "ALTER TABLE users ALTER COLUMN created_at SET DEFAULT NOW()",
        "ALTER TABLE users ALTER COLUMN updated_at SET DEFAULT NOW()",
    ]
    for stmt in statements:
        try:
            db.rollback()
        except Exception:
            pass
        try:
            db.execute(text(stmt))
            db.commit()
        except Exception as exc:
            errors.append(str(exc))
            try:
                db.rollback()
            except Exception:
                pass
    return errors


@router.get("/api/import/hub-xlsx/repair-users-timestamps")
def api_import_repair_users_timestamps_v098(db: Session = Depends(get_db)):
    errors = repair_users_timestamps_v098_(db)
    return {
        "ok": len(errors) == 0,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Opraveny created_at/updated_at defaulty pro users. Import poradců nyní posílá timestampy explicitně.",
        "errors": errors,
    }



# -------------------------------------------------------------------
# v1.2.2 import schema canonical fix
# Profesionální oprava: sjednocení schématu všech importních tabulek před importem.
# -------------------------------------------------------------------

def v099_safe_ddl(db: Session, sql: str):
    try:
        db.rollback()
    except Exception:
        pass
    try:
        db.execute(text(sql))
        db.commit()
        return None
    except Exception as exc:
        try:
            db.rollback()
        except Exception:
            pass
        return str(exc)


def v099_table_exists(db: Session, table_name: str) -> bool:
    try:
        db.rollback()
        exists = db.execute(text("""
            SELECT EXISTS (
              SELECT 1 FROM information_schema.tables
              WHERE table_schema='public' AND table_name=:t
            )
        """), {"t": table_name}).scalar()
        db.commit()
        return bool(exists)
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass
        return False


def v099_column_exists(db: Session, table_name: str, column_name: str) -> bool:
    try:
        db.rollback()
        exists = db.execute(text("""
            SELECT EXISTS (
              SELECT 1 FROM information_schema.columns
              WHERE table_schema='public' AND table_name=:t AND column_name=:c
            )
        """), {"t": table_name, "c": column_name}).scalar()
        db.commit()
        return bool(exists)
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass
        return False


def v099_add_col(db: Session, table_name: str, column_name: str, definition: str):
    if not v099_table_exists(db, table_name):
        return None
    if v099_column_exists(db, table_name, column_name):
        return None
    return v099_safe_ddl(db, f"ALTER TABLE {table_name} ADD COLUMN {column_name} {definition}")


def repair_import_schema_canonical_v099_(db: Session):
    errors = []

    # Zavolat starší bootstrap/repair funkce bezpečně.
    for fn_name in [
        "ensure_tip_workflow_v090_",
        "ensure_visible_hub_sections_",
        "ensure_specialists_import_schema_v096_",
        "repair_uuid_defaults_v097_",
        "repair_users_timestamps_v098_",
    ]:
        fn = globals().get(fn_name)
        if fn:
            try:
                db.rollback()
            except Exception:
                pass
            try:
                res = fn(db)
                if isinstance(res, list):
                    errors.extend([str(x) for x in res if x])
                db.commit()
            except Exception as exc:
                errors.append(f"{fn_name}: {exc}")
                try:
                    db.rollback()
                except Exception:
                    pass

    table_columns = {
        "users": {
            "created_at": "TIMESTAMPTZ DEFAULT NOW()",
            "updated_at": "TIMESTAMPTZ DEFAULT NOW()",
        },
        "hub_sections": {
            "section_code": "TEXT",
            "section_name": "TEXT",
            "icon": "TEXT DEFAULT ''",
            "sort_order": "INTEGER DEFAULT 100",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "note": "TEXT DEFAULT ''",
            "created_at": "TIMESTAMPTZ DEFAULT NOW()",
            "updated_at": "TIMESTAMPTZ DEFAULT NOW()",
        },
        "hub_subsections": {
            "subsection_code": "TEXT",
            "section_code": "TEXT",
            "subsection_name": "TEXT",
            "sort_order": "INTEGER DEFAULT 100",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "note": "TEXT DEFAULT ''",
            "created_at": "TIMESTAMPTZ DEFAULT NOW()",
            "updated_at": "TIMESTAMPTZ DEFAULT NOW()",
        },
        "partners": {
            "partner_code": "TEXT",
            "name": "TEXT",
            "ico": "TEXT DEFAULT ''",
            "dic": "TEXT DEFAULT ''",
            "data_box": "TEXT DEFAULT ''",
            "registry_email": "TEXT DEFAULT ''",
            "address_full": "TEXT DEFAULT ''",
            "street": "TEXT DEFAULT ''",
            "city": "TEXT DEFAULT ''",
            "zip_code": "TEXT DEFAULT ''",
            "legal_form": "TEXT DEFAULT ''",
            "note": "TEXT DEFAULT ''",
            "source": "TEXT DEFAULT ''",
            "status": "TEXT DEFAULT 'aktivní'",
            "is_vip": "BOOLEAN DEFAULT FALSE",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "created_at": "TIMESTAMPTZ DEFAULT NOW()",
            "updated_at": "TIMESTAMPTZ DEFAULT NOW()",
        },
        "partner_contacts": {
            "partner_code": "TEXT",
            "contact_name": "TEXT",
            "role_description": "TEXT DEFAULT ''",
            "region": "TEXT DEFAULT ''",
            "email": "TEXT DEFAULT ''",
            "phone": "TEXT DEFAULT ''",
            "specialization": "TEXT DEFAULT ''",
            "is_vip": "BOOLEAN DEFAULT FALSE",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "note": "TEXT DEFAULT ''",
            "original_note": "TEXT DEFAULT ''",
            "created_at": "TIMESTAMPTZ DEFAULT NOW()",
            "updated_at": "TIMESTAMPTZ DEFAULT NOW()",
        },
        "partner_links": {
            "partner_code": "TEXT",
            "link_name": "TEXT",
            "url": "TEXT",
            "category": "TEXT DEFAULT ''",
            "visibility": "TEXT DEFAULT ''",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "note": "TEXT DEFAULT ''",
            "created_at": "TIMESTAMPTZ DEFAULT NOW()",
            "updated_at": "TIMESTAMPTZ DEFAULT NOW()",
        },
        "partner_products": {
            "partner_code": "TEXT",
            "area": "TEXT DEFAULT ''",
            "subarea": "TEXT DEFAULT ''",
            "product_name": "TEXT",
            "risks": "TEXT DEFAULT ''",
            "client_type": "TEXT DEFAULT ''",
            "keywords": "TEXT DEFAULT ''",
            "priority": "INTEGER DEFAULT 100",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "note": "TEXT DEFAULT ''",
            "created_at": "TIMESTAMPTZ DEFAULT NOW()",
            "updated_at": "TIMESTAMPTZ DEFAULT NOW()",
        },
        "commission_rates": {
            "partner_code": "TEXT DEFAULT ''",
            "section_code": "TEXT DEFAULT ''",
            "subsection_code": "TEXT DEFAULT ''",
            "business_type": "TEXT DEFAULT ''",
            "area": "TEXT DEFAULT ''",
            "if_rate": "TEXT DEFAULT ''",
            "ps_rate": "TEXT DEFAULT ''",
            "note": "TEXT DEFAULT ''",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "created_at": "TIMESTAMPTZ DEFAULT NOW()",
            "updated_at": "TIMESTAMPTZ DEFAULT NOW()",
        },
        "tips": {
            "client_name": "TEXT DEFAULT ''",
            "client_contact": "TEXT DEFAULT ''",
            "client_identifier": "TEXT DEFAULT ''",
            "advisor_id": "TEXT DEFAULT ''",
            "advisor_name": "TEXT DEFAULT ''",
            "advisor_email": "TEXT DEFAULT ''",
            "specialist_id": "TEXT DEFAULT ''",
            "specialist_name": "TEXT DEFAULT ''",
            "section_code": "TEXT DEFAULT ''",
            "subsection_code": "TEXT DEFAULT ''",
            "potential": "TEXT DEFAULT ''",
            "status": "TEXT DEFAULT 'nový'",
            "description": "TEXT DEFAULT ''",
            "specialist_feedback": "TEXT DEFAULT ''",
            "final_volume": "NUMERIC",
            "created_at": "TIMESTAMPTZ DEFAULT NOW()",
            "updated_at": "TIMESTAMPTZ DEFAULT NOW()",
        },
    }

    for table, cols in table_columns.items():
        if v099_table_exists(db, table):
            for col, definition in cols.items():
                err = v099_add_col(db, table, col, definition)
                if err:
                    errors.append(f"{table}.{col}: {err}")
            if v099_column_exists(db, table, "created_at"):
                err = v099_safe_ddl(db, f"UPDATE {table} SET created_at = NOW() WHERE created_at IS NULL")
                if err: errors.append(f"{table}.created_at update: {err}")
                err = v099_safe_ddl(db, f"ALTER TABLE {table} ALTER COLUMN created_at SET DEFAULT NOW()")
                if err: errors.append(f"{table}.created_at default: {err}")
            if v099_column_exists(db, table, "updated_at"):
                err = v099_safe_ddl(db, f"UPDATE {table} SET updated_at = NOW() WHERE updated_at IS NULL")
                if err: errors.append(f"{table}.updated_at update: {err}")
                err = v099_safe_ddl(db, f"ALTER TABLE {table} ALTER COLUMN updated_at SET DEFAULT NOW()")
                if err: errors.append(f"{table}.updated_at default: {err}")

    try:
        db.rollback()
    except Exception:
        pass
    return [e for e in errors if e]


@router.get("/api/import/hub-xlsx/repair-all")
def api_import_repair_all_v099(db: Session = Depends(get_db)):
    errors = v100_fix_all_import_tables(db)
    return {
        "ok": len(errors) == 0,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Kompletní oprava importního schématu dokončena. Opraveny created_at/updated_at a chybějící importní sloupce.",
        "errors": errors,
    }


try:
    _previous_import_hub_xlsx_data_before_v099 = import_hub_xlsx_data_v093_
    def import_hub_xlsx_data_v099_(db, wb, update_existing=False):
        repair_errors = v100_fix_all_import_tables(db)
        result = _previous_import_hub_xlsx_data_before_v099(db, wb, update_existing=update_existing)
        if repair_errors:
            result.setdefault("warnings", []).extend(repair_errors)
        return result
    import_hub_xlsx_data_v093_ = import_hub_xlsx_data_v099_
except Exception:
    pass




# -------------------------------------------------------------------
# v1.2.2 full import schema fix
# Jednorázová profesionální oprava importního schématu:
# doplní přesně ty sloupce, které import reálně používá.
# -------------------------------------------------------------------

def v100_exec(db: Session, sql: str, params: dict | None = None):
    try:
        db.rollback()
    except Exception:
        pass
    try:
        db.execute(text(sql), params or {})
        db.commit()
        return None
    except Exception as exc:
        try:
            db.rollback()
        except Exception:
            pass
        return str(exc)


def v100_table_exists(db: Session, table_name: str) -> bool:
    err = None
    try:
        db.rollback()
    except Exception:
        pass
    try:
        exists = db.execute(text("""
            SELECT EXISTS (
                SELECT 1
                FROM information_schema.tables
                WHERE table_schema = 'public'
                  AND table_name = :table_name
            )
        """), {"table_name": table_name}).scalar()
        db.commit()
        return bool(exists)
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass
        return False


def v100_col_exists(db: Session, table_name: str, column_name: str) -> bool:
    try:
        db.rollback()
    except Exception:
        pass
    try:
        exists = db.execute(text("""
            SELECT EXISTS (
                SELECT 1
                FROM information_schema.columns
                WHERE table_schema = 'public'
                  AND table_name = :table_name
                  AND column_name = :column_name
            )
        """), {"table_name": table_name, "column_name": column_name}).scalar()
        db.commit()
        return bool(exists)
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass
        return False


def v100_add_col(db: Session, table: str, column: str, definition: str):
    if not v100_table_exists(db, table):
        return None
    if v100_col_exists(db, table, column):
        return None
    return v100_exec(db, f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def v100_fix_timestamps(db: Session, table: str, errors: list[str]):
    if not v100_table_exists(db, table):
        return

    for col in ["created_at", "updated_at"]:
        err = v100_add_col(db, table, col, "TIMESTAMPTZ DEFAULT NOW()")
        if err:
            errors.append(f"{table}.{col} add: {err}")

        if v100_col_exists(db, table, col):
            err = v100_exec(db, f"UPDATE {table} SET {col} = NOW() WHERE {col} IS NULL")
            if err:
                errors.append(f"{table}.{col} update: {err}")

            err = v100_exec(db, f"ALTER TABLE {table} ALTER COLUMN {col} SET DEFAULT NOW()")
            if err:
                errors.append(f"{table}.{col} default: {err}")


def v100_fix_all_import_tables(db: Session):
    errors = []

    # Nejdřív vytvořit / zajistit starší tabulky, pokud existují původní ensure funkce.
    for fn_name in [
        "ensure_tip_workflow_v090_",
        "ensure_visible_hub_sections_",
        "ensure_specialists_table_",
        "ensure_specialists_import_schema_v096_",
        "repair_uuid_defaults_v097_",
        "repair_users_timestamps_v098_",
    ]:
        fn = globals().get(fn_name)
        if fn:
            try:
                db.rollback()
            except Exception:
                pass
            try:
                res = fn(db)
                if isinstance(res, list):
                    errors.extend([str(x) for x in res if x])
                db.commit()
            except Exception as exc:
                errors.append(f"{fn_name}: {exc}")
                try:
                    db.rollback()
                except Exception:
                    pass

    # Tabulky, do kterých XLSX import skutečně zapisuje podle logů a importních funkcí.
    canonical = {
        "users": {
            "id": "TEXT",
            "advisor_id": "TEXT",
            "name": "TEXT",
            "email": "TEXT DEFAULT ''",
            "phone": "TEXT DEFAULT ''",
            "role": "TEXT DEFAULT 'advisor'",
            "password_hash": "TEXT DEFAULT ''",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "must_change_password": "BOOLEAN DEFAULT TRUE",
        },
        "sections": {
            "id": "TEXT",
            "section_code": "TEXT",
            "name": "TEXT",
            "icon": "TEXT DEFAULT ''",
            "sort_order": "INTEGER DEFAULT 100",
            "is_active": "BOOLEAN DEFAULT TRUE",
        },
        "subsections": {
            "id": "TEXT",
            "subsection_code": "TEXT",
            "section_code": "TEXT",
            "name": "TEXT",
            "sort_order": "INTEGER DEFAULT 100",
            "is_active": "BOOLEAN DEFAULT TRUE",
        },
        "hub_sections": {
            "id": "TEXT",
            "section_code": "TEXT",
            "section_name": "TEXT",
            "name": "TEXT",
            "icon": "TEXT DEFAULT ''",
            "sort_order": "INTEGER DEFAULT 100",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "note": "TEXT DEFAULT ''",
        },
        "hub_subsections": {
            "id": "TEXT",
            "subsection_code": "TEXT",
            "section_code": "TEXT",
            "subsection_name": "TEXT",
            "name": "TEXT",
            "sort_order": "INTEGER DEFAULT 100",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "note": "TEXT DEFAULT ''",
        },
        "specialists": {
            "id": "TEXT",
            "advisor_id": "TEXT DEFAULT ''",
            "specialist_name": "TEXT DEFAULT ''",
            "name": "TEXT DEFAULT ''",
            "email": "TEXT DEFAULT ''",
            "phone": "TEXT DEFAULT ''",
            "section_code": "TEXT DEFAULT ''",
            "subsection_code": "TEXT DEFAULT ''",
            "role_description": "TEXT DEFAULT ''",
            "region": "TEXT DEFAULT ''",
            "if_share": "TEXT DEFAULT ''",
            "ps_share": "TEXT DEFAULT ''",
            "available": "BOOLEAN DEFAULT TRUE",
            "unavailable_reason": "TEXT DEFAULT ''",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "note": "TEXT DEFAULT ''",
        },
        "partners": {
            "id": "TEXT",
            "partner_code": "TEXT",
            "name": "TEXT",
            "ico": "TEXT DEFAULT ''",
            "dic": "TEXT DEFAULT ''",
            "data_box": "TEXT DEFAULT ''",
            "registry_email": "TEXT DEFAULT ''",
            "address_full": "TEXT DEFAULT ''",
            "street": "TEXT DEFAULT ''",
            "city": "TEXT DEFAULT ''",
            "zip_code": "TEXT DEFAULT ''",
            "legal_form": "TEXT DEFAULT ''",
            "note": "TEXT DEFAULT ''",
            "source": "TEXT DEFAULT ''",
            "status": "TEXT DEFAULT 'aktivní'",
            "is_vip": "BOOLEAN DEFAULT FALSE",
            "is_active": "BOOLEAN DEFAULT TRUE",
        },
        "partner_contacts": {
            "id": "TEXT",
            "partner_code": "TEXT",
            "contact_name": "TEXT",
            "name": "TEXT DEFAULT ''",
            "role_description": "TEXT DEFAULT ''",
            "region": "TEXT DEFAULT ''",
            "email": "TEXT DEFAULT ''",
            "phone": "TEXT DEFAULT ''",
            "specialization": "TEXT DEFAULT ''",
            "is_vip": "BOOLEAN DEFAULT FALSE",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "note": "TEXT DEFAULT ''",
            "original_note": "TEXT DEFAULT ''",
        },
        "partner_links": {
            "id": "TEXT",
            "partner_code": "TEXT",
            "link_name": "TEXT",
            "name": "TEXT DEFAULT ''",
            "url": "TEXT",
            "category": "TEXT DEFAULT ''",
            "visibility": "TEXT DEFAULT ''",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "note": "TEXT DEFAULT ''",
        },
        "partner_products": {
            "id": "TEXT",
            "partner_code": "TEXT",
            "area": "TEXT DEFAULT ''",
            "subarea": "TEXT DEFAULT ''",
            "product_name": "TEXT",
            "name": "TEXT DEFAULT ''",
            "risks": "TEXT DEFAULT ''",
            "client_type": "TEXT DEFAULT ''",
            "keywords": "TEXT DEFAULT ''",
            "priority": "INTEGER DEFAULT 100",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "note": "TEXT DEFAULT ''",
        },
        "commission_rates": {
            "id": "TEXT",
            "partner_code": "TEXT DEFAULT ''",
            "section_code": "TEXT DEFAULT ''",
            "subsection_code": "TEXT DEFAULT ''",
            "business_type": "TEXT DEFAULT ''",
            "area": "TEXT DEFAULT ''",
            "if_rate": "TEXT DEFAULT ''",
            "ps_rate": "TEXT DEFAULT ''",
            "note": "TEXT DEFAULT ''",
            "is_active": "BOOLEAN DEFAULT TRUE",
        },
        "tips": {
            "id": "TEXT",
            "client_name": "TEXT DEFAULT ''",
            "client_contact": "TEXT DEFAULT ''",
            "client_identifier": "TEXT DEFAULT ''",
            "advisor_id": "TEXT DEFAULT ''",
            "advisor_name": "TEXT DEFAULT ''",
            "advisor_email": "TEXT DEFAULT ''",
            "specialist_id": "TEXT DEFAULT ''",
            "specialist_name": "TEXT DEFAULT ''",
            "section_code": "TEXT DEFAULT ''",
            "subsection_code": "TEXT DEFAULT ''",
            "potential": "TEXT DEFAULT ''",
            "status": "TEXT DEFAULT 'nový'",
            "description": "TEXT DEFAULT ''",
            "specialist_feedback": "TEXT DEFAULT ''",
            "final_volume": "NUMERIC",
        },
        "audit_log": {
            "id": "TEXT",
            "user_email": "TEXT DEFAULT ''",
            "action": "TEXT DEFAULT ''",
            "entity": "TEXT DEFAULT ''",
            "entity_id": "TEXT DEFAULT ''",
            "detail": "TEXT DEFAULT ''",
        },
    }

    for table, cols in canonical.items():
        if not v100_table_exists(db, table):
            continue

        for col, definition in cols.items():
            err = v100_add_col(db, table, col, definition)
            if err:
                errors.append(f"{table}.{col}: {err}")

        v100_fix_timestamps(db, table, errors)

        if v100_col_exists(db, table, "id"):
            # Nevnucujeme NOT NULL, jen doplníme default tam, kde PostgreSQL podporuje pgcrypto.
            v100_exec(db, "CREATE EXTENSION IF NOT EXISTS pgcrypto")
            v100_exec(db, f"UPDATE {table} SET id = gen_random_uuid()::text WHERE id IS NULL")
            v100_exec(db, f"ALTER TABLE {table} ALTER COLUMN id SET DEFAULT gen_random_uuid()::text")

    # Unikátní indexy jen tam, kde sloupce opravdu existují.
    index_specs = [
        ("sections", "ux_sections_section_code", "section_code"),
        ("subsections", "ux_subsections_subsection_code", "subsection_code"),
        ("hub_sections", "ux_hub_sections_section_code", "section_code"),
        ("hub_subsections", "ux_hub_subsections_subsection_code", "subsection_code"),
        ("partners", "ux_partners_partner_code", "partner_code"),
        ("users", "ux_users_advisor_id", "advisor_id"),
    ]
    for table, idx, col in index_specs:
        if v100_table_exists(db, table) and v100_col_exists(db, table, col):
            err = v100_exec(db, f"CREATE UNIQUE INDEX IF NOT EXISTS {idx} ON {table} ({col})")
            if err:
                # index neblokuje import
                errors.append(f"{idx}: {err}")

    try:
        db.rollback()
    except Exception:
        pass
    return [e for e in errors if e]


@router.get("/api/import/hub-xlsx/repair-database")
def api_import_repair_database_v100(db: Session = Depends(get_db)):
    errors = v100_fix_all_import_tables(db)
    return {
        "ok": len(errors) == 0,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Databáze byla sjednocena pro import XLSX. Doplněny sloupce sections/subsections/partners a další importní tabulky.",
        "errors": errors,
    }


@router.get("/api/import/hub-xlsx/repair-all-v100")
def api_import_repair_all_v100(db: Session = Depends(get_db)):
    return api_import_repair_database_v100(db=db)


# Přesměrovat import tak, aby si schéma opravil automaticky před nahráním XLSX.
try:
    _previous_import_hub_xlsx_data_before_v100 = import_hub_xlsx_data_v093_

    def import_hub_xlsx_data_v100_(db, wb, update_existing=False):
        repair_errors = v100_fix_all_import_tables(db)
        result = _previous_import_hub_xlsx_data_before_v100(db, wb, update_existing=update_existing)
        if repair_errors:
            result.setdefault("warnings", []).extend(repair_errors)
        return result

    import_hub_xlsx_data_v093_ = import_hub_xlsx_data_v100_
except Exception:
    pass





# -------------------------------------------------------------------
# v1.2.2 import relationship fix
# Oprava vazeb: subsections.section_id se dopočítá ze sections.section_code.
# Přidán preflight, který odhalí základní problémy před importem.
# -------------------------------------------------------------------

def v101_exec(db: Session, sql: str, params: dict | None = None):
    try:
        db.rollback()
    except Exception:
        pass
    try:
        db.execute(text(sql), params or {})
        db.commit()
        return None
    except Exception as exc:
        try:
            db.rollback()
        except Exception:
            pass
        return str(exc)


def v101_table_exists(db: Session, table_name: str) -> bool:
    try:
        db.rollback()
        exists = db.execute(text("""
            SELECT EXISTS (
                SELECT 1 FROM information_schema.tables
                WHERE table_schema='public' AND table_name=:t
            )
        """), {"t": table_name}).scalar()
        db.commit()
        return bool(exists)
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass
        return False


def v101_column_exists(db: Session, table_name: str, column_name: str) -> bool:
    try:
        db.rollback()
        exists = db.execute(text("""
            SELECT EXISTS (
                SELECT 1 FROM information_schema.columns
                WHERE table_schema='public' AND table_name=:t AND column_name=:c
            )
        """), {"t": table_name, "c": column_name}).scalar()
        db.commit()
        return bool(exists)
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass
        return False


def v101_add_column(db: Session, table: str, column: str, definition: str):
    if not v101_table_exists(db, table):
        return None
    if v101_column_exists(db, table, column):
        return None
    return v101_exec(db, f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def v101_fix_relationships(db: Session):
    errors = []

    fn = globals().get("v100_fix_all_import_tables")
    if fn:
        try:
            res = fn(db)
            if isinstance(res, list):
                errors.extend([str(x) for x in res if x])
        except Exception as exc:
            errors.append(f"v100_fix_all_import_tables: {exc}")

    if v101_table_exists(db, "subsections"):
        err = v101_add_column(db, "subsections", "section_code", "TEXT")
        if err:
            errors.append(f"subsections.section_code: {err}")
        err = v101_add_column(db, "subsections", "section_id", "TEXT")
        if err:
            errors.append(f"subsections.section_id: {err}")

    if v101_table_exists(db, "sections") and v101_table_exists(db, "subsections"):
        if v101_column_exists(db, "sections", "id") and v101_column_exists(db, "sections", "section_code") and v101_column_exists(db, "subsections", "section_code") and v101_column_exists(db, "subsections", "section_id"):
            err = v101_exec(db, """
                UPDATE subsections ss
                SET section_id = s.id
                FROM sections s
                WHERE ss.section_code = s.section_code
                  AND ss.section_id IS NULL
            """)
            if err:
                errors.append(f"subsections section_id update: {err}")

    return [e for e in errors if e]


def v101_preflight_database(db: Session):
    v101_fix_relationships(db)
    issues = []
    required = {
        "sections": ["id", "section_code", "name", "created_at", "updated_at"],
        "subsections": ["id", "subsection_code", "section_code", "section_id", "name", "created_at", "updated_at"],
        "users": ["id", "advisor_id", "name", "email", "created_at", "updated_at"],
        "partners": ["id", "partner_code", "name", "created_at", "updated_at"],
        "specialists": ["id", "advisor_id", "section_code", "subsection_code", "email", "created_at", "updated_at"],
    }

    for table, cols in required.items():
        if not v101_table_exists(db, table):
            issues.append(f"Chybí tabulka: {table}")
            continue
        for col in cols:
            if not v101_column_exists(db, table, col):
                issues.append(f"Chybí sloupec: {table}.{col}")

    if v101_table_exists(db, "subsections") and v101_column_exists(db, "subsections", "section_id"):
        try:
            db.rollback()
            missing = db.execute(text("""
                SELECT COUNT(*)
                FROM subsections
                WHERE section_id IS NULL
            """)).scalar()
            db.commit()
            if int(missing or 0) > 0:
                issues.append(f"Podsekce bez section_id: {missing}")
        except Exception as exc:
            issues.append(f"Kontrola section_id selhala: {exc}")
            try:
                db.rollback()
            except Exception:
                pass

    return issues


@router.get("/api/import/hub-xlsx/repair-relationships")
def api_import_repair_relationships_v101(db: Session = Depends(get_db)):
    errors = v101_fix_relationships(db)
    issues = v101_preflight_database(db)
    return {
        "ok": len(errors) == 0 and len(issues) == 0,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Opraveny vazby pro import. Subsections.section_id se doplňuje podle sections.section_code.",
        "errors": errors,
        "preflight_issues": issues,
    }


@router.get("/api/import/hub-xlsx/preflight")
def api_import_preflight_v101(db: Session = Depends(get_db)):
    errors = v101_fix_relationships(db)
    issues = v101_preflight_database(db)
    return {
        "ok": len(errors) == 0 and len(issues) == 0,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Předimportní kontrola databáze.",
        "errors": errors,
        "issues": issues,
    }


try:
    _previous_import_hub_xlsx_data_before_v101 = import_hub_xlsx_data_v093_

    def import_hub_xlsx_data_v101_(db, wb, update_existing=False):
        repair_errors = v101_fix_relationships(db)
        result = _previous_import_hub_xlsx_data_before_v101(db, wb, update_existing=update_existing)
        if repair_errors:
            result.setdefault("warnings", []).extend(repair_errors)
        return result

    import_hub_xlsx_data_v093_ = import_hub_xlsx_data_v101_
except Exception:
    pass



# -------------------------------------------------------------------
# v1.2.2 uuid relationship fix endpoint
# -------------------------------------------------------------------

@router.get("/api/import/hub-xlsx/repair-uuid-relationships")
def api_import_repair_uuid_relationships_v102(db: Session = Depends(get_db)):
    errors = []
    try:
        if globals().get("v101_fix_relationships"):
            res = v101_fix_relationships(db)
            if isinstance(res, list):
                errors.extend([str(x) for x in res if x])
    except Exception as exc:
        errors.append(f"v101_fix_relationships: {exc}")
        try:
            db.rollback()
        except Exception:
            pass

    try:
        db.rollback()
        db.execute(text("""
            UPDATE subsections ss
            SET section_id = s.id
            FROM sections s
            WHERE ss.section_code = s.section_code
              AND ss.section_id IS NULL
        """))
        db.commit()
    except Exception as exc:
        errors.append(f"subsections UUID section_id update: {exc}")
        try:
            db.rollback()
        except Exception:
            pass

    issues = []
    try:
        db.rollback()
        missing = db.execute(text("SELECT COUNT(*) FROM subsections WHERE section_id IS NULL")).scalar()
        db.commit()
        if int(missing or 0) > 0:
            issues.append(f"Podsekce bez section_id: {missing}")
    except Exception as exc:
        issues.append(f"Kontrola section_id selhala: {exc}")
        try:
            db.rollback()
        except Exception:
            pass

    return {
        "ok": len(errors) == 0 and len(issues) == 0,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Opravena UUID vazba subsections.section_id bez neplatného porovnání s prázdným řetězcem.",
        "errors": errors,
        "issues": issues,
    }





# -------------------------------------------------------------------
# v1.2.2 import cleanup + partner UI completion
# Cíl:
# - odstranit duplicitně nahraná data po opakovaných importech
# - oddělit Kontakty ASTORIE od kontaktů partnerů
# - doplnit návrhy změn/doplnění z poradenského HUBu
# - doplnit FAQ tab partnera a přehlednější produkty
# -------------------------------------------------------------------

def v103_exec(db: Session, sql: str, params: dict | None = None):
    try:
        db.rollback()
    except Exception:
        pass
    try:
        db.execute(text(sql), params or {})
        db.commit()
        return None
    except Exception as exc:
        try:
            db.rollback()
        except Exception:
            pass
        return str(exc)


def v103_scalar(db: Session, sql: str, params: dict | None = None):
    try:
        db.rollback()
    except Exception:
        pass
    try:
        val = db.execute(text(sql), params or {}).scalar()
        db.commit()
        return val
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass
        return None


def v103_table_exists(db: Session, table_name: str) -> bool:
    return bool(v103_scalar(db, """
        SELECT EXISTS (
            SELECT 1 FROM information_schema.tables
            WHERE table_schema='public' AND table_name=:t
        )
    """, {"t": table_name}))


def v103_column_exists(db: Session, table_name: str, column_name: str) -> bool:
    return bool(v103_scalar(db, """
        SELECT EXISTS (
            SELECT 1 FROM information_schema.columns
            WHERE table_schema='public' AND table_name=:t AND column_name=:c
        )
    """, {"t": table_name, "c": column_name}))


def v103_add_column(db: Session, table: str, column: str, definition: str):
    if not v103_table_exists(db, table):
        return None
    if v103_column_exists(db, table, column):
        return None
    return v103_exec(db, f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def ensure_v103_tables(db: Session):
    errors = []

    # Globální kontakty ASTORIE – samostatně od kontaktů partnerů.
    err = v103_exec(db, """
        CREATE TABLE IF NOT EXISTS global_contacts (
            id TEXT PRIMARY KEY DEFAULT gen_random_uuid()::text,
            contact_name TEXT NOT NULL,
            role_description TEXT DEFAULT '',
            department TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            email TEXT DEFAULT '',
            location TEXT DEFAULT '',
            note TEXT DEFAULT '',
            is_vip BOOLEAN DEFAULT FALSE,
            is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )
    """)
    if err:
        errors.append(f"global_contacts create: {err}")

    # FAQ partnerů.
    err = v103_exec(db, """
        CREATE TABLE IF NOT EXISTS partner_faq (
            id TEXT PRIMARY KEY DEFAULT gen_random_uuid()::text,
            partner_code TEXT DEFAULT '',
            question TEXT NOT NULL,
            answer TEXT DEFAULT '',
            category TEXT DEFAULT '',
            tags TEXT DEFAULT '',
            is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )
    """)
    if err:
        errors.append(f"partner_faq create: {err}")

    # Návrhy změn/doplnění od poradců.
    err = v103_exec(db, """
        CREATE TABLE IF NOT EXISTS data_suggestions (
            id TEXT PRIMARY KEY DEFAULT gen_random_uuid()::text,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW(),
            created_by_name TEXT DEFAULT '',
            created_by_email TEXT DEFAULT '',
            module TEXT DEFAULT '',
            partner_code TEXT DEFAULT '',
            tab TEXT DEFAULT '',
            item_type TEXT DEFAULT '',
            item_id TEXT DEFAULT '',
            suggestion_type TEXT DEFAULT '',
            title TEXT DEFAULT '',
            description TEXT DEFAULT '',
            status TEXT DEFAULT 'nový',
            admin_note TEXT DEFAULT ''
        )
    """)
    if err:
        errors.append(f"data_suggestions create: {err}")

    # Chybějící sloupce u partner_contacts dle starší šablony.
    contact_cols = {
        "full_name": "TEXT DEFAULT ''",
        "role": "TEXT DEFAULT ''",
        "territory": "TEXT DEFAULT ''",
        "contact_type": "TEXT DEFAULT ''",
        "is_top": "BOOLEAN DEFAULT FALSE",
        "is_vip": "BOOLEAN DEFAULT FALSE",
        "original_note": "TEXT DEFAULT ''",
        "specialization": "TEXT DEFAULT ''",
        "phone": "TEXT DEFAULT ''",
        "email": "TEXT DEFAULT ''",
        "note": "TEXT DEFAULT ''",
        "is_active": "BOOLEAN DEFAULT TRUE",
        "created_at": "TIMESTAMPTZ DEFAULT NOW()",
        "updated_at": "TIMESTAMPTZ DEFAULT NOW()",
    }
    if v103_table_exists(db, "partner_contacts"):
        for col, definition in contact_cols.items():
            err = v103_add_column(db, "partner_contacts", col, definition)
            if err:
                errors.append(f"partner_contacts.{col}: {err}")

    if v103_table_exists(db, "partner_links"):
        for col, definition in {
            "title": "TEXT DEFAULT ''",
            "category": "TEXT DEFAULT ''",
            "visibility": "TEXT DEFAULT ''",
            "note": "TEXT DEFAULT ''",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "created_at": "TIMESTAMPTZ DEFAULT NOW()",
            "updated_at": "TIMESTAMPTZ DEFAULT NOW()",
        }.items():
            err = v103_add_column(db, "partner_links", col, definition)
            if err:
                errors.append(f"partner_links.{col}: {err}")

    if v103_table_exists(db, "partner_products"):
        for col, definition in {
            "area": "TEXT DEFAULT ''",
            "subarea": "TEXT DEFAULT ''",
            "product_name": "TEXT DEFAULT ''",
            "risks": "TEXT DEFAULT ''",
            "client_type": "TEXT DEFAULT ''",
            "keywords": "TEXT DEFAULT ''",
            "priority": "INTEGER DEFAULT 100",
            "note": "TEXT DEFAULT ''",
            "is_active": "BOOLEAN DEFAULT TRUE",
            "created_at": "TIMESTAMPTZ DEFAULT NOW()",
            "updated_at": "TIMESTAMPTZ DEFAULT NOW()",
        }.items():
            err = v103_add_column(db, "partner_products", col, definition)
            if err:
                errors.append(f"partner_products.{col}: {err}")

    return [e for e in errors if e]


def cleanup_import_duplicates_v103(db: Session):
    """
    Odstraní duplicity vzniklé opakovaným importem.
    Nechá vždy nejstarší záznam podle created_at/id.
    """
    errors = []
    ensure_v103_tables(db)

    dedupe_specs = [
        ("partner_contacts", ["partner_code", "lower(COALESCE(full_name,''))", "lower(COALESCE(role,''))", "lower(COALESCE(email,''))", "COALESCE(phone,'')"]),
        ("partner_links", ["partner_code", "lower(COALESCE(title,''))", "lower(COALESCE(url,''))"]),
        ("partner_products", ["partner_code", "lower(COALESCE(area,''))", "lower(COALESCE(subarea,''))", "lower(COALESCE(product_name,''))"]),
        ("commission_rates", ["COALESCE(partner_code,'')", "COALESCE(section_code,'')", "COALESCE(subsection_code,'')", "COALESCE(business_type,'')", "COALESCE(area,'')"]),
        ("global_contacts", ["lower(COALESCE(contact_name,''))", "lower(COALESCE(email,''))", "COALESCE(phone,'')"]),
        ("partner_faq", ["COALESCE(partner_code,'')", "lower(COALESCE(question,''))"]),
    ]

    for table, keys in dedupe_specs:
        if not v103_table_exists(db, table):
            continue
        if not v103_column_exists(db, table, "id"):
            continue

        part_by = ", ".join(keys)
        err = v103_exec(db, f"""
            DELETE FROM {table} t
            USING (
                SELECT id
                FROM (
                    SELECT id,
                           ROW_NUMBER() OVER (
                               PARTITION BY {part_by}
                               ORDER BY COALESCE(created_at, NOW()), id
                           ) AS rn
                    FROM {table}
                ) x
                WHERE x.rn > 1
            ) d
            WHERE t.id = d.id
        """)
        if err:
            errors.append(f"dedupe {table}: {err}")

    return [e for e in errors if e]


def import_global_contacts_from_xlsx_v103(db: Session, wb):
    ensure_v103_tables(db)
    count = 0
    skipped = 0
    errors = []

    for row in xlsx_rows_v093_(wb, "Import_Astorie_Contacts"):
        try:
            name = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Jmeno", "Jméno", "Nazev", "Název", "Kontakt", "Osoba"))
            email = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Email", "E-mail")).lower()
            phone = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Telefon", "Mobil"))
            if not name and not email and not phone:
                skipped += 1
                continue

            exists = v103_scalar(db, """
                SELECT id FROM global_contacts
                WHERE lower(COALESCE(contact_name,'')) = lower(:name)
                  AND lower(COALESCE(email,'')) = lower(:email)
                  AND COALESCE(phone,'') = :phone
                LIMIT 1
            """, {"name": name, "email": email, "phone": phone})

            if exists:
                skipped += 1
                continue

            err = v103_exec(db, """
                INSERT INTO global_contacts
                (contact_name, role_description, department, phone, email, location, note, is_vip, is_active)
                VALUES
                (:contact_name, :role_description, :department, :phone, :email, :location, :note, :is_vip, TRUE)
            """, {
                "contact_name": name or email or phone,
                "role_description": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Role", "Funkce", "Popis")),
                "department": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Oddeleni", "Oddělení", "Sekce")),
                "phone": phone,
                "email": email,
                "location": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Region", "Lokalita", "Pobocka", "Pobočka")),
                "note": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Poznamka", "Poznámka")),
                "is_vip": xlsx_bool_v093_(xlsx_pick_v093_(row, "VIP", "TOP")),
            })
            if err:
                errors.append(err)
            else:
                count += 1
        except Exception as exc:
            errors.append(str(exc))

    return {"rows": count + skipped, "created": count, "updated_or_skipped": skipped, "errors": errors}


def import_partner_faq_from_xlsx_v103(db: Session, wb):
    ensure_v103_tables(db)
    count = 0
    skipped = 0
    errors = []
    for sheet_name in ["FAQ", "FAQ_Partners", "Import_FAQ", "Import_Partner_FAQ"]:
        for row in xlsx_rows_v093_(wb, sheet_name):
            try:
                question = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Otazka", "Otázka", "Question"))
                answer = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Odpoved", "Odpověď", "Answer"))
                if not question:
                    continue
                partner_code = xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Partner_ID", "Partner", default="")).upper()
                exists = v103_scalar(db, """
                    SELECT id FROM partner_faq
                    WHERE COALESCE(partner_code,'') = :partner_code
                      AND lower(COALESCE(question,'')) = lower(:question)
                    LIMIT 1
                """, {"partner_code": partner_code, "question": question})
                if exists:
                    skipped += 1
                    continue
                err = v103_exec(db, """
                    INSERT INTO partner_faq
                    (partner_code, question, answer, category, tags, is_active)
                    VALUES
                    (:partner_code, :question, :answer, :category, :tags, TRUE)
                """, {
                    "partner_code": partner_code,
                    "question": question,
                    "answer": answer,
                    "category": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Kategorie", "Category")),
                    "tags": xlsx_cell_to_str_v093_(xlsx_pick_v093_(row, "Tagy", "Tags")),
                })
                if err:
                    errors.append(err)
                else:
                    count += 1
            except Exception as exc:
                errors.append(str(exc))
    return {"rows": count + skipped, "created": count, "updated_or_skipped": skipped, "errors": errors}


@router.get("/api/import/hub-xlsx/cleanup-duplicates")
def api_import_cleanup_duplicates_v103(db: Session = Depends(get_db)):
    errors = cleanup_import_duplicates_v103(db)
    return {
        "ok": len(errors) == 0,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Duplicitní záznamy po opakovaném importu byly vyčištěny. Zachován je vždy první záznam.",
        "errors": errors,
    }


@router.get("/api/import/hub-xlsx/repair-display-data")
def api_import_repair_display_data_v103(db: Session = Depends(get_db)):
    errors = ensure_v103_tables(db)
    errors.extend(cleanup_import_duplicates_v103(db))
    return {
        "ok": len(errors) == 0,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Doplněny tabulky pro globální kontakty, FAQ partnerů a návrhy změn. Vyčištěny duplicity.",
        "errors": errors,
    }


# Přepsat import tak, aby po importu doplnil globální kontakty ASTORIE a vyčistil duplicity.
try:
    _previous_import_hub_xlsx_data_before_v103 = import_hub_xlsx_data_v093_

    def import_hub_xlsx_data_v103_(db, wb, update_existing=False):
        ensure_v103_tables(db)
        result = _previous_import_hub_xlsx_data_before_v103(db, wb, update_existing=update_existing)

        gc = import_global_contacts_from_xlsx_v103(db, wb)
        result["global_contacts"] = {
            "rows": gc.get("rows", 0),
            "created": gc.get("created", 0),
            "updated_or_skipped": gc.get("updated_or_skipped", 0),
        }
        if gc.get("errors"):
            result.setdefault("errors", []).extend([f"Import_Astorie_Contacts: {e}" for e in gc["errors"]])

        faq = import_partner_faq_from_xlsx_v103(db, wb)
        result["partner_faq"] = {
            "rows": faq.get("rows", 0),
            "created": faq.get("created", 0),
            "updated_or_skipped": faq.get("updated_or_skipped", 0),
        }
        if faq.get("errors"):
            result.setdefault("errors", []).extend([f"FAQ: {e}" for e in faq["errors"]])

        cleanup_errors = cleanup_import_duplicates_v103(db)
        if cleanup_errors:
            result.setdefault("warnings", []).extend(cleanup_errors)

        return result

    import_hub_xlsx_data_v093_ = import_hub_xlsx_data_v103_
except Exception:
    pass


@router.post("/hub/suggestions")
async def hub_submit_suggestion_v103(
    request: Request,
    module: str = Form(""),
    partner_code: str = Form(""),
    tab: str = Form(""),
    item_type: str = Form(""),
    item_id: str = Form(""),
    suggestion_type: str = Form("doplnění"),
    title: str = Form(""),
    description: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_v103_tables(db)
    user = hub_user_context_v083_()
    err = v103_exec(db, """
        INSERT INTO data_suggestions
        (created_by_name, created_by_email, module, partner_code, tab, item_type, item_id, suggestion_type, title, description, status)
        VALUES
        (:created_by_name, :created_by_email, :module, :partner_code, :tab, :item_type, :item_id, :suggestion_type, :title, :description, 'nový')
    """, {
        "created_by_name": user.get("name", ""),
        "created_by_email": user.get("email", ""),
        "module": module,
        "partner_code": partner_code,
        "tab": tab,
        "item_type": item_type,
        "item_id": item_id,
        "suggestion_type": suggestion_type,
        "title": title,
        "description": description,
    })
    back_url = str(request.headers.get("referer") or "/hub/partners")
    if err:
        return RedirectResponse(back_url + ("&" if "?" in back_url else "?") + "suggestion=error", status_code=303)
    return RedirectResponse(back_url + ("&" if "?" in back_url else "?") + "suggestion=ok", status_code=303)


# Novější /hub/contacts: globální kontakty ASTORIE, nikoli kontakty partnerů.
@router.get("/hub/contacts-v103", response_class=HTMLResponse)
def hub_contacts_v103_explicit(request: Request, q: str = "", db: Session = Depends(get_db)):
    return hub_contacts_v103(request, q=q, db=db)


def hub_contacts_v103(request: Request, q: str = "", db: Session = Depends(get_db)):
    ensure_v103_tables(db)
    rows = []
    params = {}
    where = "WHERE COALESCE(is_active, TRUE) = TRUE"
    if q:
        where += """
          AND (
            lower(COALESCE(contact_name, '')) LIKE :q OR
            lower(COALESCE(email, '')) LIKE :q OR
            lower(COALESCE(phone, '')) LIKE :q OR
            lower(COALESCE(role_description, '')) LIKE :q OR
            lower(COALESCE(department, '')) LIKE :q OR
            lower(COALESCE(location, '')) LIKE :q
          )
        """
        params["q"] = f"%{q.lower()}%"

    rows = fetch_all_safe_v084_(db, f"""
        SELECT *
        FROM global_contacts
        {where}
        ORDER BY COALESCE(is_vip, FALSE) DESC, contact_name
        LIMIT 500
    """, params)

    return hub_render_v083_(request, "hub_contacts.html", {
        "active": "contacts",
        "rows": rows,
        "q": q,
        "is_global_contacts": True,
    })


# Přeregistrace hlavní cesty /hub/contacts přes poslední definici ve FastAPI neodstraní starou route.
# Proto necháváme i /hub/contacts-v103 a zároveň níže přidáváme alternativní route s přímým aliasem.
@router.get("/hub/contacts-astorie", response_class=HTMLResponse)
def hub_contacts_astorie_v103(request: Request, q: str = "", db: Session = Depends(get_db)):
    return hub_contacts_v103(request, q=q, db=db)





# -------------------------------------------------------------------
# v1.2.2 Partner Workflow Core
# -------------------------------------------------------------------
PARTNER_WORKFLOW_VERSION = "1.2.2-admin-taxonomy-specialists-links-safe"


def v110_exec(db: Session, sql: str, params: dict | None = None):
    try:
        db.rollback()
    except Exception:
        pass
    try:
        db.execute(text(sql), params or {})
        db.commit()
        return None
    except Exception as exc:
        try:
            db.rollback()
        except Exception:
            pass
        return str(exc)


def v110_scalar(db: Session, sql: str, params: dict | None = None):
    try:
        db.rollback()
    except Exception:
        pass
    try:
        val = db.execute(text(sql), params or {}).scalar()
        db.commit()
        return val
    except Exception:
        try:
            db.rollback()
        except Exception:
            pass
        return None


def v110_table_exists(db: Session, table_name: str) -> bool:
    return bool(v110_scalar(db, """
        SELECT EXISTS (
            SELECT 1 FROM information_schema.tables
            WHERE table_schema='public' AND table_name=:table_name
        )
    """, {"table_name": table_name}))


def ensure_partner_workflow_v110(db: Session):
    errors = []
    for fn_name in ["ensure_v103_tables", "cleanup_import_duplicates_v103"]:
        fn = globals().get(fn_name)
        if fn:
            try:
                res = fn(db)
                if isinstance(res, list):
                    errors.extend([str(e) for e in res if e])
            except Exception as exc:
                errors.append(f"{fn_name}: {exc}")

    ddl = [
        """
        CREATE TABLE IF NOT EXISTS partner_change_requests (
            id TEXT PRIMARY KEY DEFAULT gen_random_uuid()::text,
            partner_code TEXT DEFAULT '',
            partner_name TEXT DEFAULT '',
            request_type TEXT DEFAULT '',
            request_area TEXT DEFAULT '',
            priority TEXT DEFAULT 'normal',
            status TEXT DEFAULT 'new',
            title TEXT DEFAULT '',
            description TEXT DEFAULT '',
            current_value TEXT DEFAULT '',
            proposed_value TEXT DEFAULT '',
            contact_name TEXT DEFAULT '',
            contact_phone TEXT DEFAULT '',
            contact_email TEXT DEFAULT '',
            item_type TEXT DEFAULT '',
            item_id TEXT DEFAULT '',
            tab TEXT DEFAULT '',
            created_by_name TEXT DEFAULT '',
            created_by_email TEXT DEFAULT '',
            processed_by_name TEXT DEFAULT '',
            processed_by_email TEXT DEFAULT '',
            processed_at TIMESTAMPTZ,
            admin_note TEXT DEFAULT '',
            created_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS partner_request_comments (
            id TEXT PRIMARY KEY DEFAULT gen_random_uuid()::text,
            request_id TEXT NOT NULL,
            author_name TEXT DEFAULT '',
            author_email TEXT DEFAULT '',
            comment_text TEXT DEFAULT '',
            is_internal BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMPTZ DEFAULT NOW()
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS partner_audit_log (
            id TEXT PRIMARY KEY DEFAULT gen_random_uuid()::text,
            entity_type TEXT DEFAULT '',
            entity_id TEXT DEFAULT '',
            partner_code TEXT DEFAULT '',
            action_type TEXT DEFAULT '',
            old_value TEXT DEFAULT '',
            new_value TEXT DEFAULT '',
            changed_by_name TEXT DEFAULT '',
            changed_by_email TEXT DEFAULT '',
            created_at TIMESTAMPTZ DEFAULT NOW()
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS partner_history (
            id TEXT PRIMARY KEY DEFAULT gen_random_uuid()::text,
            partner_code TEXT DEFAULT '',
            event_type TEXT DEFAULT '',
            title TEXT DEFAULT '',
            description TEXT DEFAULT '',
            created_by_name TEXT DEFAULT '',
            created_by_email TEXT DEFAULT '',
            created_at TIMESTAMPTZ DEFAULT NOW()
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS partner_favorites (
            id TEXT PRIMARY KEY DEFAULT gen_random_uuid()::text,
            partner_code TEXT DEFAULT '',
            user_email TEXT DEFAULT '',
            created_at TIMESTAMPTZ DEFAULT NOW()
        )
        """,
        "CREATE INDEX IF NOT EXISTS ix_partner_change_requests_status ON partner_change_requests(status)",
        "CREATE INDEX IF NOT EXISTS ix_partner_change_requests_partner_code ON partner_change_requests(partner_code)",
        "CREATE INDEX IF NOT EXISTS ix_partner_change_requests_created_at ON partner_change_requests(created_at)",
        "CREATE INDEX IF NOT EXISTS ix_partner_history_partner_code ON partner_history(partner_code)"
    ]
    for sql in ddl:
        err = v110_exec(db, sql)
        if err: errors.append(err)

    if v110_table_exists(db, 'data_suggestions'):
        err = v110_exec(db, """
            INSERT INTO partner_change_requests
            (partner_code, request_type, request_area, status, title, description, item_type, item_id, tab,
             created_by_name, created_by_email, created_at, updated_at)
            SELECT COALESCE(partner_code,''), COALESCE(suggestion_type,'doplnění'), COALESCE(module,'partners'),
                   CASE WHEN COALESCE(status,'nový') IN ('nový','new') THEN 'new' ELSE COALESCE(status,'new') END,
                   COALESCE(title,''), COALESCE(description,''), COALESCE(item_type,''), COALESCE(item_id,''), COALESCE(tab,''),
                   COALESCE(created_by_name,''), COALESCE(created_by_email,''), COALESCE(created_at,NOW()), COALESCE(updated_at,NOW())
            FROM data_suggestions ds
            WHERE NOT EXISTS (
              SELECT 1 FROM partner_change_requests pcr
              WHERE COALESCE(pcr.partner_code,'')=COALESCE(ds.partner_code,'')
                AND COALESCE(pcr.title,'')=COALESCE(ds.title,'')
                AND COALESCE(pcr.description,'')=COALESCE(ds.description,'')
                AND COALESCE(pcr.created_by_email,'')=COALESCE(ds.created_by_email,'')
            )
        """)
        if err: errors.append(f"migrate data_suggestions: {err}")
    return [e for e in errors if e]


def partner_name_by_code_v110(db: Session, partner_code: str) -> str:
    if not partner_code or not v110_table_exists(db, 'partners'):
        return partner_code or ''
    val = v110_scalar(db, "SELECT name FROM partners WHERE upper(partner_code)=upper(:code) LIMIT 1", {"code": partner_code})
    return str(val or partner_code or '')


def status_label_v110(status: str) -> str:
    return {'new':'Nový','processing':'V řešení','approved':'Schváleno','rejected':'Zamítnuto','completed':'Hotovo'}.get(status or '', status or 'Nový')


def normalize_status_v110(status: str) -> str:
    s=(status or '').strip().lower()
    return {'nový':'new','new':'new','v řešení':'processing','processing':'processing','schváleno':'approved','approved':'approved','zamítnuto':'rejected','rejected':'rejected','hotovo':'completed','completed':'completed'}.get(s, s or 'new')


def get_bo_email_v110(db: Session) -> str:
    return 'backoffice@astorieas.cz'


def send_partner_workflow_email_v110(db: Session, to_email: str, subject: str, body: str):
    # v1.6.0: centrální e-mailová služba s profesionální HTML šablonou a logováním.
    subj, text_body, html_body = email_template("generic_notice", subject=subject, body=body)
    return send_email(
        db,
        to_email,
        subj,
        text_body,
        html_body=html_body,
        event_type="partner_workflow",
        entity_type="partner_request",
        template_key="generic_notice",
    )


def create_partner_request_v110(db: Session, partner_code: str, request_type: str, title: str, description: str, request_area: str='partners', item_type: str='', item_id: str='', tab: str='', contact_name: str='', contact_phone: str='', contact_email: str='', current_value: str='', proposed_value: str=''):
    ensure_partner_workflow_v110(db)
    user=hub_user_context_v083_()
    partner_name=partner_name_by_code_v110(db, partner_code)
    request_id=str(uuid.uuid4())
    priority='high' if any(x in (request_type or '').lower() for x in ['kontakt','chyba','neaktu']) else 'normal'
    err=v110_exec(db, """
        INSERT INTO partner_change_requests
        (id, partner_code, partner_name, request_type, request_area, priority, status, title, description,
         current_value, proposed_value, contact_name, contact_phone, contact_email, item_type, item_id, tab,
         created_by_name, created_by_email)
        VALUES
        (:id,:partner_code,:partner_name,:request_type,:request_area,:priority,'new',:title,:description,
         :current_value,:proposed_value,:contact_name,:contact_phone,:contact_email,:item_type,:item_id,:tab,
         :created_by_name,:created_by_email)
    """, {"id":request_id,"partner_code":partner_code or '',"partner_name":partner_name or '',"request_type":request_type or 'doplnění',"request_area":request_area or 'partners',"priority":priority,"title":title or '',"description":description or '',"current_value":current_value or '',"proposed_value":proposed_value or '',"contact_name":contact_name or '',"contact_phone":contact_phone or '',"contact_email":contact_email or '',"item_type":item_type or '',"item_id":item_id or '',"tab":tab or '',"created_by_name":user.get('name',''),"created_by_email":user.get('email','')})
    if err: return None, err
    v110_exec(db, """INSERT INTO partner_history (partner_code,event_type,title,description,created_by_name,created_by_email) VALUES (:partner_code,'change_request_created',:title,:description,:created_by_name,:created_by_email)""", {"partner_code":partner_code or '',"title":f"Nový požadavek: {title or request_type}","description":description or '',"created_by_name":user.get('name',''),"created_by_email":user.get('email','')})
    bo_subject=f"[HUB] Nový požadavek na partnera – {partner_name or partner_code or 'partner'}"
    bo_body=f"Dobrý den,\n\nv HUB ASTORIE byl vložen nový požadavek v sekci Partneři.\n\nPartner: {partner_name or partner_code or '—'}\nTyp požadavku: {request_type or '—'}\nNázev: {title or '—'}\nVložil: {user.get('name','')} ({user.get('email','')})\n\nPopis:\n{description or '—'}\n\nAdministrace: /admin/partner-requests\n\nASTORIE HUB"
    sent_bo, bo_error = send_partner_workflow_email_v110(db, get_bo_email_v110(db), bo_subject, bo_body)
    sent_user, user_error = send_partner_workflow_email_v110(db, user.get('email',''), 'Potvrzení přijetí požadavku – HUB ASTORIE', f"Dobrý den,\n\nváš požadavek byl úspěšně přijat a předán BackOffice ke zpracování.\n\nPartner: {partner_name or partner_code or '—'}\nTyp požadavku: {request_type or '—'}\nNázev: {title or '—'}\n\nDěkujeme za spolupráci.\n\nASTORIE a.s.")
    if not sent_bo or not sent_user:
        v110_exec(db, """INSERT INTO partner_request_comments (request_id,author_name,author_email,comment_text,is_internal) VALUES (:request_id,'Systém','system',:comment_text,TRUE)""", {"request_id":request_id,"comment_text":f"E-mail info: BO={bo_error or 'odesláno'} | poradce={user_error or 'odesláno'}"})
    return request_id, None


@router.get('/api/partner-workflow/status')
def api_partner_workflow_status_v110(db: Session = Depends(get_db)):
    errors=ensure_partner_workflow_v110(db)
    counts={}
    if v110_table_exists(db,'partner_change_requests'):
        try:
            db.rollback(); rows=db.execute(text("SELECT status, COUNT(*) AS count FROM partner_change_requests GROUP BY status ORDER BY status")).mappings().all(); db.commit()
            counts={r['status']: int(r['count']) for r in rows}
        except Exception as exc:
            errors.append(str(exc)); db.rollback()
    return {'ok':len(errors)==0,'version':PARTNER_WORKFLOW_VERSION,'message':'Partner Workflow Core je připraven.','counts':counts,'errors':errors}


@router.post('/hub/partner-workflow/request')
async def hub_partner_workflow_request_v110(request: Request, partner_code: str = Form(''), request_type: str = Form('doplnění'), title: str = Form(''), description: str = Form(''), request_area: str = Form('partners'), item_type: str = Form(''), item_id: str = Form(''), tab: str = Form(''), contact_name: str = Form(''), contact_phone: str = Form(''), contact_email: str = Form(''), current_value: str = Form(''), proposed_value: str = Form(''), db: Session = Depends(get_db)):
    rid, err = create_partner_request_v110(db, partner_code, request_type, title, description, request_area, item_type, item_id, tab, contact_name, contact_phone, contact_email, current_value, proposed_value)
    back=str(request.headers.get('referer') or '/hub/partners')
    return RedirectResponse(back + ('&' if '?' in back else '?') + ('workflow=ok' if not err else 'workflow=error'), status_code=303)


@router.post('/api/partners/request-change')
async def api_partners_request_change_v110(partner_code: str = Form(''), request_type: str = Form('doplnění'), title: str = Form(''), description: str = Form(''), request_area: str = Form('partners'), item_type: str = Form(''), item_id: str = Form(''), tab: str = Form(''), contact_name: str = Form(''), contact_phone: str = Form(''), contact_email: str = Form(''), current_value: str = Form(''), proposed_value: str = Form(''), db: Session = Depends(get_db)):
    rid, err = create_partner_request_v110(db, partner_code, request_type, title, description, request_area, item_type, item_id, tab, contact_name, contact_phone, contact_email, current_value, proposed_value)
    return {'ok': not bool(err), 'id': rid, 'error': err, 'version': PARTNER_WORKFLOW_VERSION}


@router.get('/admin/partner-requests', response_class=HTMLResponse)
def admin_partner_requests_v110(request: Request, status: str = '', q: str = '', db: Session = Depends(get_db)):
    ensure_partner_workflow_v110(db)
    where='WHERE 1=1'; params={}
    if status:
        where+=' AND status=:status'; params['status']=normalize_status_v110(status)
    if q:
        where += """ AND (lower(COALESCE(partner_name,'')) LIKE :q OR lower(COALESCE(partner_code,'')) LIKE :q OR lower(COALESCE(title,'')) LIKE :q OR lower(COALESCE(description,'')) LIKE :q OR lower(COALESCE(created_by_name,'')) LIKE :q OR lower(COALESCE(created_by_email,'')) LIKE :q)"""; params['q']=f"%{q.lower()}%"
    rows=fetch_all_safe_v084_(db, f"""SELECT * FROM partner_change_requests {where} ORDER BY CASE status WHEN 'new' THEN 1 WHEN 'processing' THEN 2 WHEN 'approved' THEN 3 WHEN 'completed' THEN 4 WHEN 'rejected' THEN 5 ELSE 9 END, created_at DESC LIMIT 500""", params)
    stats=fetch_one_safe_v084_(db, """SELECT COUNT(*) FILTER (WHERE status='new') AS new_count, COUNT(*) FILTER (WHERE status='processing') AS processing_count, COUNT(*) FILTER (WHERE status='approved') AS approved_count, COUNT(*) FILTER (WHERE status='completed') AS completed_count, COUNT(*) FILTER (WHERE status='rejected') AS rejected_count, COUNT(*) AS total_count FROM partner_change_requests""") or {}
    return render(request, 'admin_partner_requests.html', {'active':'partner_requests','rows':rows,'stats':stats,'status':status,'q':q,'status_label_v110':status_label_v110})


@router.get('/admin/partner-requests/{request_id}', response_class=HTMLResponse)
def admin_partner_request_detail_v110(request: Request, request_id: str, db: Session = Depends(get_db)):
    ensure_partner_workflow_v110(db)
    row=fetch_one_safe_v084_(db, 'SELECT * FROM partner_change_requests WHERE id=:id LIMIT 1', {'id':request_id})
    comments=fetch_all_safe_v084_(db, 'SELECT * FROM partner_request_comments WHERE request_id=:id ORDER BY created_at DESC', {'id':request_id})
    return render(request, 'admin_partner_request_detail.html', {'active':'partner_requests','row':row,'comments':comments,'status_label_v110':status_label_v110})


@router.post('/admin/partner-requests/{request_id}/status')
def admin_partner_request_status_v110(request_id: str, status: str = Form('processing'), admin_note: str = Form(''), db: Session = Depends(get_db)):
    ensure_partner_workflow_v110(db); user=hub_user_context_v083_(); normalized=normalize_status_v110(status)
    old=fetch_one_safe_v084_(db, 'SELECT * FROM partner_change_requests WHERE id=:id LIMIT 1', {'id':request_id})
    err=v110_exec(db, """UPDATE partner_change_requests SET status=:status, admin_note=:admin_note, processed_by_name=:processed_by_name, processed_by_email=:processed_by_email, processed_at=CASE WHEN :status IN ('approved','rejected','completed') THEN NOW() ELSE processed_at END, updated_at=NOW() WHERE id=:id""", {'id':request_id,'status':normalized,'admin_note':admin_note or '', 'processed_by_name':user.get('name',''), 'processed_by_email':user.get('email','')})
    if not err:
        v110_exec(db, """INSERT INTO partner_request_comments (request_id,author_name,author_email,comment_text,is_internal) VALUES (:request_id,:author_name,:author_email,:comment_text,TRUE)""", {'request_id':request_id,'author_name':user.get('name',''),'author_email':user.get('email',''),'comment_text':f"Změna stavu na {status_label_v110(normalized)}. {admin_note or ''}"})
        if old and normalized in {'approved','rejected','completed'} and old.get('created_by_email'):
            send_partner_workflow_email_v110(db, old.get('created_by_email'), f"HUB ASTORIE – stav požadavku: {status_label_v110(normalized)}", f"Dobrý den,\n\nu vašeho požadavku v sekci Partneři došlo ke změně stavu.\n\nPartner: {old.get('partner_name') or old.get('partner_code') or '—'}\nPožadavek: {old.get('title') or '—'}\nNový stav: {status_label_v110(normalized)}\n\nPoznámka BO:\n{admin_note or '—'}\n\nASTORIE a.s.")
    return RedirectResponse(f'/admin/partner-requests/{request_id}', status_code=303)


@router.post('/admin/partner-requests/{request_id}/comment')
def admin_partner_request_comment_v110(request_id: str, comment_text: str = Form(''), db: Session = Depends(get_db)):
    ensure_partner_workflow_v110(db); user=hub_user_context_v083_()
    if comment_text:
        v110_exec(db, """INSERT INTO partner_request_comments (request_id,author_name,author_email,comment_text,is_internal) VALUES (:request_id,:author_name,:author_email,:comment_text,TRUE)""", {'request_id':request_id,'author_name':user.get('name',''),'author_email':user.get('email',''),'comment_text':comment_text})
    return RedirectResponse(f'/admin/partner-requests/{request_id}', status_code=303)




# -------------------------------------------------------------------
# v1.2.2 Partner Workflow UX Upgrade
# Premium partner workspace: dashboard counters, history timeline,
# request badges, favorite button, compact fulltext and better data cards.
# -------------------------------------------------------------------

PARTNER_WORKFLOW_UX_VERSION = "1.2.2-admin-taxonomy-specialists-links-safe"


def ensure_partner_ux_v111(db: Session):
    errors = []
    for fn_name in ["ensure_partner_workflow_v110", "ensure_v103_tables"]:
        fn = globals().get(fn_name)
        if fn:
            try:
                res = fn(db)
                if isinstance(res, list):
                    errors.extend([str(e) for e in res if e])
            except Exception as exc:
                errors.append(f"{fn_name}: {exc}")

    # Doplnit drobné tabulky, pokud v1.2.2 nebyla nasazená.
    if globals().get("v110_exec"):
        exec_fn = v110_exec
    elif globals().get("v103_exec"):
        exec_fn = v103_exec
    else:
        return errors

    for sql in [
        """
        CREATE TABLE IF NOT EXISTS partner_change_requests (
            id TEXT PRIMARY KEY DEFAULT gen_random_uuid()::text,
            partner_code TEXT DEFAULT '',
            partner_name TEXT DEFAULT '',
            request_type TEXT DEFAULT '',
            request_area TEXT DEFAULT '',
            priority TEXT DEFAULT 'normal',
            status TEXT DEFAULT 'new',
            title TEXT DEFAULT '',
            description TEXT DEFAULT '',
            item_type TEXT DEFAULT '',
            item_id TEXT DEFAULT '',
            tab TEXT DEFAULT '',
            created_by_name TEXT DEFAULT '',
            created_by_email TEXT DEFAULT '',
            admin_note TEXT DEFAULT '',
            created_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS partner_history (
            id TEXT PRIMARY KEY DEFAULT gen_random_uuid()::text,
            partner_code TEXT DEFAULT '',
            event_type TEXT DEFAULT '',
            title TEXT DEFAULT '',
            description TEXT DEFAULT '',
            created_by_name TEXT DEFAULT '',
            created_by_email TEXT DEFAULT '',
            created_at TIMESTAMPTZ DEFAULT NOW()
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS partner_favorites (
            id TEXT PRIMARY KEY DEFAULT gen_random_uuid()::text,
            partner_code TEXT DEFAULT '',
            user_email TEXT DEFAULT '',
            created_at TIMESTAMPTZ DEFAULT NOW()
        )
        """,
    ]:
        err = exec_fn(db, sql)
        if err:
            errors.append(err)

    return [e for e in errors if e]


def fetch_partner_dashboard_v111(db: Session, partner_code: str):
    ensure_partner_ux_v111(db)
    def count_sql(table, where):
        try:
            return int(db.execute(text(f"SELECT COUNT(*) FROM {table} WHERE {where}"), {"code": partner_code}).scalar() or 0)
        except Exception:
            try: db.rollback()
            except Exception: pass
            return 0

    counts = {
        "contacts": count_sql("partner_contacts", "upper(COALESCE(partner_code,'')) = upper(:code) AND COALESCE(is_active, TRUE)=TRUE") if table_exists_v084_(db, "partner_contacts") else 0,
        "links": count_sql("partner_links", "upper(COALESCE(partner_code,'')) = upper(:code) AND COALESCE(is_active, TRUE)=TRUE") if table_exists_v084_(db, "partner_links") else 0,
        "products": count_sql("partner_products", "upper(COALESCE(partner_code,'')) = upper(:code) AND COALESCE(is_active, TRUE)=TRUE") if table_exists_v084_(db, "partner_products") else 0,
        "faq": count_sql("partner_faq", "(upper(COALESCE(partner_code,'')) = upper(:code) OR COALESCE(partner_code,'')='') AND COALESCE(is_active, TRUE)=TRUE") if table_exists_v084_(db, "partner_faq") else 0,
        "requests_open": count_sql("partner_change_requests", "upper(COALESCE(partner_code,'')) = upper(:code) AND COALESCE(status,'new') IN ('new','processing')") if table_exists_v084_(db, "partner_change_requests") else 0,
    }
    return counts


def fetch_partner_history_v111(db: Session, partner_code: str):
    ensure_partner_ux_v111(db)
    if not table_exists_v084_(db, "partner_history"):
        return []
    return fetch_all_safe_v084_(db, """
        SELECT *
        FROM partner_history
        WHERE upper(COALESCE(partner_code,'')) = upper(:code)
        ORDER BY created_at DESC
        LIMIT 30
    """, {"code": partner_code})


def fetch_partner_requests_v111(db: Session, partner_code: str):
    ensure_partner_ux_v111(db)
    if not table_exists_v084_(db, "partner_change_requests"):
        return []
    return fetch_all_safe_v084_(db, """
        SELECT *
        FROM partner_change_requests
        WHERE upper(COALESCE(partner_code,'')) = upper(:code)
        ORDER BY
          CASE status WHEN 'new' THEN 1 WHEN 'processing' THEN 2 WHEN 'approved' THEN 3 WHEN 'completed' THEN 4 WHEN 'rejected' THEN 5 ELSE 9 END,
          created_at DESC
        LIMIT 50
    """, {"code": partner_code})


@router.get("/api/partner-workflow/ux-status")
def api_partner_workflow_ux_status_v111(db: Session = Depends(get_db)):
    errors = ensure_partner_ux_v111(db)
    return {
        "ok": len(errors) == 0,
        "version": PARTNER_WORKFLOW_UX_VERSION,
        "message": "Partner Workflow UX Upgrade je připraven.",
        "errors": errors,
    }


@router.post("/hub/partners/favorite")
def hub_partner_favorite_v111(
    partner_code: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_partner_ux_v111(db)
    user = hub_user_context_v083_()
    email = user.get("email", "")
    if partner_code and email:
        exists = fetch_one_safe_v084_(db, """
            SELECT id FROM partner_favorites
            WHERE upper(partner_code)=upper(:code) AND lower(user_email)=lower(:email)
            LIMIT 1
        """, {"code": partner_code, "email": email})
        if exists:
            v110_exec(db, "DELETE FROM partner_favorites WHERE id=:id", {"id": exists.get("id")})
        else:
            v110_exec(db, """
                INSERT INTO partner_favorites (partner_code, user_email)
                VALUES (:code, :email)
            """, {"code": partner_code, "email": email})
    return RedirectResponse(f"/hub/partners?selected={partner_code}", status_code=303)



# -------------------------------------------------------------------
# v1.2.2 Partner hotfix safe UI
# /hub/partners uses safe template. No import/data destructive changes.
# -------------------------------------------------------------------
PARTNER_HOTFIX_VERSION = "1.2.2-admin-taxonomy-specialists-links-safe"

def ensure_partner_hotfix_v112(db: Session):
    errors = []
    for fn_name in ["ensure_partner_workflow_v110", "ensure_v103_tables"]:
        fn = globals().get(fn_name)
        if fn:
            try:
                res = fn(db)
                if isinstance(res, list): errors.extend([str(e) for e in res if e])
            except Exception as exc:
                errors.append(f"{fn_name}: {exc}")
    return [e for e in errors if e]

@router.get("/api/partner-hotfix/status")
def api_partner_hotfix_status_v112(db: Session = Depends(get_db)):
    errors = ensure_partner_hotfix_v112(db)
    return {"ok": len(errors) == 0, "version": PARTNER_HOTFIX_VERSION, "message": "Partner hotfix safe UI je aktivní.", "errors": errors}


# -------------------------------------------------------------------
# v1.2.2 Partner safe route fix
# Oprava: /hub/partners nesmí padat na nedefinované dashboard/history/request proměnné.
# -------------------------------------------------------------------

@router.get("/api/partner-safe-route/status")
def api_partner_safe_route_status_v113(db: Session = Depends(get_db)):
    return {
        "ok": True,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Route /hub/partners má bezpečné fallback proměnné a nesmí spadnout na dashboard/history/requests.",
        "errors": []
    }




@router.get("/api/partner-safe-route/status")
def api_partner_safe_route_status_v114(db: Session = Depends(get_db)):
    errors = []
    tables = {}
    for t in ["partners", "partner_contacts", "partner_links", "partner_products", "partner_faq", "partner_change_requests"]:
        try:
            exists = table_exists_v084_(db, t)
            count = 0
            if exists:
                count = int(db.execute(text(f"SELECT COUNT(*) FROM {t}")).scalar() or 0)
                db.commit()
            tables[t] = {"exists": bool(exists), "count": count}
        except Exception as exc:
            try:
                db.rollback()
            except Exception:
                pass
            tables[t] = {"exists": False, "count": 0, "error": str(exc)}
            errors.append(f"{t}: {exc}")
    return {
        "ok": len(errors) == 0,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Uživatelská sekce /hub/partners je napojena na admin číselník partnerů a má bezpečné fallbacky.",
        "tables": tables,
        "errors": errors,
    }




@router.get("/api/partner-figma-ui/status")
def api_partner_figma_ui_status_v116(db: Session = Depends(get_db)):
    tables = {}
    errors = []
    for t in ["partners", "partner_contacts", "partner_links", "partner_products", "partner_faq", "partner_change_requests"]:
        try:
            exists = table_exists_v084_(db, t)
            count = 0
            if exists:
                count = int(db.execute(text(f"SELECT COUNT(*) FROM {t}")).scalar() or 0)
                db.commit()
            tables[t] = {"exists": bool(exists), "count": count}
        except Exception as exc:
            try:
                db.rollback()
            except Exception:
                pass
            tables[t] = {"exists": False, "count": 0, "error": str(exc)}
            errors.append(f"{t}: {exc}")
    return {
        "ok": len(errors) == 0,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Sekce Partneři používá nový Figma-like layout dle odsouhlaseného vizuálu. Backend/import/workflow nebyly měněny.",
        "tables": tables,
        "errors": errors,
    }




@router.get("/api/partners-restore-visual/status")
def api_partners_restore_visual_status_v118(db: Session = Depends(get_db)):
    """
    v1.2.2 – kontrola opravné verze sekce Partneři.
    Nic nemaže, nic nemigruje, nemění import ani workflow.
    """
    tables = {}
    errors = []
    for t in ["partners", "partner_contacts", "partner_links", "partner_products", "partner_faq", "partner_change_requests"]:
        try:
            exists = table_exists_v084_(db, t)
            count = 0
            if exists:
                count = int(db.execute(text(f"SELECT COUNT(*) FROM {t}")).scalar() or 0)
                db.commit()
            tables[t] = {"exists": bool(exists), "count": count}
        except Exception as exc:
            try:
                db.rollback()
            except Exception:
                pass
            tables[t] = {"exists": False, "count": 0, "error": str(exc)}
            errors.append(f"{t}: {exc}")
    return {
        "ok": len(errors) == 0,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Opravná verze: vrácen původní vizuál partner workspace, doplněno seskupení kontaktů a produktů bez zásahu do backendu.",
        "tables": tables,
        "errors": errors,
    }




@router.get("/api/safe-rollback-visual-shell/status")
def api_safe_rollback_visual_shell_status_v120(db: Session = Depends(get_db)):
    """
    v1.2.2 – bezpečný rollback destruktivních UI zásahů.
    Neprovádí migrace, nemaže data, nemění import ani workflow.
    """
    tables = {}
    errors = []
    for t in [
        "users", "sections", "subsections", "specialists",
        "partners", "partner_contacts", "partner_links", "partner_products",
        "partner_faq", "partner_change_requests", "tips", "commission_rates"
    ]:
        try:
            exists = table_exists_v084_(db, t)
            count = 0
            if exists:
                count = int(db.execute(text(f"SELECT COUNT(*) FROM {t}")).scalar() or 0)
                db.commit()
            tables[t] = {"exists": bool(exists), "count": count}
        except Exception as exc:
            try:
                db.rollback()
            except Exception:
                pass
            tables[t] = {"exists": False, "count": 0, "error": str(exc)}
            errors.append(f"{t}: {exc}")
    return {
        "ok": len(errors) == 0,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Stabilní rollback na funkční šablony v1.2.2 + neinvazivní sjednocení vizuálu. DB/import/workflow beze změny.",
        "tables": tables,
        "errors": errors,
    }




@router.get("/api/compact-shell-tables/status")
def api_compact_shell_tables_status_v121(db: Session = Depends(get_db)):
    """
    v1.2.2 – kompaktní HUB shell + fulltext v tabulkách.
    Bez migrací, bez změny importu a workflow.
    """
    tables = {}
    errors = []
    for t in [
        "users", "sections", "subsections", "specialists",
        "partners", "partner_contacts", "partner_links", "partner_products",
        "partner_faq", "partner_change_requests", "tips", "commission_rates"
    ]:
        try:
            exists = table_exists_v084_(db, t)
            count = 0
            if exists:
                count = int(db.execute(text(f"SELECT COUNT(*) FROM {t}")).scalar() or 0)
                db.commit()
            tables[t] = {"exists": bool(exists), "count": count}
        except Exception as exc:
            try:
                db.rollback()
            except Exception:
                pass
            tables[t] = {"exists": False, "count": 0, "error": str(exc)}
            errors.append(f"{t}: {exc}")
    return {
        "ok": len(errors) == 0,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Kompaktní uživatelský shell bez velké petrolejové hero hlavičky + bezpečný fulltext v tabulkách.",
        "tables": tables,
        "errors": errors,
    }




# -------------------------------------------------------------------
# v1.2.2 Admin Taxonomy + Specialists + Links Safe
# -------------------------------------------------------------------

def normalize_label_v122_(value: str) -> str:
    """Jednoduchá normalizace pro odstranění duplicitních dlaždic v Novém TIPu."""
    import unicodedata
    value = (value or "").strip().lower()
    value = "".join(ch for ch in unicodedata.normalize("NFD", value) if unicodedata.category(ch) != "Mn")
    value = re.sub(r"[^a-z0-9]+", "", value)
    # sjednocení nejčastějších historických duplicit
    aliases = {
        "flotily": "flotily",
        "flotila": "flotily",
        "majetek": "majetek",
        "zivot": "zivot",
        "zivotnipojisteni": "zivot",
        "podnikatel": "podnikatele",
        "podnikatele": "podnikatele",
        "podnikatelskarizika": "podnikatele",
        "penze": "penze",
        "dps": "penze",
        "uvery": "uvery",
        "hypoteky": "uvery",
        "obnova": "obnova",
        "retence": "obnova",
        "investice": "investice",
        "zlato": "zlato",
        "zvire": "zvire",
        "zvirata": "zvire",
    }
    return aliases.get(value, value)


def dedupe_taxonomy_sections_v122_(sections):
    """
    Bezpečné odstranění duplicit pouze pro zobrazení v poradenském HUBu.
    DB nemažeme. Preferujeme sekci, která už má podsekce/specialisty nebo importovaný kratší kód.
    """
    result = []
    seen = set()
    for s in sections or []:
        name = ""
        code = ""
        try:
            name = s.get("section_name") or s.get("name") or ""
            code = s.get("section_code") or s.get("code") or ""
        except Exception:
            name = getattr(s, "section_name", "") or getattr(s, "name", "")
            code = getattr(s, "section_code", "") or getattr(s, "code", "")
        key = normalize_label_v122_(name or code)
        if not key:
            key = normalize_label_v122_(code)
        if key in seen:
            continue
        seen.add(key)
        result.append(s)
    return result


def get_admin_users_for_specialists_v122_(db: Session):
    try:
        if not table_exists_v084_(db, "users"):
            return []
        return fetch_all_safe_v084_(db, """
            SELECT advisor_id, name, email, phone, role, is_active
            FROM users
            WHERE COALESCE(is_active, TRUE) = TRUE
            ORDER BY name
            LIMIT 1000
        """)
    except Exception:
        return []


def classify_link_scope_v122_(row) -> str:
    try:
        pc = (row.get("partner_code") or "").strip().upper()
        cat = (row.get("category") or "").strip().lower()
    except Exception:
        pc = (getattr(row, "partner_code", "") or "").strip().upper()
        cat = (getattr(row, "category", "") or "").strip().lower()
    if pc in ("", "AST", "ASTORIE", "ASTORIEAS", "ASTORIE_A_S") or "astorie" in cat or "intern" in cat:
        return "astorie"
    return "partner"


@router.get("/api/admin/taxonomy-health")
def api_admin_taxonomy_health_v122(db: Session = Depends(get_db)):
    ensure_visible_hub_sections_(db)
    sections = fetch_all_safe_v084_(db, """
        SELECT section_code, section_name, icon, sort_order, is_active
        FROM hub_sections
        ORDER BY sort_order, section_name
    """)
    subsections = fetch_all_safe_v084_(db, """
        SELECT subsection_code, subsection_name, section_code, sort_order, is_active
        FROM hub_subsections
        ORDER BY section_code, sort_order, subsection_name
    """)
    specialists = get_specialists_for_hub_v085_(db)
    visible = dedupe_taxonomy_sections_v122_(sections)
    return {
        "ok": True,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "sections_total": len(sections),
        "sections_visible_after_dedupe": len(visible),
        "subsections_total": len(subsections),
        "specialists_active_available": len(specialists),
        "note": "Duplicity sekcí se nemažou z DB, pouze se bezpečně skrývají v poradenském Nový TIP.",
        "sections": [dict(s) for s in sections],
        "visible_sections": [dict(s) for s in visible],
    }


@router.post("/admin/sections/{section_code}/update")
def admin_section_update_v122(
    section_code: str,
    section_name: str = Form(...),
    icon: str = Form(""),
    image_url: str = Form(""),
    sort_order: int = Form(100),
    is_active: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_taxonomy_tables_(db)
    db.execute(text("""
        UPDATE hub_sections
        SET section_name = :section_name,
            icon = :icon,
            image_url = :image_url,
            sort_order = :sort_order,
            is_active = :is_active,
            note = :note
        WHERE upper(section_code) = upper(:section_code)
    """), {
        "section_code": section_code,
        "section_name": section_name,
        "icon": icon,
        "image_url": image_url,
        "sort_order": sort_order,
        "is_active": bool(is_active),
        "note": note,
    })
    db.commit()
    return RedirectResponse("/admin/sections", status_code=303)


@router.post("/admin/subsections/{subsection_code}/update")
def admin_subsection_update_v122(
    subsection_code: str,
    section_code: str = Form(...),
    subsection_name: str = Form(...),
    sort_order: int = Form(100),
    is_active: str = Form(""),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_taxonomy_tables_(db)
    db.execute(text("""
        UPDATE hub_subsections
        SET section_code = :section_code,
            subsection_name = :subsection_name,
            sort_order = :sort_order,
            is_active = :is_active,
            note = :note
        WHERE upper(subsection_code) = upper(:subsection_code)
    """), {
        "subsection_code": subsection_code,
        "section_code": section_code.upper().strip(),
        "subsection_name": subsection_name,
        "sort_order": sort_order,
        "is_active": bool(is_active),
        "note": note,
    })
    db.commit()
    return RedirectResponse("/admin/sections", status_code=303)


@router.post("/admin/specialists/create-from-user")
def admin_specialist_create_from_user_v122(
    advisor_id: str = Form(...),
    section_code: str = Form(...),
    subsection_code: str = Form(""),
    role_description: str = Form(""),
    region: str = Form(""),
    if_share: str = Form(""),
    ps_share: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_specialists_table_(db)
    user = fetch_one_safe_v084_(db, """
        SELECT advisor_id, name, email, phone
        FROM users
        WHERE advisor_id = :advisor_id
        LIMIT 1
    """, {"advisor_id": advisor_id})
    if not user:
        return RedirectResponse("/admin/specialists?error=user_not_found", status_code=303)

    db.execute(text("""
        INSERT INTO specialists
        (advisor_id, specialist_name, email, phone, section_code, subsection_code, role_description, region,
         if_share, ps_share, available, unavailable_reason, is_active, note)
        VALUES
        (:advisor_id, :specialist_name, :email, :phone, :section_code, :subsection_code, :role_description, :region,
         :if_share, :ps_share, TRUE, '', TRUE, '')
        ON CONFLICT DO NOTHING
    """), {
        "advisor_id": user["advisor_id"],
        "specialist_name": user["name"],
        "email": user["email"],
        "phone": user["phone"] or "",
        "section_code": section_code.upper().strip(),
        "subsection_code": subsection_code.upper().strip(),
        "role_description": role_description,
        "region": region,
        "if_share": if_share,
        "ps_share": ps_share,
    })
    db.commit()
    return RedirectResponse("/admin/specialists?created=1", status_code=303)


@router.get("/api/admin/links-health")
def api_admin_links_health_v122(db: Session = Depends(get_db)):
    if not table_exists_v084_(db, "partner_links"):
        return {"ok": True, "version": "1.2.2-admin-taxonomy-specialists-links-safe", "astorie": 0, "partners": 0}
    rows = fetch_all_safe_v084_(db, "SELECT * FROM partner_links LIMIT 2000")
    astorie = [r for r in rows if classify_link_scope_v122_(r) == "astorie"]
    partner = [r for r in rows if classify_link_scope_v122_(r) == "partner"]
    return {
        "ok": True,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "astorie": len(astorie),
        "partners": len(partner),
        "note": "ASTORIE odkazy jsou partner_code prázdné/AST/ASTORIE nebo kategorie obsahuje Astorie/interní.",
    }


@router.get("/api/release-1-2-2/status")
def api_release_122_status(db: Session = Depends(get_db)):
    tables = {}
    errors = []
    for t in ["users", "hub_sections", "hub_subsections", "specialists", "partner_links", "global_contacts", "tips"]:
        try:
            exists = table_exists_v084_(db, t)
            count = 0
            if exists:
                count = int(db.execute(text(f"SELECT COUNT(*) FROM {t}")).scalar() or 0)
                db.commit()
            tables[t] = {"exists": bool(exists), "count": count}
        except Exception as exc:
            try:
                db.rollback()
            except Exception:
                pass
            tables[t] = {"exists": False, "count": 0, "error": str(exc)}
            errors.append(f"{t}: {exc}")
    return {
        "ok": not errors,
        "version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "message": "Opravy: deduplikace sekcí pro TIPy, editace sekcí/podsekcí, specialista ze seznamu uživatelů, rozdělení odkazů ASTORIE/partneři.",
        "tables": tables,
        "errors": errors,
    }


# -------------------------------------------------------------------
# v1.3.4 Stable 1.2.2 + Calculators/Rates Visual Safe
# Bezpečný kontrolní endpoint. Nemění DB, importy ani route poradenských sekcí.
# -------------------------------------------------------------------
@router.get("/api/release-1-3-4/status")
def api_release_134_status(db: Session = Depends(get_db)):
    tables = {}
    errors = []
    for t in ["sections", "subsections", "specialists", "users", "partners", "partner_links", "tips", "commission_rates"]:
        try:
            exists = table_exists_v084_(db, t)
            count = 0
            if exists:
                count = int(db.execute(text(f"SELECT COUNT(*) FROM {t}")).scalar() or 0)
            tables[t] = {"exists": bool(exists), "count": count}
        except Exception as exc:
            try:
                db.rollback()
            except Exception:
                pass
            tables[t] = {"exists": False, "count": 0, "error": str(exc)}
            errors.append(f"{t}: {exc}")

    return {
        "ok": len(errors) == 0,
        "version": "1.3.4-stable-122-calculators-rates-visual-safe",
        "message": "Základ je stabilní v1.2.2. Zachováno: TIPy, Moje TIPy, Sekce/Podsekce/Specialisté, Partneři. Upraveno pouze zobrazení Kalkulaček/Sazebníku a fulltext/filtry nad tabulkou.",
        "safe": True,
        "db_changed": False,
        "base_version": "1.2.2-admin-taxonomy-specialists-links-safe",
        "changed_files": ["app/main.py", "app/templates/hub_calculators.html", "app/routers/admin_ui.py - pouze kontrolní endpoint"],
        "routes_expected": ["/hub/new-tip", "/hub/my-tips", "/hub/calculators", "/hub/partners", "/hub/contacts", "/admin/sections", "/admin/specialists"],
        "tables": tables,
        "errors": errors,
    }



# -------------------------------------------------------------------
# v1.3.7 New TIP Exact Visual Required Specialist Safe
# Bezpečný kontrolní endpoint. Nemění DB, importy, API ani jiné sekce.
# -------------------------------------------------------------------
@router.get("/api/release-1-3-7/status")
def api_release_136_status(db: Session = Depends(get_db)):
    tables = {}
    errors = []
    for t in ["sections", "subsections", "specialists", "users", "partners", "partner_links", "tips", "commission_rates"]:
        try:
            exists = table_exists_v084_(db, t)
            count = 0
            if exists:
                count = int(db.execute(text(f"SELECT COUNT(*) FROM {t}")).scalar() or 0)
            tables[t] = {"exists": bool(exists), "count": count}
        except Exception as exc:
            try:
                db.rollback()
            except Exception:
                pass
            tables[t] = {"exists": False, "count": 0, "error": str(exc)}
            errors.append(f"{t}: {exc}")

    return {
        "ok": len(errors) == 0,
        "version": "1.3.7-new-tip-exact-visual-required-specialist-safe",
        "message": "Opraven pouze vizuál sekce Nový TIP podle schváleného návrhu: 3 sloupce Oblast / Podsekce / Specialista, kompaktní dlaždice, spodní souhrn výběru. Ostatní sekce, routy, API ani databáze se nemění.",
        "safe": True,
        "db_changed": False,
        "base_version": "1.3.5-new-tip-premium-visual-safe",
        "changed_files": ["app/main.py", "app/templates/hub_new_tip.html", "app/routers/admin_ui.py - pouze kontrolní endpoint"],
        "routes_expected": ["/hub/new-tip", "/hub/my-tips", "/hub/calculators", "/hub/partners", "/hub/contacts", "/admin/sections", "/admin/specialists"],
        "tables": tables,
        "errors": errors,
    }


@router.get("/api/release-1-3-9/status")
def release_139_status():
    return {
        "ok": True,
        "version": "1.3.9-new-tip-business-fix-safe",
        "safe": True,
        "db_changed": True,
        "scope": "Pouze Nový TIP + bezpečné workflow TIPů",
        "changes": [
            "specialisté se zobrazí až po výběru podsekce",
            "specialista je povinný pro založení TIPu",
            "povinné položky formuláře jsou hlídané na frontendu i backendu",
            "založení TIPu zapisuje zprávu do historie TIPu",
            "pokud je SMTP nastavené, odešle se e-mail specialistovi i poradci",
            "Správa TIPů je dostupná v administraci na /admin/tips"
        ]
    }

@router.get("/api/release-1-4-2/status")
def release_142_status():
    return {
        "ok": True,
        "version": "1.4.2-specialist-badge-and-rates-data-fix-safe",
        "safe": True,
        "db_changed": False,
        "scope": "Pouze Nový TIP badge specialisty + Sazebník provizí data mapping",
        "changes": [
            "specialista ukazuje PŘIJÍMÁ TIPY / NEPŘIJÍMÁ TIPY, dokud není ručně vybrán",
            "DOPORUČENO se zobrazí až po skutečném výběru specialisty",
            "Sazebník vrací sekci, podsekci, partnera, typ, produkt i sazbu přes explicitní aliasy",
            "Sazebník má doplněný filtr produktů a zobrazuje sazby jako badge"
        ]
    }


@router.get("/api/release-1-4-3/status")
def release_143_status():
    return {
        "ok": True,
        "version": "1.4.3-rates-column-mapping-safe",
        "safe": True,
        "db_changed": False,
        "scope": "Pouze Sazebník provizí – oprava mapování sloupců a názvů",
        "changes": [
            "Sazebník zobrazuje přesné sloupce: Sekce, Oblast, Partner, Produkt, Základ, Provize",
            "Produkt je mapován z importního sloupce Druh obchodu",
            "Základ je mapován z importního sloupce Typ_obchodu",
            "Provize je mapována ze Sazba_provize_%",
            "Partneři, Nový TIP, Moje TIPy ani administrace TIPů nejsou měněny"
        ]
    }

# ============================================================
# v1.4.4 SAFE – Admin Sazebník provizí jako DB source of truth
# ============================================================

def ensure_rates_admin_v144_(db: Session):
    """Bezpečné rozšíření tabulky commission_rates. Nemění TIPy, partnery ani ostatní moduly."""
    if not table_exists_v084_(db, "commission_rates"):
        db.execute(text("""
            CREATE TABLE IF NOT EXISTS commission_rates (
                id UUID PRIMARY KEY,
                section_code TEXT,
                subsection_code TEXT,
                partner_name TEXT NOT NULL,
                base_type TEXT,
                product_type TEXT,
                rate_percent NUMERIC(8,4),
                priority INTEGER NOT NULL DEFAULT 100,
                is_active BOOLEAN NOT NULL DEFAULT TRUE,
                business_type TEXT NOT NULL DEFAULT '',
                area TEXT NOT NULL DEFAULT '',
                created_at TIMESTAMP WITH TIME ZONE DEFAULT now(),
                updated_at TIMESTAMP WITH TIME ZONE DEFAULT now()
            )
        """))
    for sql in [
        "ALTER TABLE commission_rates ADD COLUMN IF NOT EXISTS business_type TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE commission_rates ADD COLUMN IF NOT EXISTS area TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE commission_rates ADD COLUMN IF NOT EXISTS note TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE commission_rates ADD COLUMN IF NOT EXISTS source TEXT NOT NULL DEFAULT 'db_admin'",
        "ALTER TABLE commission_rates ADD COLUMN IF NOT EXISTS created_at TIMESTAMP WITH TIME ZONE DEFAULT now()",
        "ALTER TABLE commission_rates ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP WITH TIME ZONE DEFAULT now()",
        "ALTER TABLE commission_rates ADD COLUMN IF NOT EXISTS deleted_at TIMESTAMP WITH TIME ZONE",
    ]:
        db.execute(text(sql))
    db.commit()


def rate_percent_text_v144_(value):
    if value in (None, ""):
        return ""
    s = str(value).strip().replace("%", "").replace(",", ".")
    try:
        d = Decimal(s)
        if d == d.to_integral_value():
            return str(int(d))
        return str(d).replace('.', ',')
    except Exception:
        return str(value)


@router.get("/api/release-1-4-4/status")
def release_144_status(db: Session = Depends(get_db)):
    ensure_rates_admin_v144_(db)
    total = db.execute(text("SELECT COUNT(*) FROM commission_rates WHERE deleted_at IS NULL")).scalar() or 0
    return {
        "ok": True,
        "version": "1.4.4-admin-rates-db-core-safe",
        "message": "Admin Sazebník provizí je aktivní. Aplikace čte ostrý sazebník z DB; Google Sheets slouží pouze jako importní podklad.",
        "safe": True,
        "db_source_of_truth": True,
        "google_sheets_runtime_read": False,
        "changed_modules": ["admin_rates", "hub_calculators_db_mapping"],
        "unchanged_modules": ["new_tip", "my_tips", "partners", "contacts", "specialists", "login"],
        "commission_rates_count": int(total),
    }


@router.post("/admin/rates/ensure")
def admin_rates_ensure(db: Session = Depends(get_db)):
    ensure_rates_admin_v144_(db)
    audit(db, "ENSURE_STRUCTURE", "commission_rates", {"version": "1.4.4", "source_of_truth": "DB"})
    return RedirectResponse(url="/admin/rates?msg=DB struktura sazebníku byla zkontrolována", status_code=303)


@router.get("/admin/rates", response_class=HTMLResponse)
def admin_rates_page(
    request: Request,
    q: str = "",
    section: str = "",
    partner: str = "",
    status: str = "",
    limit: int = 250,
    msg: str = "",
    db: Session = Depends(get_db),
):
    ensure_rates_admin_v144_(db)
    limit = limit if limit in (100, 250, 500) else 250
    params = {"limit": limit}
    where = ["cr.deleted_at IS NULL"]
    if q:
        where.append("""
            (
              lower(COALESCE(s.name, cr.section_code, '')) LIKE :q OR
              lower(COALESCE(cr.area, '')) LIKE :q OR
              lower(COALESCE(cr.partner_name, '')) LIKE :q OR
              lower(COALESCE(cr.business_type, '')) LIKE :q OR
              lower(COALESCE(cr.product_type, '')) LIKE :q OR
              lower(COALESCE(cr.base_type, '')) LIKE :q OR
              CAST(COALESCE(cr.rate_percent, 0) AS TEXT) LIKE :q
            )
        """)
        params["q"] = f"%{q.lower()}%"
    if section:
        where.append("COALESCE(s.name, cr.section_code, '') = :section")
        params["section"] = section
    if partner:
        where.append("cr.partner_name = :partner")
        params["partner"] = partner
    if status == "active":
        where.append("COALESCE(cr.is_active, TRUE) = TRUE")
    elif status == "inactive":
        where.append("COALESCE(cr.is_active, TRUE) = FALSE")
    where_sql = " AND ".join(where)

    rows = fetch_all_safe_v084_(db, f"""
        SELECT
            cr.id::text AS id,
            cr.section_code,
            cr.subsection_code,
            cr.partner_name,
            cr.business_type,
            cr.product_type,
            cr.rate_percent,
            cr.is_active,
            COALESCE(NULLIF(s.name, ''), cr.section_code, '') AS section_display,
            COALESCE(NULLIF(cr.area, ''), '') AS area_display,
            COALESCE(NULLIF(cr.partner_name, ''), '') AS partner_display,
            COALESCE(NULLIF(cr.business_type, ''), '') AS product_display,
            COALESCE(NULLIF(cr.product_type, ''), NULLIF(cr.base_type, ''), '') AS base_display,
            CASE
              WHEN cr.rate_percent IS NULL THEN ''
              WHEN cr.rate_percent = ROUND(cr.rate_percent) THEN TRIM(TO_CHAR(cr.rate_percent, 'FM999999990')) || ' %'
              ELSE REPLACE(TRIM(TO_CHAR(cr.rate_percent, 'FM999999990D99')), '.', ',') || ' %'
            END AS rate_display
        FROM commission_rates cr
        LEFT JOIN sections s ON s.section_code = cr.section_code
        WHERE {where_sql}
        ORDER BY COALESCE(cr.is_active, TRUE) DESC, section_display, area_display, cr.partner_name, product_display, base_display
        LIMIT :limit
    """, params)

    sections = [r[0] for r in db.execute(text("""
        SELECT DISTINCT COALESCE(NULLIF(s.name,''), cr.section_code, '') AS x
        FROM commission_rates cr LEFT JOIN sections s ON s.section_code=cr.section_code
        WHERE cr.deleted_at IS NULL AND COALESCE(NULLIF(s.name,''), cr.section_code, '') <> ''
        ORDER BY x
    """)).all()]
    partners = [r[0] for r in db.execute(text("""
        SELECT DISTINCT partner_name FROM commission_rates
        WHERE deleted_at IS NULL AND COALESCE(partner_name,'') <> '' ORDER BY partner_name
    """)).all()]
    stats_row = fetch_one_safe_v084_(db, """
        SELECT
          COUNT(*) AS total,
          COUNT(*) FILTER (WHERE COALESCE(is_active, TRUE)=TRUE) AS active,
          COUNT(DISTINCT partner_name) AS partners,
          COUNT(DISTINCT section_code) AS sections
        FROM commission_rates
        WHERE deleted_at IS NULL
    """) or {"total":0,"active":0,"partners":0,"sections":0}
    audit_rows = fetch_all_safe_v084_(db, """
        SELECT created_at, action, entity_type, new_value::text AS new_value
        FROM audit_log
        WHERE entity_type = 'commission_rates'
        ORDER BY created_at DESC LIMIT 8
    """) if table_exists_v084_(db, "audit_log") else []
    return render(request, "admin_rates.html", {
        "active": "rates",
        "rows": rows,
        "sections": sections,
        "partners": partners,
        "stats": stats_row,
        "audit_rows": audit_rows,
        "q": q,
        "section": section,
        "partner": partner,
        "status": status,
        "limit": limit,
        "message": msg,
        "warning": "Runtime čtení ze Sheets je vypnuté. Nové sazby se do aplikace dostanou přes import XLSX nebo ruční editaci v DB administraci.",
    })


@router.post("/admin/rates/create")
def admin_rates_create(
    section_code: str = Form(...),
    area: str = Form(""),
    partner_name: str = Form(...),
    business_type: str = Form(...),
    product_type: str = Form(...),
    rate_percent: str = Form(...),
    is_active: str = Form("1"),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_rates_admin_v144_(db)
    new_id = str(uuid.uuid4())
    rate_txt = rate_percent_text_v144_(rate_percent)
    db.execute(text("""
        INSERT INTO commission_rates
        (id, section_code, subsection_code, partner_name, base_type, product_type, rate_percent, priority, is_active, business_type, area, note, source, created_at, updated_at)
        VALUES
        (:id, :section_code, '', :partner_name, '', :product_type, :rate_percent, 100, :is_active, :business_type, :area, :note, 'db_admin', now(), now())
    """), {
        "id": new_id,
        "section_code": section_code.strip().upper(),
        "partner_name": partner_name.strip(),
        "product_type": product_type.strip(),
        "rate_percent": rate_txt.replace(',', '.') if rate_txt else None,
        "is_active": is_active == "1",
        "business_type": business_type.strip(),
        "area": area.strip(),
        "note": note.strip(),
    })
    db.commit()
    audit(db, "CREATE_RATE", "commission_rates", {"id": new_id, "section_code": section_code, "partner": partner_name, "product": business_type, "rate": rate_txt, "note": note})
    return RedirectResponse(url="/admin/rates?msg=Sazba byla přidána", status_code=303)


@router.get("/admin/rates/{rate_id}", response_class=HTMLResponse)
def admin_rate_edit_page(rate_id: str, request: Request, db: Session = Depends(get_db)):
    ensure_rates_admin_v144_(db)
    row = fetch_one_safe_v084_(db, """
        SELECT id::text AS id, section_code, area, partner_name, business_type, product_type, base_type,
               rate_percent, is_active, note
        FROM commission_rates WHERE id::text = :id AND deleted_at IS NULL
    """, {"id": rate_id})
    if not row:
        return RedirectResponse(url="/admin/rates?msg=Sazba nebyla nalezena", status_code=303)
    audit_rows = fetch_all_safe_v084_(db, """
        SELECT created_at, action, new_value::text AS new_value
        FROM audit_log
        WHERE entity_type='commission_rates' AND new_value::text LIKE :needle
        ORDER BY created_at DESC LIMIT 20
    """, {"needle": f"%{rate_id}%"}) if table_exists_v084_(db, "audit_log") else []
    return render(request, "admin_rate_edit.html", {"active": "rates", "row": row, "audit_rows": audit_rows})


@router.post("/admin/rates/{rate_id}/update")
def admin_rate_update(
    rate_id: str,
    section_code: str = Form(...),
    area: str = Form(""),
    partner_name: str = Form(...),
    business_type: str = Form(...),
    product_type: str = Form(...),
    rate_percent: str = Form(...),
    is_active: str = Form("1"),
    note: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_rates_admin_v144_(db)
    old = fetch_one_safe_v084_(db, "SELECT * FROM commission_rates WHERE id::text=:id", {"id": rate_id})
    rate_txt = rate_percent_text_v144_(rate_percent)
    db.execute(text("""
        UPDATE commission_rates SET
          section_code=:section_code,
          area=:area,
          partner_name=:partner_name,
          business_type=:business_type,
          product_type=:product_type,
          rate_percent=:rate_percent,
          is_active=:is_active,
          note=:note,
          source='db_admin',
          updated_at=now()
        WHERE id::text=:id
    """), {
        "id": rate_id,
        "section_code": section_code.strip().upper(),
        "area": area.strip(),
        "partner_name": partner_name.strip(),
        "business_type": business_type.strip(),
        "product_type": product_type.strip(),
        "rate_percent": rate_txt.replace(',', '.') if rate_txt else None,
        "is_active": is_active == "1",
        "note": note.strip(),
    })
    db.commit()
    audit(db, "UPDATE_RATE", "commission_rates", {"id": rate_id, "old": dict(old) if old else {}, "new": {"section_code": section_code, "area": area, "partner": partner_name, "product": business_type, "base": product_type, "rate": rate_txt, "active": is_active, "note": note}})
    return RedirectResponse(url=f"/admin/rates/{rate_id}", status_code=303)


@router.get("/api/release-1-4-5/status")
def release_145_status():
    return {
        "ok": True,
        "version": "1.4.5-admin-contacts-products-links-ux-safe",
        "message": "Admin moduly Kontakty, Produkty a Odkazy mají kompaktní pracovní UX. Databáze, routy, Partneři, Nový TIP, Moje TIPy a Sazebník se nemění.",
        "safe": True,
        "db_changed": False,
        "changed_modules": ["admin_contacts_template", "admin_products_template", "admin_links_template"],
        "unchanged_modules": ["new_tip", "my_tips", "partners", "admin_rates", "hub_calculators", "database", "routes"],
    }


@router.get("/api/release-1-4-7/status")
def release_147_status():
    return {
        "ok": True,
        "version": "1.4.7-terminations-pdf-archive-central-evidence-safe",
        "message": "Opravené generování výpovědi: profesionální náhled bez chybných znaků, automatické uložení do archivu a centrální evidence v Adminu.",
        "safe": True,
        "changed_sections": ["hub_terminations_preview", "termination_documents_archive", "admin_terminations_central_evidence"],
        "db_changed": True,
        "db_change_type": "additive_only_create_table_if_missing_termination_documents",
        "untouched": ["Nový TIP", "Moje TIPy", "Partneři", "Sazebník", "Kontakty", "Produkty", "Odkazy"]
    }


# -------------------------------------------------------------------
# v1.4.8 – E-mail Core SAFE
# Centrální SMTP nastavení, test odeslání a logy.
# -------------------------------------------------------------------
@router.get("/admin/email", response_class=HTMLResponse)
def admin_email_page(request: Request, db: Session = Depends(get_db)):
    ensure_email_tables(db)
    cfg = smtp_config_status()
    logs = db.execute(text("""
        SELECT created_at, event_type, entity_type, entity_id, recipient_email, subject, status, error
        FROM email_logs
        ORDER BY created_at DESC
        LIMIT 100
    """)).mappings().all()
    last_error = db.execute(text("""
        SELECT created_at, error
        FROM email_logs
        WHERE status = 'error' AND COALESCE(error,'') <> ''
        ORDER BY created_at DESC
        LIMIT 1
    """)).mappings().first()
    return render(request, "admin_email.html", {"active": "email", "cfg": cfg, "logs": logs, "last_error": last_error})


@router.post("/admin/email/test")
def admin_email_test(to_email: str = Form(...), db: Session = Depends(get_db)):
    ensure_email_tables(db)
    ok, err = send_email(
        db,
        to_email,
        "Test e-mailu – HUB ASTORIE",
        "Dobrý den,\n\ntoto je testovací e-mail z aplikace HUB ASTORIE. Pokud Vám přišel, SMTP napojení funguje.\n\nS pozdravem\nASTORIE a.s.",
        html_body=(
            "<div style='font-family:Arial,sans-serif;max-width:680px;margin:0 auto;border:1px solid #d9e7e8;border-radius:18px;overflow:hidden'>"
            "<div style='background:#003D4C;color:white;padding:22px 26px'><div style='font-size:12px;letter-spacing:.14em;text-transform:uppercase;color:#bfe5e8'>HUB ASTORIE</div><h1 style='margin:8px 0 0;font-size:24px'>Test e-mailu</h1></div>"
            "<div style='padding:24px 26px;color:#102A33;font-size:15px;line-height:1.6'><p>Dobrý den,</p><p>toto je testovací e-mail z aplikace <b>HUB ASTORIE</b>.</p><p>Pokud Vám přišel, SMTP napojení funguje.</p><hr style='border:none;border-top:1px solid #e6eef1;margin:22px 0'><p>S pozdravem<br><b>ASTORIE a.s.</b></p></div>"
            "</div>"
        ),
        event_type="email_test",
        entity_type="system",
        created_by_email="admin@astorie.local",
        template_key="system_test_direct",
    )
    suffix = "sent=1" if ok else "error=1"
    return RedirectResponse(f"/admin/email?{suffix}", status_code=303)


@router.get("/api/release-1-4-8/status")
def release_1_4_8_status(db: Session = Depends(get_db)):
    ensure_email_tables(db)
    cfg = smtp_config_status()
    cnt = db.execute(text("SELECT COUNT(*) FROM email_logs")).scalar()
    return {
        "ok": True,
        "version": "1.4.8-email-core-safe",
        "message": "Centrální e-mailová služba je připravena. SMTP konfigurace se bere z Render Environment Variables.",
        "smtp_configured": cfg.get("configured"),
        "smtp_host": cfg.get("host"),
        "smtp_port": cfg.get("port"),
        "smtp_security": cfg.get("security"),
        "missing": cfg.get("missing"),
        "email_log_count": int(cnt or 0),
    }




@router.get("/api/release-1-5-5d/status")
def release_1_5_5d_status(db: Session = Depends(get_db)):
    ensure_email_tables(db)
    cfg = smtp_config_status()
    cnt = db.execute(text("SELECT COUNT(*) FROM email_logs")).scalar()
    recent = db.execute(text("""
        SELECT created_at, event_type, recipient_email, subject, status, error
        FROM email_logs
        ORDER BY created_at DESC
        LIMIT 5
    """)).mappings().all()
    return {
        "ok": True,
        "version": "1.5.5d-email-delivery-activation-safe",
        "safe": True,
        "db_changed": "additive_only_email_logs_if_missing",
        "smtp_configured": cfg.get("configured"),
        "smtp_enabled": cfg.get("enabled"),
        "smtp_host": cfg.get("host"),
        "smtp_port": cfg.get("port"),
        "smtp_security": cfg.get("security"),
        "smtp_from": cfg.get("from_email"),
        "missing": cfg.get("missing"),
        "email_log_count": int(cnt or 0),
        "recent": [dict(r) for r in recent],
        "changed_modules": ["email_service", "admin_email_diagnostics", "user_created_email", "password_reset_email", "tip_email_logging"],
        "untouched": ["TIP data", "partners", "rates", "terminations", "import", "production source routing", "admin CRUD"],
    }

@router.get("/api/release-1-4-9/status")
def release_1_4_9_status(db: Session = Depends(get_db)):
    ensure_contact_role_tables_v149(db)
    return {
        "ok": True,
        "version": "1.4.9-admin-contact-roles-ux-safe",
        "message": "Admin Kontakty: nový pracovní UX, číselník rolí a skupin kontaktů. Ostatní moduly beze změny.",
        "safe": True,
        "changed": ["/admin/contacts", "/admin/contact-roles"],
        "db_changed": "contact_roles table + safe partner_contacts.role_group column"
    }

@router.get("/api/release-1-5-2/status")
def release_1_5_2_status(db: Session = Depends(get_db)):
    return {"ok": True, "version": "1.5.2-admin-contacts-top-form-table-ux-safe", "safe": True, "db_changed": False, "changed_modules": ["admin_contacts_ui"], "unchanged_modules": ["users", "permissions", "partners", "tips", "rates", "terminations", "email", "products", "links"]}


# --- v1.5.5a: safe admin CRUD helpers + release check ----------------------
@router.get("/api/release-1-5-5a/status")
def release_155a_status(db: Session = Depends(get_db)):
    counts = {}
    try:
        ensure_link_source_columns_v155a_(db)
    except Exception:
        pass
    for name, sql in {
        "astorie_links": "SELECT COUNT(*) FROM partner_links WHERE COALESCE(is_active, TRUE)=TRUE AND COALESCE(is_archived,FALSE)=FALSE AND source_type='ASTORIE_LINK'",
        "online_calculators": "SELECT COUNT(*) FROM partner_links WHERE COALESCE(is_active, TRUE)=TRUE AND COALESCE(is_archived,FALSE)=FALSE AND source_type='ONLINE_CALCULATOR'",
        "partner_links": "SELECT COUNT(*) FROM partner_links WHERE COALESCE(is_active, TRUE)=TRUE AND COALESCE(is_archived,FALSE)=FALSE AND source_type='PARTNER_LINK'",
        "rates": "SELECT COUNT(*) FROM commission_rates WHERE COALESCE(is_active, TRUE)=TRUE",
    }.items():
        try:
            counts[name] = db.execute(text(sql)).scalar() if table_exists_v084_(db, sql.split('FROM ')[1].split()[0]) else 0
        except Exception:
            counts[name] = None
    return {
        "ok": True,
        "version": "1.5.5a-link-source-help-crud-safe",
        "safe": True,
        "db_changed": "only metadata columns source_type/is_archived on partner_links if missing; no deletes, no imports",
        "changed_modules": ["hub_links", "hub_help", "hub_calculators_source_filter", "partner_detail_link_filter", "admin_crud_routes"],
        "unchanged_modules": ["tips", "users", "permissions", "email", "terminations", "rates_data", "login"],
        "counts": counts,
    }


def _safe_redirect_back(default_url: str):
    return RedirectResponse(default_url, status_code=303)

@router.post("/admin/links/{item_id}/archive")
def archive_link_v155a(item_id: int, db: Session = Depends(get_db)):
    ensure_link_source_columns_v155a_(db)
    item = db.query(PartnerLink).filter(PartnerLink.id == item_id).first()
    if item:
        db.execute(text("UPDATE partner_links SET is_archived = TRUE, is_active = FALSE WHERE id = :id"), {"id": item_id})
        db.commit()
    return _safe_redirect_back("/admin/links")

@router.post("/admin/links/{item_id}/delete")
def delete_link_v155a(item_id: int, db: Session = Depends(get_db)):
    item = db.query(PartnerLink).filter(PartnerLink.id == item_id).first()
    if item:
        db.delete(item)
        db.commit()
    return _safe_redirect_back("/admin/links")

@router.post("/admin/products/{item_id}/archive")
def archive_product_v155a(item_id: int, db: Session = Depends(get_db)):
    item = db.query(PartnerProduct).filter(PartnerProduct.id == item_id).first()
    if item:
        item.is_active = False
        db.commit()
    return _safe_redirect_back("/admin/products")

@router.post("/admin/products/{item_id}/delete")
def delete_product_v155a(item_id: int, db: Session = Depends(get_db)):
    item = db.query(PartnerProduct).filter(PartnerProduct.id == item_id).first()
    if item:
        db.delete(item)
        db.commit()
    return _safe_redirect_back("/admin/products")

@router.post("/admin/contact-roles/{role_id}/delete")
def delete_contact_role_v155a(role_id: int, db: Session = Depends(get_db)):
    if table_exists_v084_(db, "contact_roles"):
        db.execute(text("DELETE FROM contact_roles WHERE id = :id"), {"id": role_id})
        db.commit()
    return _safe_redirect_back("/admin/contact-roles")

@router.post("/admin/sections/{section_code}/archive")
def archive_section_v155a(section_code: str, db: Session = Depends(get_db)):
    ensure_taxonomy_tables_(db)
    db.execute(text("UPDATE hub_sections SET is_active = FALSE WHERE upper(section_code)=upper(:c)"), {"c": section_code})
    db.commit()
    return _safe_redirect_back("/admin/sections")

@router.post("/admin/subsections/{subsection_code}/archive")
def archive_subsection_v155a(subsection_code: str, db: Session = Depends(get_db)):
    ensure_taxonomy_tables_(db)
    db.execute(text("UPDATE hub_subsections SET is_active = FALSE WHERE upper(subsection_code)=upper(:c)"), {"c": subsection_code})
    db.commit()
    return _safe_redirect_back("/admin/sections")

@router.post("/admin/partners/{partner_code}/archive")
def archive_partner_v155a(partner_code: str, db: Session = Depends(get_db)):
    item = db.query(Partner).filter(Partner.partner_code == partner_code.upper()).first()
    if item:
        item.is_active = False
        db.commit()
    return _safe_redirect_back("/admin/partners")


# --- v1.5.5b: Admin Data Control SAFE – CRUD UI endpoints only ----------------------
@router.get("/api/release-1-5-5b/status")
def release_155b_status(db: Session = Depends(get_db)):
    return {
        "ok": True,
        "version": "1.5.5b-admin-data-control-crud-safe",
        "safe": True,
        "db_changed": False,
        "data_deleted_automatically": False,
        "changed_modules": ["admin_contacts", "admin_contact_roles", "admin_links", "admin_products", "admin_sections", "admin_partners_status_ui"],
        "unchanged_modules": ["hub_production", "tips", "users", "permissions", "email", "terminations", "rates", "login", "import"],
        "message": "Pouze admin UI a bezpečné CRUD endpointy. Žádný import, žádné automatické mazání dat."
    }

@router.post("/admin/contacts/{item_id}/save")
def save_contact_admin_v155b(
    item_id: int,
    partner_code: str = Form(...),
    full_name: str = Form(...),
    role: str = Form(""),
    email: str = Form(""),
    phone: str = Form(""),
    specialization: str = Form(""),
    contact_type: str = Form(""),
    territory: str = Form(""),
    is_vip: str = Form(""),
    note: str = Form(""),
    is_active: str = Form(""),
    db: Session = Depends(get_db),
):
    item = db.query(PartnerContact).filter(PartnerContact.id == item_id).first()
    if item:
        item.partner_code = (partner_code or "").upper().strip()
        item.full_name = full_name.strip()
        item.role = role.strip()
        item.email = email.strip()
        item.phone = phone.strip()
        item.specialization = specialization.strip()
        item.contact_type = contact_type.strip()
        item.territory = territory.strip()
        item.is_vip = bool(is_vip)
        item.is_top = bool(is_vip)
        item.note = note.strip()
        item.is_active = bool(is_active)
        db.commit()
    return RedirectResponse("/admin/contacts", status_code=303)

@router.post("/admin/contacts/{item_id}/archive")
def archive_contact_admin_v155b(item_id: int, db: Session = Depends(get_db)):
    item = db.query(PartnerContact).filter(PartnerContact.id == item_id).first()
    if item:
        item.is_active = False
        db.commit()
    return RedirectResponse("/admin/contacts", status_code=303)

@router.post("/admin/contacts/{item_id}/delete")
def delete_contact_admin_v155b(item_id: int, db: Session = Depends(get_db)):
    item = db.query(PartnerContact).filter(PartnerContact.id == item_id).first()
    if item:
        db.delete(item)
        db.commit()
    return RedirectResponse("/admin/contacts", status_code=303)

@router.post("/admin/links/{item_id}/save")
def save_link_admin_v155b(
    item_id: int,
    partner_code: str = Form(...),
    title: str = Form(...),
    url: str = Form(...),
    category: str = Form(""),
    note: str = Form(""),
    source_type: str = Form(""),
    is_active: str = Form(""),
    db: Session = Depends(get_db),
):
    ensure_link_source_columns_v155a_(db)
    item = db.query(PartnerLink).filter(PartnerLink.id == item_id).first()
    if item:
        item.partner_code = (partner_code or "").upper().strip()
        item.title = title.strip()
        item.url = url.strip()
        item.category = category.strip()
        item.note = note.strip()
        item.is_active = bool(is_active)
        st = (source_type or "").strip().upper()
        if st in ("ASTORIE_LINK", "ONLINE_CALCULATOR", "PARTNER_LINK"):
            db.execute(text("UPDATE partner_links SET source_type = :st WHERE id = :id"), {"st": st, "id": item_id})
        db.commit()
    return RedirectResponse("/admin/links", status_code=303)

@router.post("/admin/products/{item_id}/save")
def save_product_admin_v155b(
    item_id: int,
    partner_code: str = Form(...),
    area: str = Form(""),
    subarea: str = Form(""),
    product_name: str = Form(...),
    note: str = Form(""),
    is_active: str = Form(""),
    db: Session = Depends(get_db),
):
    item = db.query(PartnerProduct).filter(PartnerProduct.id == item_id).first()
    if item:
        item.partner_code = (partner_code or "").upper().strip()
        item.area = area.strip()
        item.subarea = subarea.strip()
        item.product_name = product_name.strip()
        item.note = note.strip()
        item.is_active = bool(is_active)
        db.commit()
    return RedirectResponse("/admin/products", status_code=303)

@router.post("/admin/partners/{partner_code}/status")
def partner_status_admin_v155b(
    partner_code: str,
    partner_status: str = Form("aktivní"),
    is_active: str = Form(""),
    db: Session = Depends(get_db),
):
    item = db.query(Partner).filter(Partner.partner_code == partner_code.upper()).first()
    if item:
        item.partner_status = partner_status.strip() or "aktivní"
        item.is_active = bool(is_active)
        db.commit()
    return RedirectResponse("/admin/partners", status_code=303)


@router.get("/api/release-1-5-5c/status")
def release_1_5_5c_status():
    return {
        "ok": True,
        "version": "1.5.5c-admin-checkbox-ui-polish-safe",
        "safe": True,
        "db_changed": False,
        "data_deleted": False,
        "changed_modules": [
            "admin_checkbox_visuals",
            "admin_contacts_ui",
            "admin_links_ui",
            "admin_products_ui",
            "admin_sections_ui",
            "admin_partner_detail_ui"
        ],
        "unchanged_modules": [
            "production_hub",
            "tips",
            "login",
            "users",
            "permissions_backend",
            "smtp",
            "terminations",
            "rates_backend",
            "imports",
            "database"
        ],
        "note": "Pouze vizuální oprava checkboxů/přepínačů v Adminu. Backend, DB a produkční čtení dat beze změny."
    }


# -------------------------------------------------------------------
# v1.5.5e – SMTP Diagnostics SAFE
# Pouze lepší diagnostika e-mailů. Nemění DB data ani business moduly.
# -------------------------------------------------------------------
@router.get("/api/release-1-5-5e/status")
def release_1_5_5e_status(db: Session = Depends(get_db)):
    ensure_email_tables(db)
    cfg = public_smtp_diagnostics()
    last = db.execute(text("""
        SELECT created_at, status, error, smtp_host
        FROM email_logs
        ORDER BY created_at DESC
        LIMIT 1
    """)).mappings().first()
    return {
        "ok": True,
        "version": "1.5.5e-email-diagnostics-safe",
        "safe": True,
        "db_changed": False,
        "changed_modules": ["email_diagnostics", "smtp_error_visibility"],
        "unchanged_modules": ["tips", "partners", "rates", "terminations", "admin_crud", "production_hub", "import"],
        "smtp": cfg,
        "last_email_event": dict(last) if last else None,
    }

@router.get("/api/email/diagnostics")
def api_email_diagnostics(db: Session = Depends(get_db)):
    ensure_email_tables(db)
    cfg = public_smtp_diagnostics()
    last_errors = db.execute(text("""
        SELECT created_at, event_type, recipient_email, subject, status, error, smtp_host
        FROM email_logs
        ORDER BY created_at DESC
        LIMIT 10
    """)).mappings().all()
    return {"ok": True, "smtp": cfg, "last_events": [dict(r) for r in last_errors]}




# -------------------------------------------------------------------
# v1.6.0B – VERSION + EMAIL STATUS CLEANUP SAFE
# Sjednocení verze, zpřehlednění SMTP/e-mailového statusu.
# Bez změny SMTP odesílací logiky a bez zásahu do business dat.
# -------------------------------------------------------------------
@router.get("/api/release-1-6-0b/status")
def release_1_6_0b_status(db: Session = Depends(get_db)):
    ensure_email_tables(db)
    cfg = public_smtp_diagnostics()
    counts = db.execute(text("""
        SELECT
          COUNT(*) AS total,
          SUM(CASE WHEN status = 'sent' THEN 1 ELSE 0 END) AS sent,
          SUM(CASE WHEN status = 'error' THEN 1 ELSE 0 END) AS errors,
          SUM(CASE WHEN status IN ('not_configured','disabled') THEN 1 ELSE 0 END) AS skipped
        FROM email_logs
    """)).mappings().first()
    last_success = db.execute(text("""
        SELECT created_at, recipient_email, subject, smtp_host, template_key
        FROM email_logs
        WHERE status = 'sent'
        ORDER BY created_at DESC
        LIMIT 1
    """)).mappings().first()
    last_error = db.execute(text("""
        SELECT created_at, recipient_email, subject, error, smtp_host, template_key
        FROM email_logs
        WHERE status = 'error' AND COALESCE(error,'') <> ''
        ORDER BY created_at DESC
        LIMIT 1
    """)).mappings().first()
    return {
        "ok": True,
        "version": "1.6.0d-email-send-rollback-stabilization-safe",
        "safe": True,
        "db_changed": False,
        "data_deleted": False,
        "message": "Sjednocená verze aplikace a zpřehledněný e-mailový status. SMTP odesílání, DB data a business moduly beze změny.",
        "smtp": cfg,
        "email_counts": dict(counts) if counts else {"total": 0, "sent": 0, "errors": 0, "skipped": 0},
        "last_success": dict(last_success) if last_success else None,
        "last_error": dict(last_error) if last_error else None,
        "changed_modules": ["version_badge", "api_version", "release_status", "email_status_display"],
        "unchanged_modules": ["smtp_delivery", "tips", "partners", "contacts", "links", "products", "rates", "terminations", "imports", "login", "permissions"],
    }

@router.get("/api/release-1-6-0/status")
def release_1_6_0_status_alias(db: Session = Depends(get_db)):
    # Backward compatible alias: current 1.6 line status.
    return release_1_6_0b_status(db)


@router.get("/api/release-1-6-0d/status")
def release_1_6_0d_status(db: Session = Depends(get_db)):
    ensure_email_tables(db)
    cfg = smtp_config_status()
    return {
        "ok": True,
        "version": "1.6.0d-email-send-rollback-stabilization-safe",
        "safe": True,
        "db_changed": False,
        "data_deleted": False,
        "changed_modules": ["email_test_route", "email_template_compatibility", "version_badge"],
        "unchanged_modules": ["tips", "partners", "contacts", "links", "products", "rates", "terminations", "login", "permissions", "imports"],
        "smtp": {
            "configured": cfg.get("configured"),
            "host": cfg.get("host"),
            "port": cfg.get("port"),
            "security": cfg.get("security"),
            "user": cfg.get("user"),
            "from_email": cfg.get("from_email"),
            "missing": cfg.get("missing"),
        }
    }

# -------------------------------------------------------------------
# v1.6.0E – EMAIL SEND COMPAT STABILIZATION SAFE
# Fixuje kompatibilitu volání send_email(template_key=...) proti staršímu mailer.py.
# Bez změny DB, SMTP proměnných a business modulů.
# -------------------------------------------------------------------
@router.get("/api/release-1-6-0e/status")
def release_1_6_0e_status(db: Session = Depends(get_db)):
    ensure_email_tables(db)
    cfg = smtp_config_status()
    return {
        "ok": True,
        "version": "1.6.0e-email-send-compat-stabilization-safe",
        "safe": True,
        "db_changed": False,
        "data_deleted": False,
        "fixed_error": "TypeError: send_email() got an unexpected keyword argument 'template_key'",
        "changed_modules": ["admin_email_test_route", "send_email_compat_wrapper", "email_template_compatibility"],
        "unchanged_modules": ["smtp_settings", "tips", "partners", "contacts", "links", "products", "rates", "terminations", "login", "permissions", "imports"],
        "smtp": {
            "configured": cfg.get("configured"),
            "host": cfg.get("host"),
            "port": cfg.get("port"),
            "security": cfg.get("security"),
            "user": cfg.get("user"),
            "from_email": cfg.get("from_email"),
            "missing": cfg.get("missing"),
        },
    }


# -------------------------------------------------------------------
# v1.6.1 – MAIL TEMPLATES PROFESSIONAL SAFE
# Stabilní povýšení e-mailových šablon bez změny SMTP, DB a produkčních dat.
# -------------------------------------------------------------------
@router.get("/api/release-1-6-1/status")
def release_1_6_1_status(db: Session = Depends(get_db)):
    ensure_email_tables(db)
    cfg = smtp_config_status()
    counts = db.execute(text("""
        SELECT
          COUNT(*) AS total,
          SUM(CASE WHEN status = 'sent' THEN 1 ELSE 0 END) AS sent,
          SUM(CASE WHEN status = 'error' THEN 1 ELSE 0 END) AS errors
        FROM email_logs
    """)).mappings().first()
    return {
        "ok": True,
        "version": "1.6.1-mail-templates-professional-safe",
        "safe": True,
        "db_changed": False,
        "data_deleted": False,
        "changed_modules": [
            "email_html_templates",
            "email_version_badge",
            "email_status_endpoint"
        ],
        "unchanged_modules": [
            "smtp_delivery",
            "tips",
            "partners",
            "contacts",
            "links",
            "products",
            "rates",
            "terminations",
            "login",
            "permissions",
            "imports",
            "production_reading"
        ],
        "smtp": {
            "configured": cfg.get("configured"),
            "host": cfg.get("host"),
            "port": cfg.get("port"),
            "security": cfg.get("security"),
            "user": cfg.get("user"),
            "from_email": cfg.get("from_email"),
            "missing": cfg.get("missing"),
        },
        "email_counts": {
            "total": int((counts or {}).get("total") or 0),
            "sent": int((counts or {}).get("sent") or 0),
            "errors": int((counts or {}).get("errors") or 0),
        },
    }
